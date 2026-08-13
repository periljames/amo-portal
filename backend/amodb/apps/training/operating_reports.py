from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from ..accounts import models as account_models
from . import operating_models as models
from .operating_service import budget_totals


NAVY = colors.HexColor("#13233f")
GOLD = colors.HexColor("#c7a755")
PALE = colors.HexColor("#f3f6fa")


def _header(story: list, *, amo: account_models.AMO, title: str, form_reference: str | None, status: str, revision: int) -> None:
    styles = getSampleStyleSheet()
    story.append(Paragraph(amo.name, styles["Title"]))
    story.append(Paragraph(title, styles["Heading1"]))
    metadata = [["AMO", amo.amo_code], ["Status", status], ["Revision", str(revision)], ["Form reference", form_reference or "Configured reference not set"]]
    table = Table(metadata, colWidths=[38 * mm, 90 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), NAVY), ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([table, Spacer(1, 7 * mm)])


def plan_pdf(db: Session, *, plan: models.TrainingPlan, amo: account_models.AMO) -> bytes:
    output = BytesIO()
    story: list = []
    _header(story, amo=amo, title=f"Annual Training Plan {plan.plan_year}", form_reference=plan.form_reference, status=plan.status, revision=plan.revision_no)
    data = [["Course", "Kind", "Provider", "Participants", "Period", "Cost", "Justification"]]
    for item in plan.items:
        period = str(item.planned_start or f"Q{item.quarter or 1}")
        data.append([
            f"{item.course_code_snapshot or ''}\n{item.course_name_snapshot}", item.training_kind,
            item.provider or item.provider_mode, str(item.participant_count), period,
            f"{item.original_currency} {item.estimated_total_cost}", item.justification or "",
        ])
    table = Table(data, repeatRows=1, colWidths=[47 * mm, 23 * mm, 34 * mm, 20 * mm, 24 * mm, 28 * mm, 68 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PALE]), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 7), ("LEADING", (0, 0), (-1, -1), 9),
    ]))
    story.append(table)
    SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm, title=f"Training Plan {plan.plan_year}").build(story)
    return output.getvalue()


def attendance_pdf(db: Session, *, event, entries: list[models.TrainingAttendanceEntry], window: models.TrainingAttendanceWindow | None, amo: account_models.AMO) -> bytes:
    output = BytesIO()
    story: list = []
    _header(story, amo=amo, title=f"Attendance Register — {event.title}", form_reference=getattr(window, "form_reference", None), status=window.status if window else "UNCERTIFIED", revision=window.register_revision if window else 1)
    users = {str(row.id): row for row in db.query(account_models.User).filter(account_models.User.id.in_([entry.user_id for entry in entries])).all()}
    data = [["Staff code", "Participant", "Status", "Method", "Signed at (UTC)", "Signed by"]]
    for entry in entries:
        user = users.get(str(entry.user_id))
        data.append([getattr(user, "staff_code", ""), getattr(user, "full_name", entry.user_id), entry.status, entry.method, entry.signed_at.strftime("%Y-%m-%d %H:%M"), entry.signed_by_user_id or ""])
    table = Table(data, repeatRows=1, colWidths=[27 * mm, 52 * mm, 25 * mm, 28 * mm, 40 * mm, 45 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PALE]), ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    if window and window.certified_at:
        story.extend([Spacer(1, 7 * mm), Paragraph(f"Certified by user {window.certified_by_user_id} at {window.certified_at.isoformat()}. {window.certification_note or ''}", getSampleStyleSheet()["BodyText"])])
    SimpleDocTemplate(output, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm, title=f"Attendance Register {event.title}").build(story)
    return output.getvalue()


def budget_xlsx(*, budget: models.TrainingBudget, amo: account_models.AMO) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Budget"
    sheet.append([amo.name, f"Training Budget Revision {budget.revision_no}"])
    sheet.append(["Status", budget.status, "Reporting currency", budget.reporting_currency, "Form reference", budget.form_reference or ""])
    sheet.append([])
    headers = ["Course code", "Course", "Kind", "Provider", "Quarter", "Trainees", "Original currency", "Unit cost", "Planned", "Exchange rate", "Rate date", "Rate source", "Converted planned", "Approved", "Committed", "Actual"]
    sheet.append(headers)
    for line in budget.lines:
        sheet.append([
            line.course_code_snapshot, line.course_name_snapshot, line.training_kind, line.provider, line.quarter,
            line.trainee_count, line.original_currency, line.unit_cost, line.planned_amount, line.exchange_rate,
            line.rate_date, line.rate_source, line.converted_planned_amount, line.converted_approved_amount,
            line.converted_committed_amount, line.converted_actual_amount,
        ])
    quarter, annual = budget_totals(budget)
    sheet.append([])
    sheet.append(["Quarter totals", *[quarter[f"Q{index}"] for index in range(1, 5)]])
    sheet.append(["Annual totals", annual["planned"], annual["approved"], annual["committed"], annual["actual"], annual["variance_to_plan"], annual["variance_to_approved"]])
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor="13233F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:P{sheet.max_row - 3}"
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(38, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
