from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any, Optional

from sqlalchemy.orm import Session

from ..accounts import models as accounts_models
from . import models, schemas


@dataclass
class ParsedCourseRow:
    row_number: int
    course_id: str
    course_name: str
    frequency_months: Optional[int]
    status: str
    category_raw: Optional[str]
    is_mandatory: bool
    scope: Optional[str]
    reference: Optional[str]


EXPECTED_HEADERS = [
    "CourseID",
    "CourseName",
    "FrequencyMonths",
    "Status",
    "Category",
    "Mandatory",
    "Scope",
    "Reference",
]

ALLOWED_STATUSES = {
    "initial": "Initial",
    "recurrent": "Recurrent",
    "one_off": "One_Off",
}




def _normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _normalized_scope_text(value: Optional[str]) -> str:
    return _normalize_whitespace((value or "").lower())


def _parse_scope_parts(raw_scope: Optional[str]) -> tuple[Optional[models.TrainingRequirementScope], Optional[str], Optional[str], Optional[str]]:
    raw = _clean(raw_scope)
    if not raw:
        return None, None, None, None

    norm = _normalized_scope_text(raw)
    if norm in {"all", "all staff", "all employees", "all personnel", "everyone", "entire amo", "amo-wide"}:
        return models.TrainingRequirementScope.ALL, None, None, None

    department_match = re.match(r"^(?:department|dept)\s*[:=-]\s*(.+)$", raw, flags=re.IGNORECASE)
    if department_match:
        value = _normalize_whitespace(department_match.group(1)).upper()
        return models.TrainingRequirementScope.DEPARTMENT, value or None, None, None

    role_match = re.match(r"^(?:job role|role|position|title)\s*[:=-]\s*(.+)$", raw, flags=re.IGNORECASE)
    if role_match:
        value = _normalize_whitespace(role_match.group(1))
        return models.TrainingRequirementScope.JOB_ROLE, None, value or None, None

    return None, None, None, None


def _derive_requirement_rule(
    db: Session,
    *,
    amo_id: str,
    is_mandatory: bool,
    raw_scope: Optional[str],
) -> tuple[Optional[models.TrainingRequirementScope], Optional[str], Optional[str], Optional[str]]:
    if not is_mandatory:
        return None, None, None, None

    scope, department_code, job_role, user_id = _parse_scope_parts(raw_scope)
    if scope is not None:
        return scope, department_code, job_role, user_id

    raw = _clean(raw_scope)
    if not raw:
        return None, None, None, None

    norm = _normalized_scope_text(raw)

    department = (
        db.query(accounts_models.Department)
        .filter(accounts_models.Department.amo_id == amo_id, accounts_models.Department.is_active.is_(True))
        .all()
    )
    for item in department:
        if norm == _normalized_scope_text(item.code) or norm == _normalized_scope_text(item.name):
            return models.TrainingRequirementScope.DEPARTMENT, item.code.upper(), None, None

    return None, None, None, None


def _find_existing_requirement(
    db: Session,
    *,
    amo_id: str,
    course_pk: str,
    scope: models.TrainingRequirementScope,
    department_code: Optional[str],
    job_role: Optional[str],
    user_id: Optional[str],
) -> Optional[models.TrainingRequirement]:
    query = db.query(models.TrainingRequirement).filter(
        models.TrainingRequirement.amo_id == amo_id,
        models.TrainingRequirement.course_id == course_pk,
        models.TrainingRequirement.scope == scope,
    )

    query = query.filter(models.TrainingRequirement.department_code == department_code if department_code is not None else models.TrainingRequirement.department_code.is_(None))
    query = query.filter(models.TrainingRequirement.job_role == job_role if job_role is not None else models.TrainingRequirement.job_role.is_(None))
    query = query.filter(models.TrainingRequirement.user_id == user_id if user_id is not None else models.TrainingRequirement.user_id.is_(None))
    return query.first()
def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_course_name(value: Any) -> Optional[str]:
    raw = _clean(value)
    if raw is None:
        return None
    # Intentionally normalize all whitespace (including embedded newlines) to single spaces.
    return re.sub(r"\s+", " ", raw).strip()


def _parse_frequency(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ValueError("FrequencyMonths must be an integer or blank")
    if isinstance(value, (int, float)):
        if int(value) != value:
            raise ValueError("FrequencyMonths must be an integer or blank")
        return int(value)
    raw = str(value).strip()
    if not raw:
        return None
    if raw.isdigit():
        return int(raw)
    raise ValueError("FrequencyMonths must be an integer or blank")


def _parse_mandatory(value: Any) -> bool:
    raw = (_clean(value) or "").lower()
    if raw in {"", "no", "n", "false", "0"}:
        return False
    if raw in {"yes", "y", "true", "1"}:
        return True
    raise ValueError("Mandatory must be Yes/No (or blank)")


def _parse_status(value: Any) -> str:
    raw = _clean(value)
    if not raw:
        raise ValueError("Status is required")
    canonical = ALLOWED_STATUSES.get(raw.lower())
    if canonical is None:
        raise ValueError("Status must be one of: Initial, Recurrent, One_Off")
    return canonical


def parse_courses_sheet(file_bytes: bytes, *, filename: str, sheet_name: str = "Courses") -> list[dict[str, Any]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in {"xlsx", "xlsm", "xltx", "xltm"}:
        raise ValueError("Only Excel .xlsx/.xlsm files are supported for courses import.")
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for courses import. Install openpyxl.") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(
            f"Unexpected headers in '{sheet_name}'. Expected exact order {EXPECTED_HEADERS}, got {headers}"
        )

    parsed: list[dict[str, Any]] = []
    for idx, values in enumerate(rows[1:], start=2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if not any(v is not None and str(v).strip() for v in row.values()):
            continue
        parsed.append({"row_number": idx, **row})
    return parsed


def _build_parsed_course(raw: dict[str, Any]) -> ParsedCourseRow:
    course_id = (_clean(raw.get("CourseID")) or "").upper()
    course_name = _normalize_course_name(raw.get("CourseName")) or ""
    if not course_id:
        raise ValueError("CourseID is required")
    if not course_name:
        raise ValueError("CourseName is required")
    return ParsedCourseRow(
        row_number=int(raw["row_number"]),
        course_id=course_id,
        course_name=course_name,
        frequency_months=_parse_frequency(raw.get("FrequencyMonths")),
        status=_parse_status(raw.get("Status")),
        category_raw=_clean(raw.get("Category")),
        is_mandatory=_parse_mandatory(raw.get("Mandatory")),
        scope=_clean(raw.get("Scope")),
        reference=_clean(raw.get("Reference")),
    )


def import_courses_rows(
    db: Session,
    *,
    amo_id: str,
    rows: list[dict[str, Any]],
    dry_run: bool,
    actor_user_id: Optional[str] = None,
) -> schemas.CourseImportSummary:
    issues: list[schemas.CourseImportRowIssue] = []
    created_courses = 0
    updated_courses = 0
    created_requirements = 0
    updated_requirements = 0
    seen_ids: set[str] = set()

    for raw in rows:
        try:
            parsed = _build_parsed_course(raw)
        except ValueError as exc:
            issues.append(schemas.CourseImportRowIssue(row_number=int(raw.get("row_number") or 0), reason=str(exc)))
            continue

        if parsed.course_id.lower() in seen_ids:
            issues.append(
                schemas.CourseImportRowIssue(
                    row_number=parsed.row_number,
                    course_id=parsed.course_id,
                    reason="Duplicate CourseID inside import file.",
                )
            )
            continue
        seen_ids.add(parsed.course_id.lower())

        existing = (
            db.query(models.TrainingCourse)
            .filter(models.TrainingCourse.amo_id == amo_id, models.TrainingCourse.course_id == parsed.course_id)
            .first()
        )
        if existing is None:
            created_courses += 1
            if dry_run:
                scope, department_code, job_role, user_id = _derive_requirement_rule(
                    db, amo_id=amo_id, is_mandatory=parsed.is_mandatory, raw_scope=parsed.scope
                )
                if scope is not None:
                    created_requirements += 1
                continue

            existing = models.TrainingCourse(
                amo_id=amo_id,
                course_id=parsed.course_id,
                course_name=parsed.course_name,
                frequency_months=parsed.frequency_months,
                status=parsed.status,
                category_raw=parsed.category_raw,
                scope=parsed.scope,
                regulatory_reference=parsed.reference,
                is_mandatory=parsed.is_mandatory,
                mandatory_for_all=_normalized_scope_text(parsed.scope) in {"all", "all staff", "all employees", "all personnel", "everyone", "entire amo", "amo-wide"},
                created_by_user_id=actor_user_id,
                updated_by_user_id=actor_user_id,
            )
            db.add(existing)
            db.flush()
        else:
            updated_courses += 1
            if not dry_run:
                existing.course_name = parsed.course_name
                existing.frequency_months = parsed.frequency_months
                existing.status = parsed.status
                existing.category_raw = parsed.category_raw
                existing.scope = parsed.scope
                existing.regulatory_reference = parsed.reference
                existing.is_mandatory = parsed.is_mandatory
                existing.mandatory_for_all = _normalized_scope_text(parsed.scope) in {"all", "all staff", "all employees", "all personnel", "everyone", "entire amo", "amo-wide"}
                existing.updated_by_user_id = actor_user_id
                db.add(existing)

        scope, department_code, job_role, user_id = _derive_requirement_rule(
            db, amo_id=amo_id, is_mandatory=parsed.is_mandatory, raw_scope=parsed.scope
        )
        if scope is None:
            continue

        requirement = _find_existing_requirement(
            db,
            amo_id=amo_id,
            course_pk=existing.id,
            scope=scope,
            department_code=department_code,
            job_role=job_role,
            user_id=user_id,
        )

        if requirement is None:
            created_requirements += 1
            if not dry_run:
                db.add(
                    models.TrainingRequirement(
                        amo_id=amo_id,
                        course_id=existing.id,
                        scope=scope,
                        department_code=department_code,
                        job_role=job_role,
                        user_id=user_id,
                        is_mandatory=True,
                        is_active=True,
                        created_by_user_id=actor_user_id,
                    )
                )
            continue

        updated_requirements += 1
        if not dry_run:
            requirement.is_mandatory = True
            requirement.is_active = True
            db.add(requirement)

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return schemas.CourseImportSummary(
        dry_run=dry_run,
        total_rows=len(rows),
        created_courses=created_courses,
        updated_courses=updated_courses,
        created_requirements=created_requirements,
        updated_requirements=updated_requirements,
        skipped_rows=len(issues),
        issues=issues,
    )
