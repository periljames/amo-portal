from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import Any, Dict, Iterable, Optional

from sqlalchemy.orm import Session

from ..accounts import models as accounts_models
from . import models, schemas
from .compliance import add_months, is_initial_course

EXPECTED_HEADERS = [
    "RecordID",
    "PersonID",
    "PersonName",
    "CourseID",
    "CourseName",
    "LastTrainingDate",
    "NextDueDate",
    "DaysToDue",
    "Status",
]


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _sheet_name_candidates(sheet_name: str) -> set[str]:
    base = (sheet_name or "Training").strip().lower()
    return {
        base,
        base.rstrip("s"),
        f"{base}s",
        "training",
        "trainings",
        "traininghistory",
        "training history",
    }


def _locate_training_sheet(workbook, requested_sheet_name: str):
    if requested_sheet_name in workbook.sheetnames:
        return workbook[requested_sheet_name]

    candidate_names = _sheet_name_candidates(requested_sheet_name)
    for name in workbook.sheetnames:
        if name.strip().lower() in candidate_names:
            return workbook[name]

    expected = {_normalize_header(header) for header in EXPECTED_HEADERS}
    for name in workbook.sheetnames:
        ws = workbook[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        normalized_headers = {_normalize_header(cell) for cell in rows[0] if cell not in (None, "")}
        if expected.issubset(normalized_headers):
            return ws

    raise ValueError(f"Sheet '{requested_sheet_name}' not found and no compatible training sheet was detected.")


def _resolve_training_headers(header_row: list[Any]) -> tuple[list[str], dict[str, int]]:
    actual_headers = [str(c).strip() if c is not None else "" for c in header_row]
    normalized_to_index = {_normalize_header(value): idx for idx, value in enumerate(actual_headers)}
    missing = [header for header in EXPECTED_HEADERS if _normalize_header(header) not in normalized_to_index]
    if missing:
        raise ValueError(
            f"Unexpected headers in training sheet. Missing required columns {missing}. Found {actual_headers}"
        )
    return actual_headers, {header: normalized_to_index[_normalize_header(header)] for header in EXPECTED_HEADERS}


@dataclass(frozen=True)
class ParsedTrainingImportRow:
    row_number: int
    legacy_record_id: Optional[str]
    person_id: str
    person_name: Optional[str]
    course_id: str
    course_name: str
    completion_date: date
    next_due_date: Optional[date]
    days_to_due: Optional[int]
    source_status: Optional[str]


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_text(value: Any) -> str:
    raw = _clean(value) or ""
    return re.sub(r"\s+", " ", raw).strip().lower()


def _parse_date(value: Any, *, field_name: str, required: bool) -> Optional[date]:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field_name} is required")
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        if required:
            raise ValueError(f"{field_name} is required")
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name} must be a valid date")


def _parse_optional_int(value: Any, *, field_name: str) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field_name} must be an integer or blank")
    if isinstance(value, (int, float)):
        if int(value) != value:
            raise ValueError(f"{field_name} must be an integer or blank")
        return int(value)
    raw = str(value).strip()
    if not raw:
        return None
    if re.fullmatch(r"-?\d+", raw):
        return int(raw)
    raise ValueError(f"{field_name} must be an integer or blank")


def parse_training_records_sheet(file_bytes: bytes, *, filename: str, sheet_name: str = "Training") -> list[dict[str, Any]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in {"xlsx", "xlsm", "xltx", "xltm"}:
        raise ValueError("Only Excel .xlsx/.xlsm files are supported for training history import.")
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for training history import. Install openpyxl.") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _locate_training_sheet(wb, sheet_name)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers, header_index = _resolve_training_headers(list(rows[0]))

    parsed: list[dict[str, Any]] = []
    for idx, values in enumerate(rows[1:], start=2):
        row = {header: (values[position] if position < len(values) else None) for header, position in header_index.items()}
        if not any(v is not None and str(v).strip() for v in row.values()):
            continue
        parsed.append({"row_number": idx, **row})
    return parsed


def _build_parsed_training_row(raw: dict[str, Any]) -> ParsedTrainingImportRow:
    person_id = (_clean(raw.get("PersonID")) or "").upper()
    person_name = _clean(raw.get("PersonName"))
    course_id = (_clean(raw.get("CourseID")) or "").upper()
    course_name = re.sub(r"\s+", " ", _clean(raw.get("CourseName")) or "").strip()
    if not person_id:
        raise ValueError("PersonID is required")
    if not course_id:
        raise ValueError("CourseID is required")
    if not course_name:
        raise ValueError("CourseName is required")
    completion_date = _parse_date(raw.get("LastTrainingDate"), field_name="LastTrainingDate", required=True)
    next_due_date = _parse_date(raw.get("NextDueDate"), field_name="NextDueDate", required=False)
    return ParsedTrainingImportRow(
        row_number=int(raw["row_number"]),
        legacy_record_id=_clean(raw.get("RecordID")),
        person_id=person_id,
        person_name=person_name,
        course_id=course_id,
        course_name=course_name,
        completion_date=completion_date,
        next_due_date=next_due_date,
        days_to_due=_parse_optional_int(raw.get("DaysToDue"), field_name="DaysToDue"),
        source_status=_clean(raw.get("Status")),
    )


def _legacy_import_remark(parsed: ParsedTrainingImportRow) -> str:
    parts = ["Imported from TRAINING.xlsx"]
    if parsed.legacy_record_id:
        parts.append(f"RecordID={parsed.legacy_record_id}")
    if parsed.source_status:
        parts.append(f"Status={parsed.source_status}")
    if parsed.days_to_due is not None:
        parts.append(f"DaysToDue={parsed.days_to_due}")
    if parsed.person_name:
        parts.append(f"SourceName={parsed.person_name}")
    return " | ".join(parts)


def _index_users(users: Iterable[accounts_models.User]) -> tuple[Dict[str, accounts_models.User], Dict[str, accounts_models.User], Dict[str, Optional[accounts_models.User]]]:
    by_staff: Dict[str, accounts_models.User] = {}
    by_id: Dict[str, accounts_models.User] = {}
    name_bucket: Dict[str, list[accounts_models.User]] = {}
    for user in users:
        staff = _clean(getattr(user, "staff_code", None))
        if staff:
            by_staff[staff.upper()] = user
        by_id[str(user.id).upper()] = user
        norm_name = _normalize_text(getattr(user, "full_name", None))
        if norm_name:
            name_bucket.setdefault(norm_name, []).append(user)
    by_name: Dict[str, Optional[accounts_models.User]] = {}
    for key, bucket in name_bucket.items():
        by_name[key] = bucket[0] if len(bucket) == 1 else None
    return by_staff, by_id, by_name


def _index_courses(courses: Iterable[models.TrainingCourse]) -> tuple[Dict[str, models.TrainingCourse], Dict[str, Optional[models.TrainingCourse]]]:
    by_code: Dict[str, models.TrainingCourse] = {}
    name_bucket: Dict[str, list[models.TrainingCourse]] = {}
    for course in courses:
        code = _clean(course.course_id)
        if code:
            by_code[code.upper()] = course
        norm_name = _normalize_text(course.course_name)
        if norm_name:
            name_bucket.setdefault(norm_name, []).append(course)
    by_name: Dict[str, Optional[models.TrainingCourse]] = {}
    for key, bucket in name_bucket.items():
        by_name[key] = bucket[0] if len(bucket) == 1 else None
    return by_code, by_name


def _match_user(
    parsed: ParsedTrainingImportRow,
    *,
    by_staff: Dict[str, accounts_models.User],
    by_id: Dict[str, accounts_models.User],
    by_name: Dict[str, Optional[accounts_models.User]],
) -> Optional[accounts_models.User]:
    if parsed.person_id in by_staff:
        return by_staff[parsed.person_id]
    if parsed.person_id in by_id:
        return by_id[parsed.person_id]
    if parsed.person_name:
        return by_name.get(_normalize_text(parsed.person_name))
    return None


def _match_course(
    parsed: ParsedTrainingImportRow,
    *,
    by_code: Dict[str, models.TrainingCourse],
    by_name: Dict[str, Optional[models.TrainingCourse]],
) -> Optional[models.TrainingCourse]:
    if parsed.course_id in by_code:
        return by_code[parsed.course_id]
    return by_name.get(_normalize_text(parsed.course_name))


def _is_one_off_course(course: models.TrainingCourse) -> bool:
    status = (_clean(getattr(course, "status", None)) or "").upper()
    if status == "ONE_OFF":
        return True
    if is_initial_course(course):
        return True
    return not bool(getattr(course, "frequency_months", None))


def _derive_latest_status(valid_until: Optional[date], *, course: models.TrainingCourse, today: date) -> str:
    if _is_one_off_course(course):
        return "OK"
    if not valid_until:
        return "OK"
    days = (valid_until - today).days
    if days < 0:
        return "OVERDUE"
    if days <= 60:
        return "DUE_SOON"
    return "OK"




def _upsert_remark_token(remarks: Optional[str], key: str, value: Optional[str]) -> str:
    tokens: list[str] = []
    if remarks:
        tokens = [part.strip() for part in str(remarks).split("|") if part and str(part).strip()]
    normalized_key = key.strip().lower()
    kept: list[str] = []
    replaced = False
    for token in tokens:
        if "=" not in token:
            kept.append(token)
            continue
        token_key, _token_value = token.split("=", 1)
        if token_key.strip().lower() == normalized_key:
            if value is not None:
                kept.append(f"{key}={value}")
                replaced = True
            continue
        kept.append(token)
    if value is not None and not replaced:
        kept.append(f"{key}={value}")
    return " | ".join(kept)

def _compute_purge_after(user: accounts_models.User, completion_date: date) -> date:
    deactivated_at = getattr(user, "deactivated_at", None)
    if deactivated_at is not None:
        base_date = max(completion_date, deactivated_at.date())
        return add_months(base_date, 24)
    if not bool(getattr(user, "is_active", True)):
        return add_months(completion_date, 24)
    return add_months(completion_date, 60)


def _apply_record_lifecycle(
    db: Session,
    *,
    amo_id: str,
    user: accounts_models.User,
    course: models.TrainingCourse,
    actor_user_id: Optional[str],
    dry_run: bool,
) -> Dict[str, str]:
    rows = (
        db.query(models.TrainingRecord)
        .filter(
            models.TrainingRecord.amo_id == amo_id,
            models.TrainingRecord.user_id == str(user.id),
            models.TrainingRecord.course_id == str(course.id),
        )
        .order_by(models.TrainingRecord.completion_date.asc(), models.TrainingRecord.created_at.asc(), models.TrainingRecord.id.asc())
        .all()
    )
    if not rows:
        return {}
    latest = rows[-1]
    today = date.today()
    status_map: Dict[str, str] = {}

    for row in rows[:-1]:
        status_map[row.id] = "RENEWED"
        if not dry_run:
            row.valid_until = None
            row.remarks = _upsert_remark_token(row.remarks, "LifecycleStatus", "RENEWED")
            db.add(row)

    computed_valid_until = latest.valid_until
    if computed_valid_until is None and not _is_one_off_course(course) and course.frequency_months:
        computed_valid_until = add_months(latest.completion_date, int(course.frequency_months))
    latest_status = _derive_latest_status(computed_valid_until, course=course, today=today)
    status_map[latest.id] = latest_status
    if not dry_run:
        latest.valid_until = computed_valid_until if latest_status != "RENEWED" else None
        latest.remarks = _upsert_remark_token(latest.remarks, "LifecycleStatus", latest_status)
        latest.verified_at = latest.verified_at or datetime.now(timezone.utc)
        latest.verified_by_user_id = latest.verified_by_user_id or actor_user_id
        db.add(latest)

    return status_map


def import_training_records_rows(
    db: Session,
    *,
    amo_id: str,
    rows: list[dict[str, Any]],
    dry_run: bool,
    actor_user_id: Optional[str] = None,
) -> schemas.TrainingRecordImportSummary:
    issues: list[schemas.TrainingRecordImportRowIssue] = []
    preview_rows: list[schemas.TrainingRecordImportRowPreview] = []
    created_records = 0
    updated_records = 0
    unchanged_rows = 0
    skipped_rows = 0
    matched_inactive_rows = 0

    parsed_rows: list[ParsedTrainingImportRow] = []
    seen_import_keys: set[tuple[str, str, date]] = set()

    for raw in rows:
        try:
            parsed = _build_parsed_training_row(raw)
        except ValueError as exc:
            issues.append(
                schemas.TrainingRecordImportRowIssue(
                    row_number=int(raw.get("row_number") or 0),
                    legacy_record_id=_clean(raw.get("RecordID")),
                    person_id=_clean(raw.get("PersonID")),
                    course_id=_clean(raw.get("CourseID")),
                    reason=str(exc),
                )
            )
            skipped_rows += 1
            continue

        dedupe_key = (parsed.person_id, parsed.course_id, parsed.completion_date)
        if dedupe_key in seen_import_keys:
            issues.append(
                schemas.TrainingRecordImportRowIssue(
                    row_number=parsed.row_number,
                    legacy_record_id=parsed.legacy_record_id,
                    person_id=parsed.person_id,
                    course_id=parsed.course_id,
                    reason="Duplicate person/course/completion combination inside import file.",
                )
            )
            skipped_rows += 1
            continue
        seen_import_keys.add(dedupe_key)
        parsed_rows.append(parsed)

    users = (
        db.query(accounts_models.User)
        .filter(accounts_models.User.amo_id == amo_id, accounts_models.User.is_system_account.is_(False))
        .all()
    )
    courses = db.query(models.TrainingCourse).filter(models.TrainingCourse.amo_id == amo_id).all()

    by_staff, by_user_id, by_name = _index_users(users)
    by_course_code, by_course_name = _index_courses(courses)

    matched_pairs: list[tuple[ParsedTrainingImportRow, accounts_models.User, models.TrainingCourse]] = []
    affected_exact_pairs: set[tuple[str, str]] = set()

    for parsed in parsed_rows:
        user = _match_user(parsed, by_staff=by_staff, by_id=by_user_id, by_name=by_name)
        if user is None:
            issues.append(
                schemas.TrainingRecordImportRowIssue(
                    row_number=parsed.row_number,
                    legacy_record_id=parsed.legacy_record_id,
                    person_id=parsed.person_id,
                    course_id=parsed.course_id,
                    reason="PersonID could not be matched to an AMO user (staff code, user id, or unique full name).",
                )
            )
            preview_rows.append(
                schemas.TrainingRecordImportRowPreview(
                    row_number=parsed.row_number,
                    legacy_record_id=parsed.legacy_record_id,
                    person_id=parsed.person_id,
                    person_name=parsed.person_name,
                    course_id=parsed.course_id,
                    course_name=parsed.course_name,
                    completion_date=parsed.completion_date,
                    next_due_date=parsed.next_due_date,
                    days_to_due=parsed.days_to_due,
                    source_status=parsed.source_status,
                    action="SKIP",
                    matched_user_id=None,
                    matched_user_name=None,
                    matched_user_active=None,
                    matched_course_pk=None,
                    matched_course_name=None,
                    existing_record_id=None,
                    changes=[],
                    reason="Unmatched user",
                )
            )
            skipped_rows += 1
            continue

        course = _match_course(parsed, by_code=by_course_code, by_name=by_course_name)
        if course is None:
            issues.append(
                schemas.TrainingRecordImportRowIssue(
                    row_number=parsed.row_number,
                    legacy_record_id=parsed.legacy_record_id,
                    person_id=parsed.person_id,
                    course_id=parsed.course_id,
                    reason="CourseID could not be matched to an existing course catalog entry.",
                )
            )
            preview_rows.append(
                schemas.TrainingRecordImportRowPreview(
                    row_number=parsed.row_number,
                    legacy_record_id=parsed.legacy_record_id,
                    person_id=parsed.person_id,
                    person_name=parsed.person_name,
                    course_id=parsed.course_id,
                    course_name=parsed.course_name,
                    completion_date=parsed.completion_date,
                    next_due_date=parsed.next_due_date,
                    days_to_due=parsed.days_to_due,
                    source_status=parsed.source_status,
                    action="SKIP",
                    matched_user_id=str(user.id),
                    matched_user_name=getattr(user, "full_name", None),
                    matched_user_active=bool(getattr(user, "is_active", False)),
                    matched_course_pk=None,
                    matched_course_name=None,
                    existing_record_id=None,
                    changes=[],
                    reason="Unmatched course",
                )
            )
            skipped_rows += 1
            continue

        matched_pairs.append((parsed, user, course))
        affected_exact_pairs.add((str(user.id), str(course.id)))
        if not bool(getattr(user, "is_active", False)):
            matched_inactive_rows += 1

    existing_records = []
    if affected_exact_pairs:
        affected_user_ids = sorted({user_id for user_id, _ in affected_exact_pairs})
        affected_course_ids = sorted({course_id for _, course_id in affected_exact_pairs})
        existing_records = (
            db.query(models.TrainingRecord)
            .filter(
                models.TrainingRecord.amo_id == amo_id,
                models.TrainingRecord.user_id.in_(affected_user_ids),
                models.TrainingRecord.course_id.in_(affected_course_ids),
            )
            .all()
        )
    existing_by_key: Dict[tuple[str, str, date], models.TrainingRecord] = {}
    for record in existing_records:
        key = (str(record.user_id), str(record.course_id), record.completion_date)
        current = existing_by_key.get(key)
        if current is None or record.created_at > current.created_at:
            existing_by_key[key] = record

    lifecycle_status_by_id: Dict[str, str] = {}
    now = datetime.now(timezone.utc)

    for parsed, user, course in matched_pairs:
        target_valid_until = parsed.next_due_date or (
            add_months(parsed.completion_date, int(course.frequency_months)) if course.frequency_months and not _is_one_off_course(course) else None
        )
        target_remarks = _legacy_import_remark(parsed)
        existing = existing_by_key.get((str(user.id), str(course.id), parsed.completion_date))
        changes: list[schemas.TrainingRecordImportChange] = []
        if existing is None:
            action = "CREATE"
            created_records += 1
            if not dry_run:
                record = models.TrainingRecord(
                    amo_id=amo_id,
                    user_id=str(user.id),
                    course_id=str(course.id),
                    event_id=None,
                    completion_date=parsed.completion_date,
                    valid_until=target_valid_until,
                    hours_completed=getattr(course, 'nominal_hours', None),
                    exam_score=None,
                    certificate_reference=None,
                    remarks=target_remarks,
                    is_manual_entry=True,
                    verification_status=models.TrainingRecordVerificationStatus.VERIFIED,
                    verified_at=now,
                    verified_by_user_id=actor_user_id,
                    verification_comment="Verified on import from TRAINING.xlsx.",
                    created_by_user_id=actor_user_id,
                )
                db.add(record)
                db.flush()
                existing = record
                existing_by_key[(str(user.id), str(course.id), parsed.completion_date)] = record
        else:
            action = "UNCHANGED"
            compare_pairs = [
                ("valid_until", "Next due date", existing.valid_until.isoformat() if existing.valid_until else None, target_valid_until.isoformat() if target_valid_until else None),
                ("remarks", "Import note", existing.remarks or None, target_remarks),
                ("hours_completed", "Nominal hours", str(existing.hours_completed) if existing.hours_completed is not None else None, str(getattr(course, 'nominal_hours', None)) if getattr(course, 'nominal_hours', None) is not None else None),
            ]
            for field, label, old_value, new_value in compare_pairs:
                if old_value != new_value:
                    changes.append(schemas.TrainingRecordImportChange(field=field, label=label, old_value=old_value, new_value=new_value))
            if not existing.is_manual_entry:
                changes.append(schemas.TrainingRecordImportChange(field="is_manual_entry", label="Manual entry flag", old_value="false", new_value="true"))
            if existing.verification_status != models.TrainingRecordVerificationStatus.VERIFIED:
                changes.append(schemas.TrainingRecordImportChange(field="verification_status", label="Verification status", old_value=str(existing.verification_status), new_value=str(models.TrainingRecordVerificationStatus.VERIFIED)))
            if changes:
                action = "UPDATE"
                updated_records += 1
                if not dry_run:
                    existing.valid_until = target_valid_until
                    existing.remarks = target_remarks
                    if getattr(course, 'nominal_hours', None) is not None:
                        existing.hours_completed = existing.hours_completed or getattr(course, 'nominal_hours', None)
                    existing.is_manual_entry = True
                    existing.verification_status = models.TrainingRecordVerificationStatus.VERIFIED
                    existing.verified_at = now
                    existing.verified_by_user_id = actor_user_id
                    existing.verification_comment = "Verified on import from TRAINING.xlsx."
                    if actor_user_id:
                        existing.created_by_user_id = existing.created_by_user_id or actor_user_id
                    db.add(existing)
            else:
                unchanged_rows += 1

        preview_rows.append(
            schemas.TrainingRecordImportRowPreview(
                row_number=parsed.row_number,
                legacy_record_id=parsed.legacy_record_id,
                person_id=parsed.person_id,
                person_name=parsed.person_name,
                course_id=parsed.course_id,
                course_name=parsed.course_name,
                completion_date=parsed.completion_date,
                next_due_date=target_valid_until,
                days_to_due=parsed.days_to_due,
                source_status=parsed.source_status,
                action=action,
                matched_user_id=str(user.id),
                matched_user_name=getattr(user, "full_name", None),
                matched_user_active=bool(getattr(user, "is_active", False)),
                matched_course_pk=str(course.id),
                matched_course_name=course.course_name,
                existing_record_id=existing.id if existing is not None else None,
                changes=changes,
                reason="Matched inactive/dormant user; record will be retained without alerting." if not bool(getattr(user, "is_active", False)) else None,
            )
        )

    if not dry_run:
        db.flush()

    affected_user_objs = {str(user.id): user for _parsed, user, _course in matched_pairs}
    affected_course_objs = {str(course.id): course for _parsed, _user, course in matched_pairs}
    for user_id, course_id in sorted(affected_exact_pairs):
        user = affected_user_objs[user_id]
        course = affected_course_objs[course_id]
        status_map = _apply_record_lifecycle(db, amo_id=amo_id, user=user, course=course, actor_user_id=actor_user_id, dry_run=dry_run)
        lifecycle_status_by_id.update(status_map)

    if not dry_run:
        for record in existing_by_key.values():
            if record.id in lifecycle_status_by_id:
                db.add(record)
        db.commit()

    for preview in preview_rows:
        if preview.existing_record_id and preview.existing_record_id in lifecycle_status_by_id:
            lifecycle = lifecycle_status_by_id[preview.existing_record_id]
            if preview.action != "SKIP":
                preview.source_status = lifecycle
                preview.next_due_date = None if lifecycle == "RENEWED" else preview.next_due_date

    return schemas.TrainingRecordImportSummary(
        dry_run=dry_run,
        total_rows=len(rows),
        created_records=created_records,
        updated_records=updated_records,
        unchanged_rows=unchanged_rows,
        skipped_rows=skipped_rows,
        matched_inactive_rows=matched_inactive_rows,
        issues=issues,
        preview_rows=preview_rows,
    )
