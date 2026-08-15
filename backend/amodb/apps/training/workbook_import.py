from __future__ import annotations

import hashlib
import os
import re
import secrets
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from time import perf_counter
from typing import Any, Callable, Iterable, Optional

from sqlalchemy import func, or_
from sqlalchemy.exc import DBAPIError, IntegrityError, OperationalError
from sqlalchemy.orm import Session

from ...database import SessionLocal
from ...user_id import generate_user_id
from ..accounts import models as account_models
from ..accounts.services import get_password_hash
from ..audit import services as audit_services
from . import models as training_models
from . import records_import
from .licence_rules import infer_licence_authority
from .workbook_models import (
    PersonnelLicence,
    TrainingCourseRoleRule,
    TrainingPersonRole,
    TrainingRoleGroup,
    TrainingWorkbookImportJob,
    TrainingWorkbookImportRow,
    TrainingWorkbookImportSheet,
)


WORKBOOK_SHEETS: dict[str, dict[str, Any]] = {
    "People": {"classification": "PERSONNEL", "destination": "Personnel register, access review and licence register", "operational": True},
    "Courses": {"classification": "COURSES", "destination": "Course catalogue", "operational": True},
    "Training": {"classification": "TRAINING_HISTORY", "destination": "Immutable personnel training history", "operational": True},
    "tblRoleGroups": {"classification": "ROLE_GROUPS", "destination": "Training applicability groups", "operational": True},
    "tblPersonRoles": {"classification": "PERSON_ROLES", "destination": "Personnel role-group assignments", "operational": True},
    "tblCourseMatrix": {"classification": "COURSE_MATRIX", "destination": "Course requirement matrix", "operational": True},
    "Params": {"classification": "CONFIGURATION", "destination": "Training policy settings", "operational": False},
    "Overdue": {"classification": "DERIVED_VIEW", "destination": "Portal overdue and due-soon queues", "operational": False},
    "Next_Batch": {"classification": "DERIVED_VIEW", "destination": "Smart scheduling and roster builder", "operational": False},
    "Individual_Lookup": {"classification": "DERIVED_VIEW", "destination": "Personnel training profile and licence view", "operational": False},
    "Course_Audit": {"classification": "DERIVED_VIEW", "destination": "Training data rectification queue", "operational": False},
    "Sheet1": {"classification": "HELPER", "destination": "Reference-only workbook helper", "operational": False},
}

OPERATIONAL_ORDER = ["Courses", "People", "tblRoleGroups", "tblPersonRoles", "tblCourseMatrix", "Training"]


def _positive_int_env(name: str, default: int) -> int:
    try:
        return max(1, int(os.getenv(name, str(default))))
    except (TypeError, ValueError):
        return default


PREVIEW_PROGRESS_BATCH = _positive_int_env("TRAINING_WORKBOOK_PREVIEW_PROGRESS_BATCH", 40)
COMMIT_PROGRESS_BATCH = _positive_int_env("TRAINING_WORKBOOK_COMMIT_PROGRESS_BATCH", 25)
LICENCE_CATEGORY_MAX_CHARS = _positive_int_env("TRAINING_LICENCE_CATEGORY_MAX_CHARS", 32767)


class PersonnelIdentityChanged(RuntimeError):
    def __init__(self, row_id: str, message: str):
        super().__init__(message)
        self.row_id = row_id


class WorkbookRowCommitError(RuntimeError):
    """Surface the original row failure after the atomic transaction rolls back."""

    def __init__(self, row_id: str, sheet: str, source_row: int, message: str):
        self.row_id = row_id
        self.sheet = sheet
        self.source_row = source_row
        super().__init__(f"{sheet} row {source_row} could not be committed: {message}")


class WorkbookCommitLeaseLost(RuntimeError):
    """Stop an obsolete worker before it can publish or commit more work."""


def new_commit_attempt_token() -> str:
    """Return an opaque lease token used to fence superseded commit workers."""
    return secrets.token_urlsafe(24)


def _commit_attempt_token(job: TrainingWorkbookImportJob) -> str:
    return str((job.summary_json or {}).get("active_commit_token") or "")


def _is_transient_database_error(exc: BaseException) -> bool:
    """Connection loss/restart is retryable; data-integrity failures are not."""
    if isinstance(exc, IntegrityError):
        return False
    if isinstance(exc, OperationalError):
        return True
    return isinstance(exc, DBAPIError) and bool(exc.connection_invalidated)


@dataclass
class PersonnelCommitIndexes:
    profiles_by_person: dict[str, account_models.PersonnelProfile]
    profiles_by_email: dict[str, account_models.PersonnelProfile]
    users_by_staff: dict[str, account_models.User]
    users_by_email: dict[str, account_models.User]
    users_by_id: dict[str, account_models.User]
    department_ids_by_token: dict[str, str]
    licences_by_profile_authority: dict[tuple[str, str], list[PersonnelLicence]]


@dataclass(frozen=True)
class PersonCommitResult:
    entity_id: Optional[str]
    action: str
    profile_created: bool = False
    portal_account_created: bool = False
    non_login_identity_created: bool = False


def _licence_reconciliation_status(current_number: Optional[str], imported_number: Optional[str]) -> str:
    """Classify an existing authority credential when a workbook is reapplied."""
    if not imported_number:
        return "RETIRED"
    return "ACTIVE" if current_number == imported_number else "SUPERSEDED"


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def upper(value: Any) -> str:
    return (clean(value) or "").upper()


def _licence_category(value: Any, label: str) -> Optional[str]:
    category = clean(value)
    if category and len(category) > LICENCE_CATEGORY_MAX_CHARS:
        raise ValueError(
            f"{label} exceeds the {LICENCE_CATEGORY_MAX_CHARS:,}-character licence-category limit."
        )
    return category


def bool_value(value: Any, default: bool = False) -> bool:
    raw = (clean(value) or "").lower()
    if not raw:
        return default
    return raw in {"yes", "y", "true", "1", "active", "required"}


def date_value(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date '{raw}'")


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _load_workbook(path: str, *, data_only: bool = True):
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Training workbook imports.") from exc
    return openpyxl.load_workbook(path, data_only=data_only, read_only=True, keep_vba=path.lower().endswith(".xlsm"))


def _sheet_rows(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(cell) or "" for cell in rows[0]]
    while headers and not headers[-1]:
        headers.pop()
    result: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows[1:], start=2):
        payload = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]}
        if not any(clean(value) for value in payload.values()):
            continue
        result.append({"row_number": row_number, **payload})
    return result


def _sheet_visibility(workbook, name: str) -> str:
    return str(getattr(workbook[name], "sheet_state", "visible") or "visible").upper()


def _set_job_progress(
    db: Session,
    job: TrainingWorkbookImportJob,
    *,
    stage: Optional[str] = None,
    sheet: Optional[str] = None,
    label: Optional[str] = None,
    processed_delta: int = 0,
) -> None:
    if stage:
        job.stage = stage
    if sheet is not None:
        job.current_sheet = sheet
    if label is not None:
        job.current_record_label = label[:255]
    if processed_delta:
        job.processed_rows += processed_delta
    job.updated_at = utcnow()
    db.add(job)
    should_publish = job.processed_rows % PREVIEW_PROGRESS_BATCH == 0 or job.processed_rows >= job.total_rows
    if should_publish:
        db.commit()
        db.refresh(job)
        if job.cancel_requested:
            raise RuntimeError("IMPORT_CANCELLED")
    else:
        db.flush()


def _row(
    *,
    job_id: str,
    sheet: str,
    row_number: int,
    entity_type: str,
    source_key: Optional[str],
    label: Optional[str],
    action: str,
    status: str = "READY",
    decision_required: bool = False,
    decision_options: Optional[list[str]] = None,
    payload: Optional[dict[str, Any]] = None,
    changes: Optional[list[dict[str, Any]]] = None,
    issue_code: Optional[str] = None,
    issue_message: Optional[str] = None,
) -> TrainingWorkbookImportRow:
    return TrainingWorkbookImportRow(
        job_id=job_id,
        sheet_name=sheet,
        source_row=row_number,
        entity_type=entity_type,
        source_key=source_key,
        display_label=label,
        proposed_action=action,
        status=status,
        decision_required=decision_required,
        decision_options=decision_options or [],
        payload_json={key: json_value(value) for key, value in (payload or {}).items()},
        changes_json=changes or [],
        issue_code=issue_code,
        issue_message=issue_message,
    )


def _counter(sheet: TrainingWorkbookImportSheet, action: str, *, review: bool = False, failed: bool = False) -> None:
    sheet.processed_rows += 1
    if action == "CREATE":
        sheet.created_count += 1
    elif action == "UPDATE":
        sheet.updated_count += 1
    elif action == "UNCHANGED":
        sheet.unchanged_count += 1
    elif action == "SKIP":
        sheet.skipped_count += 1
    if review:
        sheet.review_count += 1
    if failed:
        sheet.failed_count += 1


def _changes(existing: Any, payload: dict[str, Any], mapping: dict[str, str]) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for source_field, model_field in mapping.items():
        new = payload.get(source_field)
        old = getattr(existing, model_field, None)
        old_cmp = old.isoformat() if isinstance(old, (date, datetime)) else old
        new_cmp = new.isoformat() if isinstance(new, (date, datetime)) else new
        if (old_cmp or None) != (new_cmp or None):
            changes.append({"field": model_field, "old": old_cmp, "new": new_cmp})
    return changes


def _person_payload(raw: dict[str, Any]) -> dict[str, Any]:
    person_id = upper(raw.get("PersonID"))
    first = upper(raw.get("FIRSTNAME"))
    last = upper(raw.get("LASTNAME"))
    if not person_id or person_id == "TOTAL":
        raise ValueError("PersonID is required")
    if not first or not last:
        raise ValueError("FIRSTNAME and LASTNAME are required")
    email = (clean(raw.get("Email")) or "").lower() or None
    return {
        "person_id": person_id,
        "first_name": first.title(),
        "last_name": last.title(),
        "full_name": clean(raw.get("PersonName")) or f"{first.title()} {last.title()}",
        "national_id": clean(raw.get("nid")),
        "category_reg_2013": _licence_category(raw.get("Category (Reg. 2013)"), "Category (Reg. 2013)"),
        "category_reg_2018": _licence_category(raw.get("Category (Reg. 2018)"), "Category (Reg. 2018)"),
        "kamel_no": clean(raw.get("KAMEL NO:")) or clean(raw.get("AMEL NO:")),
        "internal_stamp_no": clean(raw.get("Internal Certification Stamp No:")),
        "initial_authorization_date": date_value(raw.get("initial_auth")),
        "department": upper(raw.get("Department")) or None,
        "position_title": upper(raw.get("Position")) or None,
        "phone_number": clean(raw.get("PhoneNumber")),
        "secondary_phone": clean(raw.get("secondary_phone")),
        "email": email,
        "hire_date": date_value(raw.get("HireDate")),
        "employment_status": clean(raw.get("Employment_Status")),
        "status": clean(raw.get("Status")) or "Active",
        "date_of_birth": date_value(raw.get("DOB")),
        "birth_place": clean(raw.get("birthplace")),
        "e_amel": clean(raw.get("E-AMEL")),
        "g_amel": clean(raw.get("G-AMEL")),
    }


def _course_payload(raw: dict[str, Any], *, default_frequency_months: Optional[int] = None) -> dict[str, Any]:
    course_id = upper(raw.get("CourseID"))
    name = clean(raw.get("CourseName"))
    if not course_id or not name:
        raise ValueError("CourseID and CourseName are required")
    status = clean(raw.get("Status")) or clean(raw.get("CourseType")) or "One_Off"
    canonical = {"initial": "Initial", "recurrent": "Recurrent", "one_off": "One_Off", "one-off": "One_Off", "one off": "One_Off"}.get(status.lower())
    if not canonical:
        raise ValueError("Course type must be Initial, Recurrent or One_Off")
    frequency = raw.get("FrequencyMonths")
    if frequency in (None, ""):
        months = default_frequency_months if canonical == "Recurrent" else None
    else:
        months = int(float(frequency))
        if months < 0:
            raise ValueError("FrequencyMonths cannot be negative")
    return {
        "course_id": course_id,
        "course_name": name,
        "frequency_months": months,
        "status": canonical,
        "category_raw": clean(raw.get("Category")),
        "is_mandatory": bool_value(raw.get("Mandatory")),
        "scope": clean(raw.get("Scope")),
        "regulatory_reference": clean(raw.get("Reference")),
        "is_active": bool_value(raw.get("Active"), True),
    }



def _workbook_params(rows: list[dict[str, Any]]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for row in rows:
        setting = clean(row.get("Setting"))
        if not setting:
            continue
        values[setting] = row.get("Value")
    return values


def _default_frequency_months(params: dict[str, Any]) -> Optional[int]:
    raw = params.get("Default Frequency (months)")
    if raw in (None, ""):
        return None
    value = int(float(raw))
    if value <= 0:
        raise ValueError("Default Frequency (months) must be a positive integer")
    return value


def _role_from_position(position: Optional[str]) -> account_models.AccountRole:
    text = upper(position)
    mapping = {
        "QUALITY MANAGER": account_models.AccountRole.QUALITY_MANAGER,
        "QUALITY INSPECTOR": account_models.AccountRole.QUALITY_INSPECTOR,
        "AUDITOR": account_models.AccountRole.AUDITOR,
        "CERTIFYING ENGINEER": account_models.AccountRole.CERTIFYING_ENGINEER,
        "CERTIFYING TECHNICIAN": account_models.AccountRole.CERTIFYING_TECHNICIAN,
        "PLANNING ENGINEER": account_models.AccountRole.PLANNING_ENGINEER,
        "PRODUCTION ENGINEER": account_models.AccountRole.PRODUCTION_ENGINEER,
        "STORES MANAGER": account_models.AccountRole.STORES_MANAGER,
        "STOREKEEPER": account_models.AccountRole.STOREKEEPER,
        "PROCUREMENT OFFICER": account_models.AccountRole.PROCUREMENT_OFFICER,
    }
    return mapping.get(text, account_models.AccountRole.TECHNICIAN)


def _preview_people(db: Session, job: TrainingWorkbookImportJob, sheet: TrainingWorkbookImportSheet, rows: list[dict[str, Any]]) -> None:
    profiles = db.query(account_models.PersonnelProfile).filter(account_models.PersonnelProfile.amo_id == job.amo_id).all()
    users = db.query(account_models.User).filter(account_models.User.amo_id == job.amo_id).all()
    by_person = {upper(item.person_id): item for item in profiles}
    by_profile_email = {(item.email or "").lower(): item for item in profiles if item.email}
    by_staff = {upper(item.staff_code): item for item in users}
    by_user_email = {(item.email or "").lower(): item for item in users if item.email}

    for raw in rows:
        row_number = int(raw["row_number"])
        try:
            payload = _person_payload(raw)
            existing_profile = by_person.get(payload["person_id"])
            email_profile = by_profile_email.get(payload["email"] or "") if payload["email"] else None
            existing_user = by_staff.get(payload["person_id"])
            email_user = by_user_email.get(payload["email"] or "") if payload["email"] else None
            profile_email_conflict = bool(
                email_profile
                and upper(email_profile.person_id) != payload["person_id"]
            )
            user_email_conflict = bool(
                email_user
                and upper(email_user.staff_code) != payload["person_id"]
            )
            split_profile_conflict = bool(
                existing_profile
                and email_profile
                and existing_profile.id != email_profile.id
            )
            split_user_conflict = bool(
                existing_user
                and email_user
                and existing_user.id != email_user.id
            )
            if profile_email_conflict or user_email_conflict or split_profile_conflict or split_user_conflict:
                item = _row(
                    job_id=job.id, sheet="People", row_number=row_number, entity_type="PERSON",
                    source_key=payload["person_id"], label=payload["full_name"], action="UPDATE", status="REVIEW",
                    decision_required=True,
                    decision_options=["KEEP_EXISTING_EMAIL", "SKIP"] if (existing_profile or existing_user) else ["SKIP"],
                    payload=payload, issue_code="IDENTITY_CONFLICT",
                    issue_message="PersonID and email resolve to different existing personnel/account records.",
                )
                _counter(sheet, "UPDATE", review=True)
            else:
                profile = existing_profile or email_profile
                user = existing_user or email_user or (db.get(account_models.User, profile.user_id) if profile and profile.user_id else None)
                mapping = {
                    "first_name": "first_name", "last_name": "last_name", "full_name": "full_name",
                    "national_id": "national_id", "kamel_no": "amel_no", "internal_stamp_no": "internal_certification_stamp_no",
                    "initial_authorization_date": "initial_authorization_date", "department": "department",
                    "position_title": "position_title", "phone_number": "phone_number", "secondary_phone": "secondary_phone",
                    "email": "email", "hire_date": "hire_date", "employment_status": "employment_status",
                    "status": "status", "date_of_birth": "date_of_birth", "birth_place": "birth_place",
                }
                change_list = _changes(profile, payload, mapping) if profile else []
                if profile and user:
                    action = "UPDATE" if change_list else "UNCHANGED"
                    item = _row(job_id=job.id, sheet="People", row_number=row_number, entity_type="PERSON", source_key=payload["person_id"], label=payload["full_name"], action=action, payload=payload, changes=change_list)
                    _counter(sheet, action)
                elif user and not profile:
                    item = _row(
                        job_id=job.id,
                        sheet="People",
                        row_number=row_number,
                        entity_type="PERSON",
                        source_key=payload["person_id"],
                        label=payload["full_name"],
                        action="UPDATE",
                        status="REVIEW",
                        decision_required=True,
                        decision_options=["LINK_EXISTING_ACCOUNT", "SKIP"],
                        payload=payload,
                        changes=change_list,
                        issue_code="PROFILE_NOT_LINKED",
                        issue_message="A portal account exists for this staff code or email. Review and link the personnel profile, or skip the row.",
                    )
                    _counter(sheet, "UPDATE", review=True)
                else:
                    action = "UPDATE" if profile else "CREATE"
                    options = ["PROFILE_ONLY", "SKIP"] if not payload["email"] else ["CREATE_ACCOUNT", "PROFILE_ONLY", "SKIP"]
                    item = _row(
                        job_id=job.id, sheet="People", row_number=row_number, entity_type="PERSON",
                        source_key=payload["person_id"], label=payload["full_name"], action=action, status="REVIEW",
                        decision_required=True, decision_options=options, payload=payload, changes=change_list,
                        issue_code="NEW_PERSON" if not profile else "ACCOUNT_NOT_LINKED",
                        issue_message="Review whether this person should enter approval/onboarding or remain a non-login personnel identity.",
                    )
                    _counter(sheet, action, review=True)
            db.add(item)
        except Exception as exc:
            db.add(_row(job_id=job.id, sheet="People", row_number=row_number, entity_type="PERSON", source_key=upper(raw.get("PersonID")) or None, label=clean(raw.get("PersonName")), action="SKIP", status="FAILED", payload=raw, issue_code="INVALID_PERSON", issue_message=str(exc)))
            _counter(sheet, "SKIP", failed=True)
        _set_job_progress(db, job, stage="VALIDATING", sheet="People", label=f"{clean(raw.get('PersonName')) or clean(raw.get('PersonID')) or 'Personnel row'}", processed_delta=1)


def _preview_courses(db: Session, job: TrainingWorkbookImportJob, sheet: TrainingWorkbookImportSheet, rows: list[dict[str, Any]], *, default_frequency_months: Optional[int] = None) -> None:
    existing = {upper(item.course_id): item for item in db.query(training_models.TrainingCourse).filter(training_models.TrainingCourse.amo_id == job.amo_id).all()}
    seen: set[str] = set()
    mapping = {"course_name": "course_name", "frequency_months": "frequency_months", "status": "status", "category_raw": "category_raw", "is_mandatory": "is_mandatory", "scope": "scope", "regulatory_reference": "regulatory_reference", "is_active": "is_active"}
    for raw in rows:
        row_number = int(raw["row_number"])
        try:
            payload = _course_payload(raw, default_frequency_months=default_frequency_months)
            if payload["course_id"] in seen:
                raise ValueError("Duplicate CourseID inside workbook")
            seen.add(payload["course_id"])
            current = existing.get(payload["course_id"])
            change_list = _changes(current, payload, mapping) if current else []
            action = "CREATE" if not current else "UPDATE" if change_list else "UNCHANGED"
            db.add(_row(job_id=job.id, sheet="Courses", row_number=row_number, entity_type="COURSE", source_key=payload["course_id"], label=payload["course_name"], action=action, payload=payload, changes=change_list))
            _counter(sheet, action)
        except Exception as exc:
            db.add(_row(job_id=job.id, sheet="Courses", row_number=row_number, entity_type="COURSE", source_key=upper(raw.get("CourseID")) or None, label=clean(raw.get("CourseName")), action="SKIP", status="FAILED", payload=raw, issue_code="INVALID_COURSE", issue_message=str(exc)))
            _counter(sheet, "SKIP", failed=True)
        _set_job_progress(db, job, stage="VALIDATING", sheet="Courses", label=f"{clean(raw.get('CourseID')) or 'Course row'}", processed_delta=1)


def _preview_training(
    db: Session,
    job: TrainingWorkbookImportJob,
    sheet: TrainingWorkbookImportSheet,
    rows: list[dict[str, Any]],
    *,
    workbook_people: set[str],
    workbook_courses: set[str],
) -> None:
    users = db.query(account_models.User).filter(
        account_models.User.amo_id == job.amo_id,
        account_models.User.is_system_account.is_(False),
    ).all()
    courses = db.query(training_models.TrainingCourse).filter(
        training_models.TrainingCourse.amo_id == job.amo_id,
    ).all()
    by_staff, by_user_id, by_name = records_import._index_users(users)
    by_code, by_course_name = records_import._index_courses(courses)
    existing = {
        (str(item.user_id), str(item.course_id), item.completion_date): item
        for item in db.query(training_models.TrainingRecord)
        .filter(training_models.TrainingRecord.amo_id == job.amo_id)
        .all()
    }
    seen: set[tuple[str, str, date]] = set()
    for raw in rows:
        row_number = int(raw["row_number"])
        try:
            parsed = records_import._build_parsed_training_row(raw)
            source_key = f"{parsed.person_id}:{parsed.course_id}:{parsed.completion_date.isoformat()}"
            dedupe = (parsed.person_id, parsed.course_id, parsed.completion_date)
            if dedupe in seen:
                raise ValueError("Duplicate person/course/completion combination inside workbook")
            seen.add(dedupe)
            user = records_import._match_user(parsed, by_staff=by_staff, by_id=by_user_id, by_name=by_name)
            course = records_import._match_course(parsed, by_code=by_code, by_name=by_course_name)
            payload = {
                "RecordID": parsed.legacy_record_id,
                "PersonID": parsed.person_id,
                "PersonName": parsed.person_name,
                "CourseID": parsed.course_id,
                "CourseName": parsed.course_name,
                "LastTrainingDate": parsed.completion_date,
                "NextDueDate": parsed.next_due_date,
                "DaysToDue": parsed.days_to_due,
                "Status": parsed.source_status,
            }
            person_resolves_from_workbook = user is None and parsed.person_id in workbook_people
            course_resolves_from_workbook = course is None and parsed.course_id in workbook_courses
            if user is None and not person_resolves_from_workbook:
                item = _row(
                    job_id=job.id,
                    sheet="Training",
                    row_number=row_number,
                    entity_type="TRAINING_RECORD",
                    source_key=source_key,
                    label=f"{parsed.person_name or parsed.person_id} · {parsed.course_name}",
                    action="SKIP",
                    status="FAILED",
                    payload=payload,
                    issue_code="UNMATCHED_PERSON",
                    issue_message="PersonID is not present in the People sheet or the AMO personnel register.",
                )
                _counter(sheet, "SKIP", failed=True)
            elif course is None and not course_resolves_from_workbook:
                item = _row(
                    job_id=job.id,
                    sheet="Training",
                    row_number=row_number,
                    entity_type="TRAINING_RECORD",
                    source_key=source_key,
                    label=f"{parsed.person_name or parsed.person_id} · {parsed.course_name}",
                    action="SKIP",
                    status="FAILED",
                    payload=payload,
                    issue_code="UNMATCHED_COURSE",
                    issue_message="CourseID is not present in the Courses sheet or the AMO course catalogue.",
                )
                _counter(sheet, "SKIP", failed=True)
            elif person_resolves_from_workbook or course_resolves_from_workbook:
                dependencies = []
                if person_resolves_from_workbook:
                    dependencies.append("People")
                if course_resolves_from_workbook:
                    dependencies.append("Courses")
                item = _row(
                    job_id=job.id,
                    sheet="Training",
                    row_number=row_number,
                    entity_type="TRAINING_RECORD",
                    source_key=source_key,
                    label=f"{parsed.person_name or parsed.person_id} · {parsed.course_name}",
                    action="CREATE",
                    status="PENDING_DEPENDENCY",
                    payload=payload,
                    issue_code="WORKBOOK_DEPENDENCY",
                    issue_message=f"Will resolve after accepted {' and '.join(dependencies)} rows are committed.",
                )
                _counter(sheet, "CREATE")
            else:
                current = existing.get((str(user.id), str(course.id), parsed.completion_date))
                action = "CREATE" if current is None else "UPDATE"
                item = _row(
                    job_id=job.id,
                    sheet="Training",
                    row_number=row_number,
                    entity_type="TRAINING_RECORD",
                    source_key=source_key,
                    label=f"{getattr(user, 'full_name', parsed.person_id)} · {course.course_name}",
                    action=action,
                    payload=payload,
                )
                _counter(sheet, action)
            db.add(item)
        except Exception as exc:
            db.add(_row(
                job_id=job.id,
                sheet="Training",
                row_number=row_number,
                entity_type="TRAINING_RECORD",
                source_key=None,
                label=f"{clean(raw.get('PersonName')) or clean(raw.get('PersonID')) or 'Training row'} · {clean(raw.get('CourseName')) or clean(raw.get('CourseID')) or ''}",
                action="SKIP",
                status="FAILED",
                payload=raw,
                issue_code="INVALID_TRAINING_RECORD",
                issue_message=str(exc),
            ))
            _counter(sheet, "SKIP", failed=True)
        _set_job_progress(
            db,
            job,
            stage="MATCHING",
            sheet="Training",
            label=f"{clean(raw.get('PersonName')) or clean(raw.get('PersonID')) or 'Training row'} · {clean(raw.get('CourseID')) or ''}",
            processed_delta=1,
        )


def _preview_role_groups(db: Session, job: TrainingWorkbookImportJob, sheet: TrainingWorkbookImportSheet, rows: list[dict[str, Any]]) -> None:
    existing = {upper(item.code): item for item in db.query(TrainingRoleGroup).filter(TrainingRoleGroup.amo_id == job.amo_id).all()}
    for raw in rows:
        code = upper(raw.get("RoleGroup"))
        description = clean(raw.get("Description"))
        try:
            if not code:
                raise ValueError("RoleGroup is required")
            current = existing.get(code)
            change_list = [] if not current or current.description == description else [{"field": "description", "old": current.description, "new": description}]
            action = "CREATE" if not current else "UPDATE" if change_list else "UNCHANGED"
            db.add(_row(job_id=job.id, sheet="tblRoleGroups", row_number=int(raw["row_number"]), entity_type="ROLE_GROUP", source_key=code, label=description or code, action=action, payload={"code": code, "description": description}, changes=change_list))
            _counter(sheet, action)
        except Exception as exc:
            db.add(_row(job_id=job.id, sheet="tblRoleGroups", row_number=int(raw["row_number"]), entity_type="ROLE_GROUP", source_key=code or None, label=description, action="SKIP", status="FAILED", payload=raw, issue_code="INVALID_ROLE_GROUP", issue_message=str(exc)))
            _counter(sheet, "SKIP", failed=True)
        _set_job_progress(db, job, stage="VALIDATING", sheet="tblRoleGroups", label=code or "Role group", processed_delta=1)


def _preview_person_roles(db: Session, job: TrainingWorkbookImportJob, sheet: TrainingWorkbookImportSheet, rows: list[dict[str, Any]], known_people: set[str], known_groups: set[str]) -> None:
    for raw in rows:
        person_id = upper(raw.get("PersonID"))
        role_group = upper(raw.get("RoleGroup"))
        payload = {"person_id": person_id, "role_group": role_group, "department": upper(raw.get("Department")) or None, "position": upper(raw.get("Position")) or None, "is_active": bool_value(raw.get("ActiveRole"), True), "notes": clean(raw.get("Notes"))}
        issue = None
        if not person_id or not role_group:
            issue = "PersonID and RoleGroup are required"
        elif person_id not in known_people:
            issue = "PersonID is not present in the People sheet or personnel register"
        elif role_group not in known_groups:
            issue = "RoleGroup is not defined in tblRoleGroups"
        if issue:
            db.add(_row(job_id=job.id, sheet="tblPersonRoles", row_number=int(raw["row_number"]), entity_type="PERSON_ROLE", source_key=f"{person_id}:{role_group}", label=f"{person_id} · {role_group}", action="SKIP", status="FAILED", payload=payload, issue_code="INVALID_PERSON_ROLE", issue_message=issue))
            _counter(sheet, "SKIP", failed=True)
        else:
            db.add(_row(job_id=job.id, sheet="tblPersonRoles", row_number=int(raw["row_number"]), entity_type="PERSON_ROLE", source_key=f"{person_id}:{role_group}", label=f"{person_id} · {role_group}", action="UPDATE", payload=payload))
            _counter(sheet, "UPDATE")
        _set_job_progress(db, job, stage="VALIDATING", sheet="tblPersonRoles", label=f"{person_id} · {role_group}", processed_delta=1)


def _preview_matrix(db: Session, job: TrainingWorkbookImportJob, sheet: TrainingWorkbookImportSheet, rows: list[dict[str, Any]], known_courses: set[str], known_groups: set[str]) -> None:
    for raw in rows:
        course_id = upper(raw.get("CourseID"))
        role_group = upper(raw.get("RoleGroup"))
        payload = {"course_id": course_id, "role_group": role_group, "is_required": bool_value(raw.get("Required"), True), "requirement_type": upper(raw.get("RequirementType")) or "GENERAL", "notes": clean(raw.get("Notes"))}
        issue = None
        if not course_id or not role_group:
            issue = "CourseID and RoleGroup are required"
        elif course_id not in known_courses:
            issue = "CourseID is not present in the Courses sheet or course catalogue"
        elif role_group not in known_groups:
            issue = "RoleGroup is not defined in tblRoleGroups"
        if issue:
            db.add(_row(job_id=job.id, sheet="tblCourseMatrix", row_number=int(raw["row_number"]), entity_type="COURSE_ROLE_RULE", source_key=f"{course_id}:{role_group}", label=f"{course_id} · {role_group}", action="SKIP", status="FAILED", payload=payload, issue_code="INVALID_MATRIX_RULE", issue_message=issue))
            _counter(sheet, "SKIP", failed=True)
        else:
            db.add(_row(job_id=job.id, sheet="tblCourseMatrix", row_number=int(raw["row_number"]), entity_type="COURSE_ROLE_RULE", source_key=f"{course_id}:{role_group}", label=f"{course_id} · {role_group}", action="UPDATE", payload=payload))
            _counter(sheet, "UPDATE")
        _set_job_progress(db, job, stage="VALIDATING", sheet="tblCourseMatrix", label=f"{course_id} · {role_group}", processed_delta=1)


def process_workbook_preview(job_id: str) -> None:
    db = SessionLocal()
    try:
        claimed = (
            db.query(TrainingWorkbookImportJob)
            .filter(
                TrainingWorkbookImportJob.id == job_id,
                TrainingWorkbookImportJob.status == "QUEUED",
            )
            .update(
                {
                    TrainingWorkbookImportJob.status: "PARSING",
                    TrainingWorkbookImportJob.stage: "DISCOVERING_SHEETS",
                    TrainingWorkbookImportJob.started_at: utcnow(),
                    TrainingWorkbookImportJob.processed_rows: 0,
                    TrainingWorkbookImportJob.error_message: None,
                    TrainingWorkbookImportJob.completed_at: None,
                    TrainingWorkbookImportJob.updated_at: utcnow(),
                },
                synchronize_session=False,
            )
        )
        db.commit()
        if claimed != 1:
            return
        db.expire_all()
        job = db.get(TrainingWorkbookImportJob, job_id)
        if not job:
            return
        db.query(TrainingWorkbookImportRow).filter(TrainingWorkbookImportRow.job_id == job.id).delete(synchronize_session=False)
        db.query(TrainingWorkbookImportSheet).filter(TrainingWorkbookImportSheet.job_id == job.id).delete(synchronize_session=False)
        db.commit()

        workbook = _load_workbook(job.storage_path)
        rows_by_sheet: dict[str, list[dict[str, Any]]] = {}
        total_rows = 0
        sheets: dict[str, TrainingWorkbookImportSheet] = {}
        for index, name in enumerate(workbook.sheetnames):
            config = WORKBOOK_SHEETS.get(name, {"classification": "UNMAPPED", "destination": "Review and classify", "operational": False})
            rows = _sheet_rows(workbook[name])
            rows_by_sheet[name] = rows
            if name == "People":
                operational_rows = sum(1 for row in rows if upper(row.get("PersonID")) != "TOTAL")
            else:
                operational_rows = len(rows) if config["operational"] else 0
            total_rows += operational_rows
            sheet = TrainingWorkbookImportSheet(
                job_id=job.id,
                sheet_name=name,
                visibility=_sheet_visibility(workbook, name),
                classification=config["classification"],
                portal_destination=config["destination"],
                is_operational=config["operational"],
                display_order=index,
                status="PENDING" if config["operational"] else "MAPPED",
                total_rows=operational_rows if config["operational"] else len(rows),
                message=None if config["operational"] else "This worksheet is represented by a live portal view or configuration and is not copied as duplicate operational data.",
            )
            db.add(sheet)
            sheets[name] = sheet
        job.total_rows = total_rows
        job.summary_json = {
            "sheet_count": len(workbook.sheetnames),
            "operational_sheet_count": sum(1 for item in sheets.values() if item.is_operational),
            "mapped_derived_sheet_count": sum(1 for item in sheets.values() if not item.is_operational),
        }
        db.commit()

        params = _workbook_params(rows_by_sheet.get("Params", []))
        default_frequency_months = _default_frequency_months(params)
        job.summary_json = {
            **(job.summary_json or {}),
            "policy_defaults": {
                "default_frequency_months": default_frequency_months,
            },
        }
        db.add(job)
        db.commit()

        course_rows = rows_by_sheet.get("Courses", [])
        people_rows = [item for item in rows_by_sheet.get("People", []) if upper(item.get("PersonID")) != "TOTAL"]
        workbook_courses = {upper(item.get("CourseID")) for item in course_rows if upper(item.get("CourseID"))}
        workbook_people = {upper(item.get("PersonID")) for item in people_rows if upper(item.get("PersonID"))}
        known_courses = set(workbook_courses)
        known_courses.update(upper(item.course_id) for item in db.query(training_models.TrainingCourse).filter(training_models.TrainingCourse.amo_id == job.amo_id).all())
        known_people = set(workbook_people)
        known_people.update(upper(item.person_id) for item in db.query(account_models.PersonnelProfile).filter(account_models.PersonnelProfile.amo_id == job.amo_id).all())
        role_rows = rows_by_sheet.get("tblRoleGroups", [])
        known_groups = {upper(item.get("RoleGroup")) for item in role_rows if upper(item.get("RoleGroup"))}
        known_groups.update(upper(item.code) for item in db.query(TrainingRoleGroup).filter(TrainingRoleGroup.amo_id == job.amo_id).all())

        processors: list[tuple[str, Callable[[], None]]] = []
        if "Courses" in sheets:
            processors.append(("Courses", lambda: _preview_courses(db, job, sheets["Courses"], course_rows, default_frequency_months=default_frequency_months)))
        if "People" in sheets:
            sheets["People"].total_rows = len(people_rows)
            processors.append(("People", lambda: _preview_people(db, job, sheets["People"], people_rows)))
        if "tblRoleGroups" in sheets:
            processors.append(("tblRoleGroups", lambda: _preview_role_groups(db, job, sheets["tblRoleGroups"], role_rows)))
        if "tblPersonRoles" in sheets:
            processors.append(("tblPersonRoles", lambda: _preview_person_roles(db, job, sheets["tblPersonRoles"], rows_by_sheet.get("tblPersonRoles", []), known_people, known_groups)))
        if "tblCourseMatrix" in sheets:
            processors.append(("tblCourseMatrix", lambda: _preview_matrix(db, job, sheets["tblCourseMatrix"], rows_by_sheet.get("tblCourseMatrix", []), known_courses, known_groups)))
        if "Training" in sheets:
            processors.append(("Training", lambda: _preview_training(db, job, sheets["Training"], rows_by_sheet.get("Training", []), workbook_people=workbook_people, workbook_courses=workbook_courses)))

        for name, processor in processors:
            sheets[name].status = "PROCESSING"
            db.commit()
            processor()
            sheets[name].status = "READY"
            db.add(sheets[name])
            db.commit()

        all_sheets = db.query(TrainingWorkbookImportSheet).filter(TrainingWorkbookImportSheet.job_id == job.id).all()
        job.created_count = sum(item.created_count for item in all_sheets)
        job.updated_count = sum(item.updated_count for item in all_sheets)
        job.unchanged_count = sum(item.unchanged_count for item in all_sheets)
        job.skipped_count = sum(item.skipped_count for item in all_sheets)
        job.failed_count = sum(item.failed_count for item in all_sheets)
        job.review_count = sum(item.review_count for item in all_sheets)
        job.status = "REVIEW_REQUIRED" if job.review_count else "PREVIEW_READY"
        job.stage = "REVIEW"
        job.current_sheet = None
        job.current_record_label = None
        job.preview_completed_at = utcnow()
        job.completed_at = utcnow()
        job.summary_json = {
            **(job.summary_json or {}),
            "workbook_functions": {
                "People": "Personnel, access review and licences",
                "Courses": "Course catalogue",
                "Training": "Training history",
                "tblRoleGroups": "Applicability groups",
                "tblPersonRoles": "Person-to-group assignments",
                "tblCourseMatrix": "Requirement rules",
                "Overdue": "Live overdue queue",
                "Next_Batch": "Smart scheduler",
                "Individual_Lookup": "Individual profile",
                "Course_Audit": "Rectification queue",
            },
        }
        db.add(job)
        db.commit()
    except Exception as exc:
        db.rollback()
        job = db.get(TrainingWorkbookImportJob, job_id)
        if job:
            if str(exc) == "IMPORT_CANCELLED":
                job.status = "CANCELLED"
                job.stage = "CANCELLED"
            else:
                job.status = "FAILED"
                job.stage = "FAILED"
                job.error_message = str(exc)
            job.completed_at = utcnow()
            db.add(job)
            db.commit()
    finally:
        db.close()


def _upsert_course(
    db: Session,
    amo_id: str,
    payload: dict[str, Any],
    actor_user_id: Optional[str],
    *,
    courses_by_code: Optional[dict[str, training_models.TrainingCourse]] = None,
) -> training_models.TrainingCourse:
    code = upper(payload["course_id"])
    course = courses_by_code.get(code) if courses_by_code is not None else None
    if course is None and courses_by_code is None:
        course = db.query(training_models.TrainingCourse).filter(
            training_models.TrainingCourse.amo_id == amo_id,
            training_models.TrainingCourse.course_id == payload["course_id"],
        ).first()
    if course is None:
        course = training_models.TrainingCourse(amo_id=amo_id, course_id=payload["course_id"], created_by_user_id=actor_user_id)
        db.add(course)
    course.course_name = payload["course_name"]
    course.frequency_months = payload.get("frequency_months")
    course.status = payload.get("status") or "One_Off"
    course.category_raw = payload.get("category_raw")
    course.is_mandatory = bool(payload.get("is_mandatory"))
    course.scope = payload.get("scope")
    course.regulatory_reference = payload.get("regulatory_reference")
    course.licence_authority = course.licence_authority or infer_licence_authority(
        payload.get("course_id"), payload.get("course_name")
    )
    course.is_active = payload.get("is_active", True)
    course.updated_by_user_id = actor_user_id
    db.flush()
    if courses_by_code is not None:
        courses_by_code[code] = course
    return course


def _build_personnel_commit_indexes(db: Session, amo_id: str) -> PersonnelCommitIndexes:
    profiles = db.query(account_models.PersonnelProfile).filter(
        account_models.PersonnelProfile.amo_id == amo_id,
    ).all()
    users = db.query(account_models.User).filter(account_models.User.amo_id == amo_id).all()
    departments = db.query(account_models.Department).filter(
        account_models.Department.amo_id == amo_id,
        account_models.Department.is_active.is_(True),
    ).all()
    licences = db.query(PersonnelLicence).filter(PersonnelLicence.amo_id == amo_id).all()

    licence_index: dict[tuple[str, str], list[PersonnelLicence]] = {}
    for licence in licences:
        licence_index.setdefault((str(licence.personnel_profile_id), upper(licence.authority)), []).append(licence)

    department_ids: dict[str, str] = {}
    for department in departments:
        if department.code:
            department_ids[str(department.code).strip().lower()] = str(department.id)
        if department.name:
            department_ids[str(department.name).strip().lower()] = str(department.id)

    return PersonnelCommitIndexes(
        profiles_by_person={upper(item.person_id): item for item in profiles},
        profiles_by_email={(item.email or "").strip().lower(): item for item in profiles if item.email},
        users_by_staff={upper(item.staff_code): item for item in users},
        users_by_email={(item.email or "").strip().lower(): item for item in users if item.email},
        users_by_id={str(item.id): item for item in users},
        department_ids_by_token=department_ids,
        licences_by_profile_authority=licence_index,
    )


def _department_id(
    db: Session,
    amo_id: str,
    value: Optional[str],
    *,
    indexes: Optional[PersonnelCommitIndexes] = None,
) -> Optional[str]:
    if not value:
        return None
    normalized = value.strip().lower()
    if indexes is not None:
        return indexes.department_ids_by_token.get(normalized)
    items = db.query(account_models.Department).filter(account_models.Department.amo_id == amo_id, account_models.Department.is_active.is_(True)).all()
    for item in items:
        if str(item.code or "").strip().lower() == normalized or str(item.name or "").strip().lower() == normalized:
            return str(item.id)
    return None


def _upsert_licence(
    db: Session,
    *,
    job: TrainingWorkbookImportJob,
    profile: account_models.PersonnelProfile,
    user: Optional[account_models.User],
    authority: str,
    country: str,
    number: Optional[str],
    category: Optional[str],
    category_source: Optional[str],
    payload: dict[str, Any],
    source_row: int,
    primary: bool,
    indexes: Optional[PersonnelCommitIndexes] = None,
) -> None:
    licence_key = (str(profile.id), upper(authority))
    if indexes is not None:
        authority_licences = indexes.licences_by_profile_authority.setdefault(licence_key, [])
    else:
        authority_licences = db.query(PersonnelLicence).filter(
            PersonnelLicence.amo_id == job.amo_id,
            PersonnelLicence.personnel_profile_id == profile.id,
            PersonnelLicence.authority == authority,
        ).all()
    licence = next((item for item in authority_licences if item.licence_number == number), None) if number else None
    for previous in authority_licences:
        if previous is licence:
            continue
        previous.status = _licence_reconciliation_status(previous.licence_number, number)
        previous.is_primary = False
    if not number:
        if primary and user:
            user.regulatory_authority = None
            user.licence_number = None
            user.licence_state_or_country = None
        return
    if licence is None:
        licence = PersonnelLicence(
            amo_id=job.amo_id,
            personnel_profile_id=profile.id,
            authority=authority,
            licence_number=number,
        )
        db.add(licence)
        authority_licences.append(licence)
    licence.user_id = user.id if user else None
    licence.country = country
    licence.category_code = category
    licence.category_source = category_source
    licence.internal_stamp_no = payload.get("internal_stamp_no")
    licence.initial_authorization_date = date_value(payload.get("initial_authorization_date"))
    licence.status = "ACTIVE" if str(payload.get("status") or "Active").lower() == "active" else "DORMANT"
    licence.is_primary = primary
    licence.source_job_id = job.id
    licence.source_row = source_row


def _upsert_person(
    db: Session,
    job: TrainingWorkbookImportJob,
    row: TrainingWorkbookImportRow,
    *,
    indexes: Optional[PersonnelCommitIndexes] = None,
    inactive_password_hash: Optional[str] = None,
) -> PersonCommitResult:
    payload = dict(row.payload_json or {})
    person_id = upper(payload.get("person_id"))
    decision = (row.decision or "").upper()
    if decision == "SKIP":
        return PersonCommitResult(entity_id=None, action="SKIP")
    if decision == "USE_IMPORTED_EMAIL":
        raise ValueError("Imported email conflicts must be reconciled outside the import; choose KEEP_EXISTING_EMAIL or SKIP.")

    email_key = str(payload.get("email") or "").strip().lower()
    if indexes is not None:
        profile_by_person = indexes.profiles_by_person.get(person_id)
        profile_by_email = indexes.profiles_by_email.get(email_key) if email_key else None
        existing_staff_user = indexes.users_by_staff.get(person_id)
    else:
        profile_by_person = db.query(account_models.PersonnelProfile).filter(
            account_models.PersonnelProfile.amo_id == job.amo_id,
            account_models.PersonnelProfile.person_id == person_id,
        ).first()
        profile_by_email = None
        if email_key:
            profile_by_email = db.query(account_models.PersonnelProfile).filter(
                account_models.PersonnelProfile.amo_id == job.amo_id,
                func.lower(account_models.PersonnelProfile.email) == email_key,
            ).first()
        existing_staff_user = db.query(account_models.User).filter(
            account_models.User.amo_id == job.amo_id,
            account_models.User.staff_code == person_id,
        ).first()
    profile = profile_by_person if decision == "KEEP_EXISTING_EMAIL" else profile_by_person or profile_by_email
    is_new = profile is None
    if profile is None:
        profile = account_models.PersonnelProfile(
            amo_id=job.amo_id,
            person_id=person_id,
            first_name=payload["first_name"],
            last_name=payload["last_name"],
        )
        db.add(profile)

    imported_email = payload.get("email")
    if decision == "KEEP_EXISTING_EMAIL":
        selected_email = (profile_by_person.email if profile_by_person else None) or (existing_staff_user.email if existing_staff_user else None)
    else:
        selected_email = imported_email or profile.email

    selected_email_key = str(selected_email or "").strip().lower()
    if selected_email_key and decision != "KEEP_EXISTING_EMAIL":
        claimed_profile = indexes.profiles_by_email.get(selected_email_key) if indexes is not None else profile_by_email
        if claimed_profile is not None and claimed_profile is not profile:
            raise PersonnelIdentityChanged(row.id, "The imported email is now assigned to another personnel profile. Review this People row again.")

    profile.person_id = person_id
    profile.first_name = payload["first_name"]
    profile.last_name = payload["last_name"]
    profile.full_name = payload.get("full_name")
    profile.national_id = payload.get("national_id")
    profile.amel_no = payload.get("kamel_no")
    profile.internal_certification_stamp_no = payload.get("internal_stamp_no")
    profile.initial_authorization_date = date_value(payload.get("initial_authorization_date"))
    profile.department = payload.get("department")
    profile.position_title = payload.get("position_title")
    profile.phone_number = payload.get("phone_number")
    profile.secondary_phone = payload.get("secondary_phone")
    profile.email = selected_email
    imported_hire_date = date_value(payload.get("hire_date"))
    previous_hire_date = profile.hire_date
    if imported_hire_date and imported_hire_date != previous_hire_date:
        profile.hire_date = imported_hire_date
        audit_services.log_event(
            db,
            amo_id=job.amo_id,
            actor_user_id=job.actor_user_id,
            entity_type="PersonnelProfile",
            entity_id=str(profile.id),
            action="HIRE_DATE_IMPORT_APPLIED",
            before={"hire_date": previous_hire_date.isoformat() if previous_hire_date else None},
            after={"hire_date": imported_hire_date.isoformat()},
            metadata={
                "module": "training",
                "imported_hire_date": imported_hire_date.isoformat(),
                "reason": "Imported personnel hire date is the authoritative Workforce start",
            },
            critical=True,
        )
    profile.employment_status = payload.get("employment_status")
    profile.status = payload.get("status") or "Active"
    profile.date_of_birth = date_value(payload.get("date_of_birth"))
    profile.birth_place = payload.get("birth_place")
    db.flush()

    existing_profile_user = (
        indexes.users_by_id.get(str(profile.user_id)) if indexes is not None and profile.user_id else
        db.get(account_models.User, profile.user_id) if profile.user_id else None
    )
    existing_email_user = None
    if selected_email and decision != "KEEP_EXISTING_EMAIL":
        if indexes is not None:
            existing_email_user = indexes.users_by_email.get(selected_email_key)
        else:
            existing_email_user = db.query(account_models.User).filter(
                account_models.User.amo_id == job.amo_id,
                func.lower(account_models.User.email) == selected_email_key,
            ).first()

    if decision == "PROFILE_ONLY":
        if existing_profile_user or existing_staff_user or existing_email_user:
            raise ValueError("A portal account now exists for this person. Re-run preview and choose LINK_EXISTING_ACCOUNT or another reviewed action.")
        user = account_models.User(
            id=generate_user_id(),
            amo_id=job.amo_id,
            department_id=_department_id(db, job.amo_id, payload.get("department"), indexes=indexes),
            staff_code=person_id,
            email=f"{person_id.lower()}@personnel.invalid",
            first_name=payload["first_name"],
            last_name=payload["last_name"],
            full_name=payload.get("full_name") or f"{payload['first_name']} {payload['last_name']}",
            role=_role_from_position(payload.get("position_title")),
            position_title=payload.get("position_title"),
            phone=payload.get("phone_number"),
            secondary_phone=payload.get("secondary_phone"),
            hashed_password=inactive_password_hash or get_password_hash(secrets.token_urlsafe(48)),
            is_active=False,
            is_amo_admin=False,
            is_auditor=False,
            must_change_password=True,
            approved_by_user_id=None,
            approved_at=None,
            approval_notes="Personnel-only identity imported for licence and training history; portal access disabled.",
        )
        db.add(user)
        try:
            db.flush()
        except IntegrityError as exc:
            raise PersonnelIdentityChanged(row.id, "A portal account was created after preview. Review this People row again.") from exc
        profile.user_id = user.id
    else:
        user = existing_profile_user or existing_staff_user or existing_email_user
        if decision == "CREATE_ACCOUNT" and user is not None:
            raise PersonnelIdentityChanged(row.id, "A portal account was created after preview. Review this People row again.")
        if user is None and decision == "CREATE_ACCOUNT":
            if not selected_email:
                raise ValueError("A portal-access candidate requires a valid email address")
            user = account_models.User(
                id=generate_user_id(),
                amo_id=job.amo_id,
                department_id=_department_id(db, job.amo_id, payload.get("department"), indexes=indexes),
                staff_code=person_id,
                email=selected_email,
                first_name=payload["first_name"],
                last_name=payload["last_name"],
                full_name=payload.get("full_name") or f"{payload['first_name']} {payload['last_name']}",
                role=_role_from_position(payload.get("position_title")),
                position_title=payload.get("position_title"),
                phone=payload.get("phone_number"),
                secondary_phone=payload.get("secondary_phone"),
                hashed_password=inactive_password_hash or get_password_hash(secrets.token_urlsafe(48)),
                is_active=False,
                is_amo_admin=False,
                is_auditor=False,
                must_change_password=True,
                approved_by_user_id=None,
                approved_at=None,
                approval_notes="Imported from Training Tracker; pending administrator approval and onboarding.",
            )
            db.add(user)
            try:
                db.flush()
            except IntegrityError as exc:
                raise PersonnelIdentityChanged(row.id, "A portal account was created after preview. Review this People row again.") from exc
        elif user is None and decision == "LINK_EXISTING_ACCOUNT":
            raise ValueError("The account selected for linking no longer exists. Re-run the workbook preview.")
        elif user is not None:
            user.department_id = _department_id(db, job.amo_id, payload.get("department"), indexes=indexes) or user.department_id
            user.first_name = payload["first_name"]
            user.last_name = payload["last_name"]
            user.full_name = payload.get("full_name") or user.full_name
            user.position_title = payload.get("position_title")
            user.phone = payload.get("phone_number")
            user.secondary_phone = payload.get("secondary_phone")
            if str(payload.get("status") or "Active").lower() != "active":
                user.is_active = False
            if selected_email and decision != "KEEP_EXISTING_EMAIL":
                user.email = selected_email
        if user is not None:
            profile.user_id = user.id

    if user and payload.get("kamel_no"):
        user.regulatory_authority = account_models.RegulatoryAuthority.KCAA
        user.licence_number = payload.get("kamel_no")
        user.licence_state_or_country = "Kenya"

    if user and profile.hire_date:
        from ..workforce import services as workforce_services

        workforce_services.sync_contract_start_from_hire_date(
            db,
            amo_id=job.amo_id,
            user_id=str(user.id),
            hire_date=profile.hire_date,
            actor_user_id=job.actor_user_id,
            source="TRAINING_WORKBOOK_PEOPLE_HIREDATE",
        )

    category = payload.get("category_reg_2018") or payload.get("category_reg_2013")
    category_source = "Reg. 2018" if payload.get("category_reg_2018") else "Reg. 2013" if payload.get("category_reg_2013") else None
    _upsert_licence(db, job=job, profile=profile, user=user, authority="KCAA", country="Kenya", number=payload.get("kamel_no"), category=category, category_source=category_source, payload=payload, source_row=row.source_row, primary=True, indexes=indexes)
    _upsert_licence(db, job=job, profile=profile, user=user, authority="ETHIOPIAN_CAA", country="Ethiopia", number=payload.get("e_amel"), category=None, category_source=None, payload=payload, source_row=row.source_row, primary=False, indexes=indexes)
    _upsert_licence(db, job=job, profile=profile, user=user, authority="GHANA_CAA", country="Ghana", number=payload.get("g_amel"), category=None, category_source=None, payload=payload, source_row=row.source_row, primary=False, indexes=indexes)
    db.flush()
    if indexes is not None:
        indexes.profiles_by_person[person_id] = profile
        for key, indexed_profile in list(indexes.profiles_by_email.items()):
            if indexed_profile is profile and key != selected_email_key:
                indexes.profiles_by_email.pop(key, None)
        if selected_email_key:
            indexes.profiles_by_email[selected_email_key] = profile
        if user is not None:
            indexes.users_by_id[str(user.id)] = user
            indexes.users_by_staff[person_id] = user
            current_user_email_key = str(user.email or "").strip().lower()
            for key, indexed_user in list(indexes.users_by_email.items()):
                if indexed_user is user and key != current_user_email_key:
                    indexes.users_by_email.pop(key, None)
            if user.email:
                indexes.users_by_email[current_user_email_key] = user
    return PersonCommitResult(
        entity_id=str(user.id if user else profile.id),
        action="CREATE" if is_new else "UPDATE",
        profile_created=is_new,
        portal_account_created=decision == "CREATE_ACCOUNT" and user is not None,
        non_login_identity_created=decision == "PROFILE_ONLY" and user is not None,
    )


def _progress_callback(job_id: str, base_processed: int, attempt_token: str) -> Callable[[int, int, str], None]:
    last_published = 0

    def callback(processed: int, total: int, label: str) -> None:
        nonlocal last_published
        if processed < total and processed - last_published < COMMIT_PROGRESS_BATCH:
            return
        last_published = processed
        progress_db = SessionLocal()
        try:
            _commit_progress(
                progress_db,
                job_id,
                attempt_token,
                base_processed + processed,
                "COMMITTING_TRAINING",
                "Training",
                label,
            )
        finally:
            progress_db.close()
    return callback


def _materialize_mandatory_catalogue_requirements(db: Session, job: TrainingWorkbookImportJob) -> None:
    """Preserve catalogue fallback before the tenant enters explicit requirement mode."""
    existing = db.query(training_models.TrainingRequirement.id).filter(
        training_models.TrainingRequirement.amo_id == job.amo_id,
        training_models.TrainingRequirement.is_active.is_(True),
    ).first()
    if existing:
        return
    courses = db.query(training_models.TrainingCourse).filter(
        training_models.TrainingCourse.amo_id == job.amo_id,
        training_models.TrainingCourse.is_active.is_(True),
        training_models.TrainingCourse.is_mandatory.is_(True),
    ).all()
    for course in courses:
        db.add(training_models.TrainingRequirement(
            amo_id=job.amo_id,
            course_id=course.id,
            scope=training_models.TrainingRequirementScope.ALL,
            is_mandatory=True,
            is_active=True,
            created_by_user_id=job.actor_user_id,
        ))
    db.flush()


def _refresh_identity_review_options(
    db: Session,
    job: TrainingWorkbookImportJob,
    row: TrainingWorkbookImportRow,
) -> list[str]:
    """Rebuild valid choices when identity state changes after preview."""
    payload = dict(row.payload_json or {})
    person_id = upper(payload.get("person_id"))
    email = str(payload.get("email") or "").strip().lower()
    profile_by_person = db.query(account_models.PersonnelProfile).filter(
        account_models.PersonnelProfile.amo_id == job.amo_id,
        account_models.PersonnelProfile.person_id == person_id,
    ).first()
    profile_by_email = None
    if email:
        profile_by_email = db.query(account_models.PersonnelProfile).filter(
            account_models.PersonnelProfile.amo_id == job.amo_id,
            func.lower(account_models.PersonnelProfile.email) == email,
        ).first()
    identity_filters = [account_models.User.staff_code == person_id]
    if email:
        identity_filters.append(func.lower(account_models.User.email) == email)
    user = db.query(account_models.User).filter(
        account_models.User.amo_id == job.amo_id,
        or_(*identity_filters),
    ).first()
    if user:
        return ["LINK_EXISTING_ACCOUNT", "SKIP"]
    if profile_by_person:
        if profile_by_email is not None and profile_by_email.id != profile_by_person.id:
            return ["KEEP_EXISTING_EMAIL", "SKIP"]
        return (["CREATE_ACCOUNT", "PROFILE_ONLY", "SKIP"] if email else ["PROFILE_ONLY", "SKIP"])
    return list(row.decision_options or ["SKIP"])


def commit_workbook_import(
    job_id: str,
    *,
    force_reimport: bool = False,
    attempt_token: Optional[str] = None,
) -> None:
    progress_db = SessionLocal()
    work_db = SessionLocal()
    commit_started = perf_counter()
    total_processed = 0
    accounts_created = 0
    profiles_created = 0
    non_login_identities_created = 0
    try:
        job = progress_db.get(TrainingWorkbookImportJob, job_id)
        if not job:
            return
        expected_token = attempt_token or _commit_attempt_token(job)
        if not expected_token or _commit_attempt_token(job) != expected_token:
            return
        claimed = (
            progress_db.query(TrainingWorkbookImportJob)
            .filter(
                TrainingWorkbookImportJob.id == job_id,
                TrainingWorkbookImportJob.status == "QUEUED_COMMIT",
            )
            .update(
                {
                    TrainingWorkbookImportJob.status: "COMMITTING",
                    TrainingWorkbookImportJob.stage: "COMMITTING_COURSES",
                    TrainingWorkbookImportJob.processed_rows: 0,
                    TrainingWorkbookImportJob.current_sheet: "Courses",
                    TrainingWorkbookImportJob.current_record_label: None,
                    TrainingWorkbookImportJob.error_message: None,
                    TrainingWorkbookImportJob.completed_at: None,
                    TrainingWorkbookImportJob.updated_at: utcnow(),
                },
                synchronize_session=False,
            )
        )
        progress_db.commit()
        if claimed != 1:
            return
        progress_db.expire_all()
        job = progress_db.get(TrainingWorkbookImportJob, job_id)
        if not job or _commit_attempt_token(job) != expected_token:
            return
        duplicate = progress_db.query(TrainingWorkbookImportJob).filter(
            TrainingWorkbookImportJob.amo_id == job.amo_id,
            TrainingWorkbookImportJob.file_sha256 == job.file_sha256,
            TrainingWorkbookImportJob.status == "COMPLETED",
            TrainingWorkbookImportJob.committed_at.isnot(None),
            TrainingWorkbookImportJob.id != job.id,
        ).order_by(TrainingWorkbookImportJob.committed_at.desc()).first()
        if duplicate and not force_reimport:
            raise ValueError(f"This workbook was already committed by import {duplicate.id}. Use force re-import only after reviewing the reconciliation.")

        unresolved = progress_db.query(TrainingWorkbookImportRow).filter(
            TrainingWorkbookImportRow.job_id == job.id,
            TrainingWorkbookImportRow.decision_required.is_(True),
            TrainingWorkbookImportRow.decision.is_(None),
        ).count()
        if unresolved:
            raise ValueError(f"{unresolved} review decision(s) are still required before commit.")

        job.summary_json = {
            **(job.summary_json or {}),
            "commit_started_at": utcnow().isoformat(),
            "last_commit_attempt": {"status": "RUNNING", "processed_rows": 0},
        }
        progress_db.add(job)
        progress_db.commit()

        rows = progress_db.query(TrainingWorkbookImportRow).filter(TrainingWorkbookImportRow.job_id == job.id).order_by(TrainingWorkbookImportRow.sheet_name, TrainingWorkbookImportRow.source_row).all()
        rows_by_sheet: dict[str, list[TrainingWorkbookImportRow]] = {}
        for item in rows:
            rows_by_sheet.setdefault(item.sheet_name, []).append(item)
            progress_db.expunge(item)

        with work_db.begin():
            # Course catalogue first so matrix and history can resolve CourseID.
            courses = {
                upper(item.course_id): item
                for item in work_db.query(training_models.TrainingCourse).filter(
                    training_models.TrainingCourse.amo_id == job.amo_id,
                ).all()
            }
            for item in rows_by_sheet.get("Courses", []):
                if item.status == "FAILED" or item.proposed_action == "SKIP":
                    total_processed += 1
                    continue
                entity = _upsert_course(
                    work_db,
                    job.amo_id,
                    dict(item.payload_json or {}),
                    job.actor_user_id,
                    courses_by_code=courses,
                )
                item.committed_entity_id = entity.id
                total_processed += 1
                if total_processed % COMMIT_PROGRESS_BATCH == 0:
                    _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_COURSES", "Courses", item.display_label)

            # Personnel + explicit access decisions + multi-authority licences.
            _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_PEOPLE", "People", None)
            personnel_indexes = _build_personnel_commit_indexes(work_db, job.amo_id)
            needs_inactive_identity = any(
                item.status != "FAILED"
                and (item.decision or "").upper() in {"CREATE_ACCOUNT", "PROFILE_ONLY"}
                for item in rows_by_sheet.get("People", [])
            )
            inactive_password_hash = get_password_hash(secrets.token_urlsafe(48)) if needs_inactive_identity else None
            for item in rows_by_sheet.get("People", []):
                if item.status == "FAILED":
                    total_processed += 1
                    continue
                try:
                    # A row savepoint ensures an identity race cannot poison the
                    # surrounding atomic import transaction. Any failure exits
                    # the outer context before another SQL command is issued.
                    with work_db.begin_nested():
                        result = _upsert_person(
                            work_db,
                            job,
                            item,
                            indexes=personnel_indexes,
                            inactive_password_hash=inactive_password_hash,
                        )
                    item.committed_entity_id = result.entity_id
                    profiles_created += int(result.profile_created)
                    accounts_created += int(result.portal_account_created)
                    non_login_identities_created += int(result.non_login_identity_created)
                    if result.action == "SKIP":
                        item.status = "SKIPPED"
                except PersonnelIdentityChanged:
                    raise
                except IntegrityError as exc:
                    raise PersonnelIdentityChanged(
                        item.id,
                        "A personnel profile or portal account changed after review. Review this People row again.",
                    ) from exc
                except (OperationalError, DBAPIError):
                    raise
                except Exception as exc:
                    raise WorkbookRowCommitError(
                        item.id,
                        item.sheet_name,
                        item.source_row,
                        str(exc),
                    ) from exc
                total_processed += 1
                if total_processed % COMMIT_PROGRESS_BATCH == 0:
                    _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_PEOPLE", "People", item.display_label)

            # Applicability groups.
            _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_ROLE_GROUPS", "tblRoleGroups", None)
            groups: dict[str, TrainingRoleGroup] = {
                upper(item.code): item
                for item in work_db.query(TrainingRoleGroup).filter(
                    TrainingRoleGroup.amo_id == job.amo_id,
                ).all()
            }
            for item in rows_by_sheet.get("tblRoleGroups", []):
                if item.status == "FAILED":
                    total_processed += 1
                    continue
                payload = dict(item.payload_json or {})
                code = upper(payload.get("code"))
                group = groups.get(code)
                if group is None:
                    group = TrainingRoleGroup(amo_id=job.amo_id, code=code)
                    work_db.add(group)
                group.description = payload.get("description")
                group.is_active = True
                group.source_job_id = job.id
                work_db.flush()
                groups[code] = group
                item.committed_entity_id = group.id
                total_processed += 1
                if total_processed % COMMIT_PROGRESS_BATCH == 0:
                    _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_ROLE_GROUPS", "tblRoleGroups", item.display_label)

            _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_PERSON_ROLES", "tblPersonRoles", None)
            profiles = personnel_indexes.profiles_by_person
            users = personnel_indexes.users_by_staff
            assignments = {
                (upper(item.person_id), str(item.role_group_id)): item
                for item in work_db.query(TrainingPersonRole).filter(
                    TrainingPersonRole.amo_id == job.amo_id,
                ).all()
            }
            for item in rows_by_sheet.get("tblPersonRoles", []):
                if item.status == "FAILED":
                    total_processed += 1
                    continue
                payload = dict(item.payload_json or {})
                profile = profiles.get(upper(payload.get("person_id")))
                group = groups.get(upper(payload.get("role_group")))
                if not profile or not group:
                    item.status = "FAILED"
                    item.issue_message = "Person or role group was not available at commit."
                    total_processed += 1
                    continue
                assignment_key = (upper(profile.person_id), str(group.id))
                assignment = assignments.get(assignment_key)
                if assignment is None:
                    assignment = TrainingPersonRole(amo_id=job.amo_id, person_id=profile.person_id, role_group_id=group.id)
                    work_db.add(assignment)
                    assignments[assignment_key] = assignment
                assignment.personnel_profile_id = profile.id
                assignment.user_id = (users.get(upper(profile.person_id)).id if users.get(upper(profile.person_id)) else profile.user_id)
                assignment.department = payload.get("department")
                assignment.position = payload.get("position")
                assignment.notes = payload.get("notes")
                assignment.is_active = bool(payload.get("is_active", True))
                assignment.source_job_id = job.id
                work_db.flush()
                item.committed_entity_id = assignment.id
                total_processed += 1
                if total_processed % COMMIT_PROGRESS_BATCH == 0:
                    _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_PERSON_ROLES", "tblPersonRoles", item.display_label)

            _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_COURSE_MATRIX", "tblCourseMatrix", None)
            rules = {
                (str(item.course_id), str(item.role_group_id), upper(item.requirement_type) or "GENERAL"): item
                for item in work_db.query(TrainingCourseRoleRule).filter(
                    TrainingCourseRoleRule.amo_id == job.amo_id,
                ).all()
            }
            for item in rows_by_sheet.get("tblCourseMatrix", []):
                if item.status == "FAILED":
                    total_processed += 1
                    continue
                payload = dict(item.payload_json or {})
                course = courses.get(upper(payload.get("course_id")))
                group = groups.get(upper(payload.get("role_group")))
                if not course or not group:
                    item.status = "FAILED"
                    item.issue_message = "Course or role group was not available at commit."
                    total_processed += 1
                    continue
                requirement_type = upper(payload.get("requirement_type")) or "GENERAL"
                rule_key = (str(course.id), str(group.id), requirement_type)
                rule = rules.get(rule_key)
                if rule is None:
                    rule = TrainingCourseRoleRule(amo_id=job.amo_id, course_id=course.id, role_group_id=group.id, requirement_type=requirement_type)
                    work_db.add(rule)
                    rules[rule_key] = rule
                rule.is_required = bool(payload.get("is_required", True))
                rule.notes = payload.get("notes")
                rule.is_active = True
                rule.source_job_id = job.id
                work_db.flush()
                item.committed_entity_id = rule.id
                # Keep the canonical ALL requirement in exact sync for existing
                # consumers, including deactivation when a later matrix makes
                # the course optional.
                if group.code == "ALL":
                    if bool(payload.get("is_required", True)):
                        _materialize_mandatory_catalogue_requirements(work_db, job)
                    canonical = work_db.query(training_models.TrainingRequirement).filter(
                        training_models.TrainingRequirement.amo_id == job.amo_id,
                        training_models.TrainingRequirement.course_id == course.id,
                        training_models.TrainingRequirement.scope == training_models.TrainingRequirementScope.ALL,
                    ).first()
                    any_required = work_db.query(TrainingCourseRoleRule.id).filter(
                        TrainingCourseRoleRule.amo_id == job.amo_id,
                        TrainingCourseRoleRule.course_id == course.id,
                        TrainingCourseRoleRule.role_group_id == group.id,
                        TrainingCourseRoleRule.is_active.is_(True),
                        TrainingCourseRoleRule.is_required.is_(True),
                    ).first() is not None
                    if canonical is None and any_required:
                        canonical = training_models.TrainingRequirement(
                            amo_id=job.amo_id,
                            course_id=course.id,
                            scope=training_models.TrainingRequirementScope.ALL,
                            is_mandatory=True,
                            is_active=True,
                            created_by_user_id=job.actor_user_id,
                        )
                        work_db.add(canonical)
                    elif canonical is not None:
                        canonical.is_mandatory = any_required
                        canonical.is_active = any_required
                total_processed += 1
                if total_processed % COMMIT_PROGRESS_BATCH == 0:
                    _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_COURSE_MATRIX", "tblCourseMatrix", item.display_label)

            _commit_progress(progress_db, job.id, expected_token, total_processed, "COMMITTING_TRAINING", "Training", None)
            training_payloads = []
            training_rows = []
            for item in rows_by_sheet.get("Training", []):
                if item.status == "FAILED" or (item.decision or "").upper() == "SKIP":
                    total_processed += 1
                    continue
                training_payloads.append({"row_number": item.source_row, **dict(item.payload_json or {})})
                training_rows.append(item)
            if training_payloads:
                result = records_import.import_training_records_rows(
                    work_db,
                    amo_id=job.amo_id,
                    rows=training_payloads,
                    dry_run=False,
                    actor_user_id=job.actor_user_id,
                    manage_transaction=False,
                    progress_callback=_progress_callback(job.id, total_processed, expected_token),
                )
                preview_by_row = {entry.row_number: entry for entry in result.preview_rows}
                for item in training_rows:
                    preview = preview_by_row.get(item.source_row)
                    if preview:
                        item.committed_entity_id = preview.existing_record_id
                        item.proposed_action = preview.action
                        if preview.action == "SKIP":
                            item.status = "FAILED"
                            item.issue_message = preview.reason
                total_processed += len(training_payloads)

            # Fence this atomic transaction immediately before commit. If a
            # stale worker was superseded while PostgreSQL was unavailable,
            # only the currently leased attempt may publish operational data.
            _commit_progress(
                progress_db,
                job.id,
                expected_token,
                total_processed,
                "FINALIZING_COMMIT",
                "Training",
                None,
            )

            audit_services.log_event(
                work_db,
                amo_id=job.amo_id,
                actor_user_id=job.actor_user_id,
                entity_type="training.workbook_import",
                entity_id=job.id,
                action="COMMIT",
                after={"filename": job.filename, "sha256": job.file_sha256, "rows": total_processed},
                metadata={"module": "training", "source": "Training_Tracker workbook"},
            )
            _require_commit_lease(progress_db, job.id, expected_token)

        # Persist all row outcomes as one executemany operation. The previous
        # per-row get/update loop issued thousands of round trips on realistic
        # workbooks and made the final percentage appear frozen.
        _commit_progress(
            progress_db,
            job.id,
            expected_token,
            total_processed,
            "FINALIZING_IMPORT",
            "Reconciliation",
            None,
        )
        progress_db.bulk_update_mappings(
            TrainingWorkbookImportRow,
            [
                {
                    "id": item.id,
                    "status": item.status if item.status in {"FAILED", "SKIPPED"} else "COMMITTED",
                    "issue_code": item.issue_code,
                    "issue_message": item.issue_message,
                    "committed_entity_id": item.committed_entity_id,
                    "proposed_action": item.proposed_action,
                    "updated_at": utcnow(),
                }
                for item in rows
            ],
        )
        progress_db.commit()

        # Keep the current-year training plan in step with the newly committed
        # personnel history. This is intentionally a separate transaction: an
        # optional planning sync must never invalidate an otherwise successful
        # governed workbook import.
        plan_sync: dict[str, Any]
        try:
            from sqlalchemy import inspect
            from . import operating_service

            if inspect(work_db.get_bind()).has_table("training_plans"):
                with work_db.begin():
                    actor = work_db.query(account_models.User).filter(
                        account_models.User.id == job.actor_user_id,
                        account_models.User.amo_id == job.amo_id,
                    ).first()
                    if actor is None:
                        plan_sync = {"action": "SKIPPED", "message": "Import actor is not available for plan provenance."}
                    else:
                        plan_sync = operating_service.sync_current_plan_from_records(work_db, actor=actor)
            else:
                plan_sync = {"action": "SKIPPED", "message": "Training planning migration is not installed yet."}
        except Exception as plan_exc:
            work_db.rollback()
            plan_sync = {"action": "FAILED", "message": str(plan_exc)}

        job = _require_commit_lease(progress_db, job.id, expected_token)
        committed_rows = progress_db.query(TrainingWorkbookImportRow).filter(TrainingWorkbookImportRow.job_id == job.id).all()
        job.created_count = sum(1 for item in committed_rows if item.status == "COMMITTED" and item.proposed_action == "CREATE")
        job.updated_count = sum(1 for item in committed_rows if item.status == "COMMITTED" and item.proposed_action == "UPDATE")
        job.unchanged_count = sum(1 for item in committed_rows if item.status == "COMMITTED" and item.proposed_action == "UNCHANGED")
        job.skipped_count = sum(1 for item in committed_rows if item.status == "SKIPPED")
        job.failed_count = sum(1 for item in committed_rows if item.status == "FAILED")
        job.review_count = 0
        job.processed_rows = job.total_rows
        job.status = "COMPLETED"
        job.stage = "COMPLETED"
        job.current_sheet = None
        job.current_record_label = None
        job.committed_by_user_id = job.actor_user_id
        job.committed_at = utcnow()
        job.completed_at = utcnow()
        job.summary_json = {
            **(job.summary_json or {}),
            "active_commit_token": None,
            "training_plan_sync": plan_sync,
            "commit_stats": {
                "portal_accounts_created": accounts_created,
                "personnel_profiles_created": profiles_created,
                "non_login_identities_created": non_login_identities_created,
                "processed_rows": total_processed,
                "elapsed_ms": int((perf_counter() - commit_started) * 1000),
            },
            "last_commit_attempt": {
                "status": "COMPLETED",
                "processed_rows": total_processed,
                "elapsed_ms": int((perf_counter() - commit_started) * 1000),
            },
        }
        progress_db.add(job)
        progress_db.commit()
    except WorkbookCommitLeaseLost:
        for session in (work_db, progress_db):
            try:
                session.rollback()
            except Exception:
                pass
        return
    except Exception as exc:
        for session in (work_db, progress_db):
            try:
                session.rollback()
            except Exception:
                pass
        # A PostgreSQL restart invalidates both sessions. Leave the durable job
        # active so the status endpoint can renew its lease and rerun the whole
        # atomic commit after connectivity returns.
        if _is_transient_database_error(exc):
            return
        job = progress_db.get(TrainingWorkbookImportJob, job_id)
        if job:
            if isinstance(exc, PersonnelIdentityChanged):
                row = progress_db.get(TrainingWorkbookImportRow, exc.row_id)
                if row:
                    row.status = "REVIEW"
                    row.decision = None
                    row.decision_required = True
                    row.decision_options = _refresh_identity_review_options(progress_db, job, row)
                    row.issue_code = "IDENTITY_CHANGED"
                    row.issue_message = str(exc)
                    progress_db.add(row)
                job.status = "REVIEW_REQUIRED"
                job.stage = "REVIEW"
                job.error_message = str(exc)
            elif isinstance(exc, WorkbookRowCommitError):
                row = progress_db.get(TrainingWorkbookImportRow, exc.row_id)
                if row:
                    row.issue_code = "COMMIT_RETRY_REQUIRED"
                    row.issue_message = str(exc)
                    progress_db.add(row)
            if str(exc) == "IMPORT_CANCELLED":
                job.status = "CANCELLED"
                job.stage = "CANCELLED"
            elif not isinstance(exc, PersonnelIdentityChanged):
                job.status = "FAILED"
                job.stage = "FAILED"
                job.error_message = str(exc)
            job.summary_json = {
                **(job.summary_json or {}),
                "active_commit_token": None,
                "last_commit_attempt": {
                    "status": job.status,
                    "processed_rows": total_processed,
                    "elapsed_ms": int((perf_counter() - commit_started) * 1000),
                    "error": str(exc),
                },
            }
            job.completed_at = utcnow()
            progress_db.add(job)
            progress_db.commit()
    finally:
        work_db.close()
        progress_db.close()


def _require_commit_lease(db: Session, job_id: str, attempt_token: str) -> TrainingWorkbookImportJob:
    db.expire_all()
    job = db.get(TrainingWorkbookImportJob, job_id)
    if (
        not job
        or job.status != "COMMITTING"
        or _commit_attempt_token(job) != attempt_token
    ):
        raise WorkbookCommitLeaseLost("This workbook commit worker was superseded.")
    if job.cancel_requested:
        raise RuntimeError("IMPORT_CANCELLED")
    return job


def _commit_progress(
    db: Session,
    job_id: str,
    attempt_token: str,
    processed: int,
    stage: str,
    sheet: str,
    label: Optional[str],
) -> None:
    job = _require_commit_lease(db, job_id, attempt_token)
    job.processed_rows = min(job.total_rows, processed)
    job.stage = stage
    job.current_sheet = sheet
    job.current_record_label = (label or "")[:255] or None
    job.updated_at = utcnow()
    db.add(job)
    db.commit()
