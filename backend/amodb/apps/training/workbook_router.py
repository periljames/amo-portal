from __future__ import annotations

import hashlib
import io
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, Response, UploadFile, status
from openpyxl import Workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ...database import get_db
from ...security import get_current_active_user
from ..accounts import models as account_models
from . import models as training_models
from .permissions import TrainingCapability, require_training_capability, tenant_id_for
from .workbook_import import (
    commit_workbook_import,
    new_commit_attempt_token,
    process_workbook_preview,
    utcnow,
)
from .workbook_models import (
    PersonnelLicence,
    TrainingCourseRoleRule,
    TrainingPersonRole,
    TrainingRoleGroup,
    TrainingWorkbookImportJob,
    TrainingWorkbookImportRow,
    TrainingWorkbookImportSheet,
)
from .workbook_schemas import (
    PersonnelLicenceRead,
    TrainingCourseRoleRuleRead,
    TrainingCourseRoleRuleWrite,
    TrainingPersonRoleRead,
    TrainingPersonRoleWrite,
    TrainingRoleGroupRead,
    TrainingRoleGroupWrite,
    TrainingWorkbookImportJobRead,
    TrainingWorkbookImportJobPage,
    WorkbookImportCommitRequest,
    WorkbookImportRowPage,
    WorkbookImportRowRead,
    WorkbookImportSheetRead,
)

router = APIRouter(prefix="/workbook-imports", tags=["training-workbook-imports"])

_IMPORT_DIR = Path(os.getenv("TRAINING_UPLOAD_DIR", "uploads/training")).resolve() / "workbook-imports"
_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
_MAX_BYTES = int(os.getenv("TRAINING_WORKBOOK_MAX_UPLOAD_BYTES", str(40 * 1024 * 1024)))
_STALE_COMMIT_SECONDS = max(30, int(os.getenv("TRAINING_WORKBOOK_STALE_COMMIT_SECONDS", "90")))
_MAX_AUTO_RECOVERIES = max(1, int(os.getenv("TRAINING_WORKBOOK_MAX_AUTO_RECOVERIES", "3")))


_editor = require_training_capability(TrainingCapability.COURSE_MANAGE)


def _job_for_user(db: Session, current_user: account_models.User, job_id: str) -> TrainingWorkbookImportJob:
    query = db.query(TrainingWorkbookImportJob).filter(
        TrainingWorkbookImportJob.id == job_id,
        TrainingWorkbookImportJob.amo_id == tenant_id_for(current_user),
    )
    job = query.first()
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Training workbook import not found.")
    return job


def _age_seconds(value: Optional[datetime]) -> float:
    if value is None:
        return float("inf")
    checkpoint = value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)
    return max(0.0, (utcnow() - checkpoint).total_seconds())


def _recover_stale_commit(
    db: Session,
    job: TrainingWorkbookImportJob,
    background_tasks: BackgroundTasks,
) -> TrainingWorkbookImportJob:
    """Renew an orphaned commit after an API/DB restart without duplicating it."""
    if job.status not in {"QUEUED_COMMIT", "COMMITTING"}:
        return job
    if _age_seconds(job.updated_at) < _STALE_COMMIT_SECONDS:
        return job

    summary = dict(job.summary_json or {})
    recoveries = int(summary.get("automatic_recovery_attempts") or 0)
    if job.cancel_requested:
        job.status = "CANCELLED"
        job.stage = "CANCELLED"
        job.completed_at = utcnow()
        job.error_message = None
        summary["active_commit_token"] = None
        job.summary_json = summary
        db.add(job)
        db.commit()
        db.refresh(job)
        return job

    if recoveries >= _MAX_AUTO_RECOVERIES:
        job.status = "FAILED"
        job.stage = "FAILED"
        job.completed_at = utcnow()
        job.error_message = (
            "The database connection interrupted this commit repeatedly. "
            "The atomic write was stopped safely; use Retry reviewed import after database health is stable."
        )
        summary["active_commit_token"] = None
        summary["last_commit_attempt"] = {
            "status": "FAILED",
            "processed_rows": job.processed_rows,
            "error": job.error_message,
        }
        job.summary_json = summary
        db.add(job)
        db.commit()
        db.refresh(job)
        return job

    token = new_commit_attempt_token()
    commit_request = summary.get("commit_request")
    force_reimport = bool(commit_request.get("force_reimport")) if isinstance(commit_request, dict) else False
    summary.update(
        {
            "active_commit_token": token,
            "automatic_recovery_attempts": recoveries + 1,
            "last_recovery_at": utcnow().isoformat(),
            "last_commit_attempt": {
                "status": "RECOVERING",
                "processed_rows": job.processed_rows,
                "reason": "Worker checkpoint expired after a database or application interruption.",
            },
        }
    )
    previous_status = job.status
    previous_updated_at = job.updated_at
    updated = (
        db.query(TrainingWorkbookImportJob)
        .filter(
            TrainingWorkbookImportJob.id == job.id,
            TrainingWorkbookImportJob.status == previous_status,
            TrainingWorkbookImportJob.updated_at == previous_updated_at,
        )
        .update(
            {
                TrainingWorkbookImportJob.status: "QUEUED_COMMIT",
                TrainingWorkbookImportJob.stage: "RECOVERING_COMMIT",
                TrainingWorkbookImportJob.processed_rows: 0,
                TrainingWorkbookImportJob.current_sheet: None,
                TrainingWorkbookImportJob.current_record_label: None,
                TrainingWorkbookImportJob.error_message: None,
                TrainingWorkbookImportJob.completed_at: None,
                TrainingWorkbookImportJob.summary_json: summary,
                TrainingWorkbookImportJob.updated_at: utcnow(),
            },
            synchronize_session=False,
        )
    )
    db.commit()
    db.expire_all()
    refreshed = db.get(TrainingWorkbookImportJob, job.id)
    if updated == 1 and refreshed is not None:
        background_tasks.add_task(
            commit_workbook_import,
            refreshed.id,
            force_reimport=force_reimport,
            attempt_token=token,
        )
    return refreshed or job


def _sheet_read(item: TrainingWorkbookImportSheet) -> WorkbookImportSheetRead:
    return WorkbookImportSheetRead.model_validate(item)


def _job_read(db: Session, job: TrainingWorkbookImportJob) -> TrainingWorkbookImportJobRead:
    sheets = (
        db.query(TrainingWorkbookImportSheet)
        .filter(TrainingWorkbookImportSheet.job_id == job.id)
        .order_by(TrainingWorkbookImportSheet.display_order.asc())
        .all()
    )
    return TrainingWorkbookImportJobRead(
        id=job.id,
        amo_id=job.amo_id,
        actor_user_id=job.actor_user_id,
        filename=job.filename,
        size_bytes=job.size_bytes,
        file_sha256=job.file_sha256,
        duplicate_of_job_id=job.duplicate_of_job_id,
        status=job.status,
        stage=job.stage,
        current_sheet=job.current_sheet,
        current_record_label=job.current_record_label,
        processed_rows=job.processed_rows,
        total_rows=job.total_rows,
        created_count=job.created_count,
        updated_count=job.updated_count,
        unchanged_count=job.unchanged_count,
        skipped_count=job.skipped_count,
        failed_count=job.failed_count,
        review_count=job.review_count,
        summary=job.summary_json or {},
        error_message=job.error_message,
        cancel_requested=job.cancel_requested,
        created_at=job.created_at,
        started_at=job.started_at,
        preview_completed_at=job.preview_completed_at,
        committed_at=job.committed_at,
        completed_at=job.completed_at,
        updated_at=job.updated_at,
        sheets=[_sheet_read(item) for item in sheets],
    )


@router.post("", response_model=TrainingWorkbookImportJobRead, status_code=status.HTTP_202_ACCEPTED)
async def create_workbook_import(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    idempotency_key: Optional[str] = Query(default=None, max_length=160),
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingWorkbookImportJobRead:
    amo_id = tenant_id_for(current_user)
    filename = Path(file.filename or "training-workbook.xlsx").name
    extension = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if extension not in {"xlsx", "xlsm", "xltx", "xltm"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload a modern Excel .xlsx or .xlsm training workbook.")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The uploaded workbook is empty.")
    if _MAX_BYTES and len(content) > _MAX_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=f"Workbook exceeds the {_MAX_BYTES // (1024 * 1024)} MB import limit.")

    sha = hashlib.sha256(content).hexdigest()
    key = (idempotency_key or f"preview:{sha}:{current_user.id}").strip()
    existing_key = db.query(TrainingWorkbookImportJob).filter(
        TrainingWorkbookImportJob.amo_id == amo_id,
        TrainingWorkbookImportJob.idempotency_key == key,
    ).first()
    if existing_key:
        return _job_read(db, existing_key)

    duplicate = db.query(TrainingWorkbookImportJob).filter(
        TrainingWorkbookImportJob.amo_id == amo_id,
        TrainingWorkbookImportJob.file_sha256 == sha,
        TrainingWorkbookImportJob.status == "COMPLETED",
        TrainingWorkbookImportJob.committed_at.isnot(None),
    ).order_by(TrainingWorkbookImportJob.committed_at.desc()).first()

    amo_dir = _IMPORT_DIR / str(amo_id)
    amo_dir.mkdir(parents=True, exist_ok=True)
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "-", filename).strip("-") or "training-workbook.xlsx"
    storage_path = amo_dir / f"{sha[:16]}-{safe_stem}"
    if not storage_path.exists():
        temporary = storage_path.with_suffix(storage_path.suffix + ".tmp")
        temporary.write_bytes(content)
        temporary.replace(storage_path)

    job = TrainingWorkbookImportJob(
        amo_id=amo_id,
        actor_user_id=current_user.id,
        filename=filename,
        content_type=file.content_type,
        size_bytes=len(content),
        file_sha256=sha,
        storage_path=str(storage_path),
        idempotency_key=key,
        duplicate_of_job_id=duplicate.id if duplicate else None,
        status="QUEUED",
        stage="UPLOAD_COMPLETE",
        summary_json={"duplicate_committed_import": duplicate.id if duplicate else None},
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(process_workbook_preview, job.id)
    return _job_read(db, job)


@router.get("", response_model=TrainingWorkbookImportJobPage)
def list_workbook_imports(
    job_status: Optional[str] = Query(default=None, alias="status"),
    q: Optional[str] = None,
    limit: int = Query(default=25, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingWorkbookImportJobPage:
    query = db.query(TrainingWorkbookImportJob).filter(TrainingWorkbookImportJob.amo_id == tenant_id_for(current_user))
    if job_status:
        query = query.filter(TrainingWorkbookImportJob.status == job_status.upper())
    if q and q.strip():
        token = f"%{q.strip()}%"
        query = query.filter(or_(TrainingWorkbookImportJob.filename.ilike(token), TrainingWorkbookImportJob.file_sha256.ilike(token), TrainingWorkbookImportJob.error_message.ilike(token)))
    total = int(query.count())
    rows = query.order_by(TrainingWorkbookImportJob.created_at.desc()).offset(offset).limit(limit).all()
    return TrainingWorkbookImportJobPage(items=[_job_read(db, row) for row in rows], total=total, limit=limit, offset=offset, has_more=offset + len(rows) < total)


@router.get("/template")
def download_workbook_template(
    current_user: account_models.User = Depends(_editor),
) -> Response:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["AMO Training Workbook Import Template", "Schema version", "2026.08"])
    readme.append(["Instructions", "Keep sheet names and headers unchanged. Dates use YYYY-MM-DD. Upload from Training > Templates / Settings > Import Excel."])
    readme.append(["Identity", "PersonID must match the canonical staff code, or the preview will require an explicit reconciliation decision."])
    readme.append(["Governance", "Every workbook is previewed, conflict-checked and retained before an atomic commit. No row activates silently."])
    sheets = {
        "People": ["PersonID", "FIRSTNAME", "LASTNAME", "PersonName", "Email", "Department", "Position", "PhoneNumber", "HireDate", "Employment_Status", "Status", "KAMEL NO:", "Internal Certification Stamp No:", "initial_auth"],
        "Courses": ["CourseID", "CourseName", "Status", "FrequencyMonths", "Category", "Mandatory", "Scope", "Reference", "Active"],
        "Training": ["RecordID", "PersonID", "PersonName", "CourseID", "CourseName", "LastTrainingDate", "NextDueDate", "DaysToDue", "Status"],
        "tblRoleGroups": ["RoleGroup", "Description"],
        "tblPersonRoles": ["PersonID", "RoleGroup", "Department", "Position", "ActiveRole", "Notes"],
        "tblCourseMatrix": ["CourseID", "RoleGroup", "Required", "RequirementType", "Notes"],
        "Params": ["Setting", "Value"],
    }
    samples = {
        "People": ["EMP-001", "ALEX", "TECHNICIAN", "Alex Technician", "alex@example.invalid", "MAINTENANCE", "TECHNICIAN", "+000000000", "2026-01-15", "EMPLOYED", "Active", "AMEL-0001", "STAMP-001", "2026-02-01"],
        "Courses": ["HF-INIT", "Human Factors Initial", "Initial", "", "HUMAN_FACTORS", "Yes", "ALL MAINTENANCE", "MPM 3.7", "Yes"],
        "Training": ["REC-001", "EMP-001", "Alex Technician", "HF-INIT", "Human Factors Initial", "2026-02-15", "", "", "CURRENT"],
        "tblRoleGroups": ["TECHNICIANS", "Maintenance technicians"],
        "tblPersonRoles": ["EMP-001", "TECHNICIANS", "MAINTENANCE", "TECHNICIAN", "Yes", ""],
        "tblCourseMatrix": ["HF-INIT", "TECHNICIANS", "Yes", "GENERAL", "MPM 3.7"],
        "Params": ["Default Frequency (months)", 24],
    }
    for title, headers in sheets.items():
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        sheet.append(samples[title])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(34, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="amo-training-import-template-2026.08.xlsx"', "X-Training-Schema-Version": "2026.08"},
    )


@router.get("/{job_id}", response_model=TrainingWorkbookImportJobRead)
def get_workbook_import(
    job_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingWorkbookImportJobRead:
    job = _job_for_user(db, current_user, job_id)
    job = _recover_stale_commit(db, job, background_tasks)
    return _job_read(db, job)


@router.get("/{job_id}/rows", response_model=WorkbookImportRowPage)
def list_workbook_import_rows(
    job_id: str,
    sheet: Optional[str] = None,
    row_status: Optional[str] = Query(default=None, alias="status"),
    review_only: bool = False,
    q: Optional[str] = None,
    limit: int = Query(default=80, ge=1, le=250),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> WorkbookImportRowPage:
    job = _job_for_user(db, current_user, job_id)
    query = db.query(TrainingWorkbookImportRow).filter(TrainingWorkbookImportRow.job_id == job.id)
    if sheet:
        query = query.filter(TrainingWorkbookImportRow.sheet_name == sheet)
    if row_status:
        query = query.filter(TrainingWorkbookImportRow.status == row_status.upper())
    if review_only:
        query = query.filter(TrainingWorkbookImportRow.decision_required.is_(True))
    if q and q.strip():
        term = f"%{q.strip()}%"
        query = query.filter(or_(TrainingWorkbookImportRow.display_label.ilike(term), TrainingWorkbookImportRow.source_key.ilike(term), TrainingWorkbookImportRow.issue_message.ilike(term)))
    total = query.count()
    rows = query.order_by(TrainingWorkbookImportRow.sheet_name.asc(), TrainingWorkbookImportRow.source_row.asc()).offset(offset).limit(limit).all()
    return WorkbookImportRowPage(
        total=total,
        limit=limit,
        offset=offset,
        items=[
            WorkbookImportRowRead(
                id=item.id,
                sheet_name=item.sheet_name,
                source_row=item.source_row,
                entity_type=item.entity_type,
                source_key=item.source_key,
                display_label=item.display_label,
                proposed_action=item.proposed_action,
                status=item.status,
                decision_required=item.decision_required,
                decision=item.decision,
                decision_options=item.decision_options or [],
                changes=item.changes_json or [],
                issue_code=item.issue_code,
                issue_message=item.issue_message,
                payload=item.payload_json or {},
                committed_entity_id=item.committed_entity_id,
            )
            for item in rows
        ],
    )


@router.post("/{job_id}/commit", response_model=TrainingWorkbookImportJobRead, status_code=status.HTTP_202_ACCEPTED)
def commit_import(
    job_id: str,
    payload: WorkbookImportCommitRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingWorkbookImportJobRead:
    job = _job_for_user(db, current_user, job_id)
    if job.status not in {"PREVIEW_READY", "REVIEW_REQUIRED", "FAILED"}:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Import is currently {job.status.lower()} and cannot be committed.")
    rows_by_id = {
        item.id: item
        for item in db.query(TrainingWorkbookImportRow).filter(
            TrainingWorkbookImportRow.job_id == job.id,
            TrainingWorkbookImportRow.id.in_([decision.row_id for decision in payload.decisions] or [""]),
        ).all()
    }
    for decision in payload.decisions:
        row = rows_by_id.get(decision.row_id)
        if not row:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Decision row {decision.row_id} does not belong to this import.")
        selected = decision.decision.upper()
        if row.decision_options and selected not in {str(option).upper() for option in row.decision_options}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Decision {selected} is not valid for row {row.source_row}.")
        row.decision = selected
        row.updated_at = utcnow()
        db.add(row)
    db.commit()

    unresolved = db.query(TrainingWorkbookImportRow).filter(
        TrainingWorkbookImportRow.job_id == job.id,
        TrainingWorkbookImportRow.decision_required.is_(True),
        TrainingWorkbookImportRow.decision.is_(None),
    ).count()
    if unresolved:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"{unresolved} personnel or conflict review decision(s) are still required.")

    attempt_token = new_commit_attempt_token()
    job.status = "QUEUED_COMMIT"
    job.stage = "QUEUED_COMMIT"
    job.actor_user_id = current_user.id
    job.completed_at = None
    job.error_message = None
    job.cancel_requested = False
    job.processed_rows = 0
    job.current_sheet = None
    job.current_record_label = None
    job.summary_json = {
        **(job.summary_json or {}),
        "active_commit_token": attempt_token,
        "automatic_recovery_attempts": 0,
        "commit_request": {
            "force_reimport": bool(payload.force_reimport),
            "requested_at": utcnow().isoformat(),
        },
    }
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(
        commit_workbook_import,
        job.id,
        force_reimport=payload.force_reimport,
        attempt_token=attempt_token,
    )
    return _job_read(db, job)


@router.post("/{job_id}/cancel", response_model=TrainingWorkbookImportJobRead)
def cancel_import(
    job_id: str,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingWorkbookImportJobRead:
    job = _job_for_user(db, current_user, job_id)
    if job.status in {"COMPLETED", "CANCELLED"}:
        return _job_read(db, job)
    job.cancel_requested = True
    if job.status in {"QUEUED", "PARSING", "QUEUED_COMMIT", "COMMITTING", "PREVIEW_READY", "REVIEW_REQUIRED", "FAILED"}:
        job.status = "CANCELLED"
        job.stage = "CANCELLED"
        job.completed_at = utcnow()
        job.current_sheet = None
        job.current_record_label = None
        job.summary_json = {**(job.summary_json or {}), "active_commit_token": None}
    db.add(job)
    db.commit()
    db.refresh(job)
    return _job_read(db, job)


@router.get("/users/{user_id}/licences", response_model=list[PersonnelLicenceRead])
def list_personnel_licences(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(get_current_active_user),
) -> list[PersonnelLicenceRead]:
    user_query = db.query(account_models.User).filter(
        account_models.User.id == user_id,
        account_models.User.amo_id == tenant_id_for(current_user),
    )
    user = user_query.first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Personnel user not found.")
    rows = db.query(PersonnelLicence).filter(PersonnelLicence.amo_id == user.amo_id, PersonnelLicence.user_id == user.id).order_by(PersonnelLicence.is_primary.desc(), PersonnelLicence.authority.asc()).all()
    return [PersonnelLicenceRead.model_validate(item) for item in rows]


@router.get("/catalog/role-groups", response_model=list[TrainingRoleGroupRead])
def list_role_groups(
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(get_current_active_user),
) -> list[TrainingRoleGroupRead]:
    rows = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.amo_id == tenant_id_for(current_user), TrainingRoleGroup.is_active.is_(True)).order_by(TrainingRoleGroup.code.asc()).all()
    return [TrainingRoleGroupRead.model_validate(item) for item in rows]


@router.get("/catalog/role-rules", response_model=list[TrainingCourseRoleRuleRead])
def list_role_rules(
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(get_current_active_user),
) -> list[TrainingCourseRoleRuleRead]:
    rows = (
        db.query(TrainingCourseRoleRule, TrainingRoleGroup, training_models.TrainingCourse)
        .join(TrainingRoleGroup, TrainingCourseRoleRule.role_group_id == TrainingRoleGroup.id)
        .join(training_models.TrainingCourse, TrainingCourseRoleRule.course_id == training_models.TrainingCourse.id)
        .filter(
            TrainingCourseRoleRule.amo_id == tenant_id_for(current_user),
            TrainingCourseRoleRule.is_active.is_(True),
        )
        .order_by(TrainingRoleGroup.code.asc(), training_models.TrainingCourse.course_id.asc())
        .all()
    )
    return [TrainingCourseRoleRuleRead(id=rule.id, course_id=rule.course_id, course_code=course.course_id, course_name=course.course_name, role_group_id=rule.role_group_id, role_group_code=group.code, is_required=rule.is_required, requirement_type=rule.requirement_type, notes=rule.notes, is_active=rule.is_active) for rule, group, course in rows]


@router.get("/users/{user_id}/roles", response_model=list[TrainingPersonRoleRead])
def list_person_roles(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(get_current_active_user),
) -> list[TrainingPersonRoleRead]:
    amo_id = tenant_id_for(current_user)
    user = db.query(account_models.User).filter(account_models.User.id == user_id, account_models.User.amo_id == amo_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Personnel user not found.")
    rows = db.query(TrainingPersonRole, TrainingRoleGroup).join(TrainingRoleGroup, TrainingPersonRole.role_group_id == TrainingRoleGroup.id).filter(TrainingPersonRole.amo_id == amo_id, TrainingPersonRole.user_id == user.id, TrainingPersonRole.is_active.is_(True)).all()
    return [TrainingPersonRoleRead(id=assignment.id, person_id=assignment.person_id, personnel_profile_id=assignment.personnel_profile_id, user_id=assignment.user_id, role_group_id=assignment.role_group_id, role_group_code=group.code, person_name=user.full_name, staff_code=user.staff_code, department=assignment.department, position=assignment.position, notes=assignment.notes, is_active=assignment.is_active) for assignment, group in rows]


@router.get("/catalog/person-roles", response_model=list[TrainingPersonRoleRead])
def list_catalog_person_roles(
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(get_current_active_user),
) -> list[TrainingPersonRoleRead]:
    amo_id = tenant_id_for(current_user)
    rows = (
        db.query(TrainingPersonRole, TrainingRoleGroup, account_models.User)
        .join(TrainingRoleGroup, TrainingPersonRole.role_group_id == TrainingRoleGroup.id)
        .outerjoin(account_models.User, TrainingPersonRole.user_id == account_models.User.id)
        .filter(TrainingPersonRole.amo_id == amo_id, TrainingPersonRole.is_active.is_(True))
        .order_by(TrainingRoleGroup.code.asc(), TrainingPersonRole.person_id.asc())
        .all()
    )
    return [TrainingPersonRoleRead(id=assignment.id, person_id=assignment.person_id, personnel_profile_id=assignment.personnel_profile_id, user_id=assignment.user_id, role_group_id=assignment.role_group_id, role_group_code=group.code, person_name=user.full_name if user else assignment.person_id, staff_code=user.staff_code if user else assignment.person_id, department=assignment.department, position=assignment.position, notes=assignment.notes, is_active=assignment.is_active) for assignment, group, user in rows]


@router.post("/catalog/role-groups", response_model=TrainingRoleGroupRead)
def save_role_group(
    payload: TrainingRoleGroupWrite,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingRoleGroupRead:
    amo_id = tenant_id_for(current_user)
    code = payload.code.strip().upper()
    row = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.amo_id == amo_id, TrainingRoleGroup.code == code).first()
    if row is None:
        row = TrainingRoleGroup(amo_id=amo_id, code=code)
    row.description = payload.description.strip() if payload.description else None
    row.is_active = payload.is_active
    db.add(row); db.commit(); db.refresh(row)
    return TrainingRoleGroupRead.model_validate(row)


@router.delete("/catalog/role-groups/{group_id}", response_model=TrainingRoleGroupRead)
def deactivate_role_group(
    group_id: str,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingRoleGroupRead:
    row = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.id == group_id, TrainingRoleGroup.amo_id == tenant_id_for(current_user)).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Training role group not found.")
    row.is_active = False; db.add(row); db.commit(); db.refresh(row)
    return TrainingRoleGroupRead.model_validate(row)


@router.post("/catalog/person-roles", response_model=TrainingPersonRoleRead)
def save_person_role(
    payload: TrainingPersonRoleWrite,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingPersonRoleRead:
    amo_id = tenant_id_for(current_user)
    user = db.query(account_models.User).filter(account_models.User.id == payload.user_id, account_models.User.amo_id == amo_id).first()
    group = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.id == payload.role_group_id, TrainingRoleGroup.amo_id == amo_id).first()
    if user is None or group is None:
        raise HTTPException(status_code=422, detail="Select a valid person and role group in this AMO.")
    profile = db.query(account_models.PersonnelProfile).filter(account_models.PersonnelProfile.amo_id == amo_id, account_models.PersonnelProfile.user_id == user.id).first()
    person_id = str(user.staff_code or (profile.person_id if profile else user.id)).strip().upper()
    row = db.query(TrainingPersonRole).filter(TrainingPersonRole.amo_id == amo_id, TrainingPersonRole.person_id == person_id, TrainingPersonRole.role_group_id == group.id).first()
    if row is None:
        row = TrainingPersonRole(amo_id=amo_id, person_id=person_id, role_group_id=group.id)
    row.user_id = user.id; row.personnel_profile_id = profile.id if profile else None
    row.department = payload.department.strip() if payload.department else None
    row.position = payload.position.strip() if payload.position else None
    row.notes = payload.notes.strip() if payload.notes else None
    row.is_active = payload.is_active
    db.add(row); db.commit(); db.refresh(row)
    return TrainingPersonRoleRead(id=row.id, person_id=row.person_id, personnel_profile_id=row.personnel_profile_id, user_id=row.user_id, role_group_id=row.role_group_id, role_group_code=group.code, person_name=user.full_name, staff_code=user.staff_code, department=row.department, position=row.position, notes=row.notes, is_active=row.is_active)


@router.delete("/catalog/person-roles/{assignment_id}", response_model=TrainingPersonRoleRead)
def deactivate_person_role(
    assignment_id: str,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingPersonRoleRead:
    amo_id = tenant_id_for(current_user)
    row = db.query(TrainingPersonRole).filter(TrainingPersonRole.id == assignment_id, TrainingPersonRole.amo_id == amo_id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Personnel role assignment not found.")
    group = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.id == row.role_group_id, TrainingRoleGroup.amo_id == amo_id).one()
    user = db.query(account_models.User).filter(account_models.User.id == row.user_id, account_models.User.amo_id == amo_id).first()
    row.is_active = False; db.add(row); db.commit(); db.refresh(row)
    return TrainingPersonRoleRead(id=row.id, person_id=row.person_id, personnel_profile_id=row.personnel_profile_id, user_id=row.user_id, role_group_id=row.role_group_id, role_group_code=group.code, person_name=user.full_name if user else row.person_id, staff_code=user.staff_code if user else row.person_id, department=row.department, position=row.position, notes=row.notes, is_active=row.is_active)


@router.post("/catalog/role-rules", response_model=TrainingCourseRoleRuleRead)
def save_role_rule(
    payload: TrainingCourseRoleRuleWrite,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingCourseRoleRuleRead:
    amo_id = tenant_id_for(current_user)
    course = db.query(training_models.TrainingCourse).filter(training_models.TrainingCourse.id == payload.course_id, training_models.TrainingCourse.amo_id == amo_id).first()
    group = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.id == payload.role_group_id, TrainingRoleGroup.amo_id == amo_id).first()
    requirement_type = payload.requirement_type.strip().upper()
    if course is None or group is None:
        raise HTTPException(status_code=422, detail="Select a valid course and role group in this AMO.")
    row = db.query(TrainingCourseRoleRule).filter(TrainingCourseRoleRule.amo_id == amo_id, TrainingCourseRoleRule.course_id == course.id, TrainingCourseRoleRule.role_group_id == group.id, TrainingCourseRoleRule.requirement_type == requirement_type).first()
    if row is None:
        row = TrainingCourseRoleRule(amo_id=amo_id, course_id=course.id, role_group_id=group.id, requirement_type=requirement_type)
    row.is_required = payload.is_required; row.notes = payload.notes.strip() if payload.notes else None; row.is_active = payload.is_active
    db.add(row); db.commit(); db.refresh(row)
    return TrainingCourseRoleRuleRead(id=row.id, course_id=row.course_id, course_code=course.course_id, course_name=course.course_name, role_group_id=row.role_group_id, role_group_code=group.code, is_required=row.is_required, requirement_type=row.requirement_type, notes=row.notes, is_active=row.is_active)


@router.delete("/catalog/role-rules/{rule_id}", response_model=TrainingCourseRoleRuleRead)
def deactivate_role_rule(
    rule_id: str,
    db: Session = Depends(get_db),
    current_user: account_models.User = Depends(_editor),
) -> TrainingCourseRoleRuleRead:
    amo_id = tenant_id_for(current_user)
    row = db.query(TrainingCourseRoleRule).filter(TrainingCourseRoleRule.id == rule_id, TrainingCourseRoleRule.amo_id == amo_id).first()
    if row is None:
        raise HTTPException(status_code=404, detail="Training matrix rule not found.")
    course = db.query(training_models.TrainingCourse).filter(training_models.TrainingCourse.id == row.course_id, training_models.TrainingCourse.amo_id == amo_id).one()
    group = db.query(TrainingRoleGroup).filter(TrainingRoleGroup.id == row.role_group_id, TrainingRoleGroup.amo_id == amo_id).one()
    row.is_active = False; db.add(row); db.commit(); db.refresh(row)
    return TrainingCourseRoleRuleRead(id=row.id, course_id=row.course_id, course_code=course.course_id, course_name=course.course_name, role_group_id=row.role_group_id, role_group_code=group.code, is_required=row.is_required, requirement_type=row.requirement_type, notes=row.notes, is_active=row.is_active)
