from __future__ import annotations

import hashlib
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException

from . import ingestion_schemas


MAX_SOURCE_BYTES = 75 * 1024 * 1024
MAX_SAMPLE_ROWS = 20
MAX_SAMPLE_COLUMNS = 24
SUPPORTED_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}


Q400_MPD_CORE_SHEETS = {"section 1", "section 2", "section 3", "supplement 1"}
Q400_MPD_SUPPORT_SHEETS = {
    "appendix a",
    "appendix e",
    "appendix f",
    "appendix i",
    "appendix j",
    "appendix n",
    "appendix p",
    "appendix q",
    "appendix r",
    "appendix t",
    "appendix u",
}
LEGACY_EMP_FILENAME_SERIES = {
    "81mplm": ("100", "DHC8_100_EMP_V1", "DHC8_100_MPD_SOURCE_INTAKE"),
    "82mplm": ("200", "DHC8_200_EMP_V1", "DHC8_200_MPD_SOURCE_INTAKE"),
    "83mplm": ("300", "DHC8_300_EMP_V1", "DHC8_300_MPD_SOURCE_INTAKE"),
}


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        # Preview values are not operational counters. Render floats as strings so
        # a workbook's binary floating representation cannot silently become a
        # controlled maintenance value during preview.
        return format(value, ".15g")
    if isinstance(value, bytes):
        return f"<binary:{len(value)} bytes>"
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _sample_rows(rows: list[list[Any]]) -> list[list[Any]]:
    samples: list[list[Any]] = []
    for row in rows:
        values = [_cell_value(value) for value in row[:MAX_SAMPLE_COLUMNS]]
        while values and values[-1] is None:
            values.pop()
        if not any(value not in (None, "") for value in values):
            continue
        samples.append(values)
        if len(samples) >= MAX_SAMPLE_ROWS:
            break
    return samples


def _normalise_sheet_name(value: str) -> str:
    return " ".join(value.strip().lower().split())


def detect_profile(
    sheet_names: list[str],
    *,
    filename: str | None = None,
) -> tuple[str, str, str | None, list[str]]:
    normalised = {_normalise_sheet_name(name) for name in sheet_names}
    warnings: list[str] = []
    q400_core = Q400_MPD_CORE_SHEETS & normalised
    q400_support = Q400_MPD_SUPPORT_SHEETS & normalised
    if q400_core == Q400_MPD_CORE_SHEETS and len(q400_support) >= 5:
        missing_support = sorted(Q400_MPD_SUPPORT_SHEETS - normalised)
        if missing_support:
            warnings.append(
                "Q400 MPD profile detected, but not all known supporting appendices are present: "
                + ", ".join(missing_support)
            )
        return "DHC8_400_MPD_V1", "HIGH", "DHC8_400_MPD_SOURCE_INTAKE", warnings

    joined = " | ".join(sorted(normalised))
    legacy_emp = "work packages" in joined and (
        "cross reference" in joined or "l-check" in joined or "out of phase" in joined
    )
    if legacy_emp:
        stem = Path(filename or "").stem.lower().replace(" ", "").replace("_", "")
        for marker, (series, profile, pack_code) in LEGACY_EMP_FILENAME_SERIES.items():
            if stem.startswith(marker):
                warnings.append(
                    f"Legacy DHC-8 Series {series} equalized-maintenance workbook profile inferred from the known OEM MPLM filename. The controlled publication metadata and PDF/manual revision must confirm series and revision before materialization."
                )
                return profile, "MEDIUM", pack_code, warnings
        warnings.append(
            "Legacy DHC-8 equalized-maintenance workbook detected, but series identity was not proven. It remains preview-only until the controlled publication record identifies the series."
        )
        return "DHC8_EMP_LEGACY_V1", "LOW", None, warnings

    warnings.append(
        "No governed OEM workbook profile matched. The workbook may be previewed, but it must not materialize engineering content until an approved mapping profile exists."
    )
    return "UNMAPPED", "LOW", None, warnings


def _preview_xlsx(content: bytes) -> tuple[str, list[ingestion_schemas.WorkbookSheetPreview]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is pinned in runtime
        raise HTTPException(status_code=503, detail="XLSX reader is not installed") from exc

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Workbook could not be opened safely") from exc

    previews: list[ingestion_schemas.WorkbookSheetPreview] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for row_index, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(max(worksheet.max_row or 1, 1), 80),
                    max_col=min(max(worksheet.max_column or 1, 1), MAX_SAMPLE_COLUMNS),
                    values_only=True,
                ),
                start=1,
            ):
                rows.append(list(row))
                if row_index >= 80:
                    break
            previews.append(
                ingestion_schemas.WorkbookSheetPreview(
                    name=worksheet.title,
                    state=str(getattr(worksheet, "sheet_state", "visible")).upper(),
                    row_count=int(worksheet.max_row or 0),
                    column_count=int(worksheet.max_column or 0),
                    sample_rows=_sample_rows(rows),
                )
            )
    finally:
        workbook.close()
    return "OOXML", previews


def _preview_xls(content: bytes) -> tuple[str, list[ingestion_schemas.WorkbookSheetPreview]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency is pinned in runtime
        raise HTTPException(status_code=503, detail="Legacy XLS reader is not installed") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Legacy workbook could not be opened safely") from exc

    visibility = list(getattr(workbook, "sheet_visibility", []))
    states = {0: "VISIBLE", 1: "HIDDEN", 2: "VERY_HIDDEN"}
    previews: list[ingestion_schemas.WorkbookSheetPreview] = []
    try:
        for index, sheet_name in enumerate(workbook.sheet_names()):
            worksheet = workbook.sheet_by_index(index)
            rows = [
                worksheet.row_values(row_index, start_colx=0, end_colx=min(worksheet.ncols, MAX_SAMPLE_COLUMNS))
                for row_index in range(min(worksheet.nrows, 80))
            ]
            state = states.get(visibility[index], "UNKNOWN") if index < len(visibility) else "VISIBLE"
            previews.append(
                ingestion_schemas.WorkbookSheetPreview(
                    name=sheet_name,
                    state=state,
                    row_count=int(worksheet.nrows),
                    column_count=int(worksheet.ncols),
                    sample_rows=_sample_rows(rows),
                )
            )
            workbook.unload_sheet(index)
    finally:
        workbook.release_resources()
    return "BIFF8", previews


def preview_oem_workbook(
    *,
    filename: str,
    content: bytes,
) -> ingestion_schemas.OemWorkbookPreview:
    extension = Path(filename or "").suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=415,
            detail="OEM source workbook must be XLS, XLSX, or XLSM",
        )
    if not content:
        raise HTTPException(status_code=422, detail="OEM source workbook is empty")
    if len(content) > MAX_SOURCE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"OEM source workbook exceeds the {MAX_SOURCE_BYTES // (1024 * 1024)} MB preview limit",
        )

    checksum = hashlib.sha256(content).hexdigest()
    if extension == ".xls":
        workbook_kind, sheets = _preview_xls(content)
    else:
        workbook_kind, sheets = _preview_xlsx(content)

    profile, confidence, pack_code, warnings = detect_profile(
        [sheet.name for sheet in sheets],
        filename=filename,
    )
    hidden = [sheet.name for sheet in sheets if sheet.state != "VISIBLE"]
    if hidden:
        warnings.append(
            "Hidden workbook sheets were detected and included in the inventory: " + ", ".join(hidden)
        )
    if extension == ".xlsm":
        warnings.append(
            "Macro-enabled workbook detected. Macros are never executed; only worksheet values/formulas and metadata are inspected."
        )

    return ingestion_schemas.OemWorkbookPreview(
        filename=Path(filename).name,
        extension=extension,
        size_bytes=len(content),
        checksum_sha256=checksum,
        detected_profile=profile,
        profile_confidence=confidence,
        workbook_kind=workbook_kind,
        sheets=sheets,
        warnings=warnings,
        recommended_pack_code=pack_code,
        source_manifest={
            "filename": Path(filename).name,
            "sha256": checksum,
            "profile": profile,
            "sheet_count": len(sheets),
            "sheets": [
                {
                    "name": sheet.name,
                    "state": sheet.state,
                    "rows": sheet.row_count,
                    "columns": sheet.column_count,
                }
                for sheet in sheets
            ],
            "materialization_allowed": pack_code is not None,
            "series_confirmation_required": profile.startswith("DHC8_") and confidence != "HIGH",
        },
    )
