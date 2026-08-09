from __future__ import annotations

import hashlib
import os
from datetime import datetime, timezone
from io import BytesIO
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.exc import DBAPIError
from sqlalchemy.orm import sessionmaker

from amodb.apps.accounts import models as account_models
from amodb.apps.aircraft_architecture.content_packs import (
    backend_models,
    backend_schemas,
    backend_services,
    governance,
    models,
    schemas,
    services,
)


DATABASE_URL = os.environ.get("POSTGRES_INTEGRATION_URL")
pytestmark = pytest.mark.skipif(
    not DATABASE_URL,
    reason="POSTGRES_INTEGRATION_URL is required for OEM backend acceptance tests",
)


@pytest.fixture(scope="module")
def engine():
    value = create_engine(DATABASE_URL, future=True, pool_pre_ping=True)
    try:
        yield value
    finally:
        value.dispose()


@pytest.fixture(scope="module")
def sessions(engine):
    return sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)


def _suffix() -> str:
    return uuid4().hex[:10]


def _users(db, suffix: str):
    amo = account_models.AMO(
        amo_code=f"O{suffix}"[:32],
        name=f"OEM Backend Acceptance {suffix}",
        login_slug=f"oem-backend-{suffix}"[:64],
        is_active=True,
    )
    db.add(amo)
    db.flush()
    superuser = account_models.User(
        amo_id=amo.id,
        staff_code=f"SUP-{suffix}"[:32],
        email=f"super-{suffix}@acceptance.invalid",
        first_name="OEM",
        last_name="Superuser",
        full_name="OEM Superuser",
        role=account_models.AccountRole.SUPERUSER,
        hashed_password="not-used",
        is_active=True,
        is_superuser=True,
        is_amo_admin=True,
        is_system_account=False,
    )
    admin = account_models.User(
        amo_id=amo.id,
        staff_code=f"ADM-{suffix}"[:32],
        email=f"admin-{suffix}@acceptance.invalid",
        first_name="OEM",
        last_name="Admin",
        full_name="OEM Admin",
        role=account_models.AccountRole.AMO_ADMIN,
        hashed_password="not-used",
        is_active=True,
        is_superuser=False,
        is_amo_admin=True,
        is_system_account=False,
    )
    db.add_all([superuser, admin])
    db.flush()
    return amo, superuser, admin


def _q400_workbook() -> bytes:
    workbook = Workbook()
    first = workbook.active
    first.title = "Section 1"
    core = ["Section 1", "Section 2", "Section 3", "Supplement 1"]
    for index, name in enumerate(core):
        sheet = first if index == 0 else workbook.create_sheet(name)
        sheet.append(["MPD TASK", "TASK DESCRIPTION", "INTERVAL", "EFFECTIVITY", "ATA"])
        sheet.append(
            [
                f"21{index + 1:02d}00-20{index + 1}-A-00",
                f"Controlled Q400 acceptance task {index + 1}",
                "8000 FH OR 72 MO",
                "MSN 4001 TO 4999",
                "21",
            ]
        )
    for name in ["Appendix A", "Appendix E", "Appendix F", "Appendix I", "Appendix J"]:
        sheet = workbook.create_sheet(name)
        sheet.append(["CODE", "DESCRIPTION"])
        sheet.append([f"{name}-1", f"Controlled supporting content for {name}"])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _fixture(db, suffix: str):
    amo, superuser, admin = _users(db, suffix)
    content = _q400_workbook()
    checksum = hashlib.sha256(content).hexdigest()
    publication = models.AircraftOemPublication(
        code=f"DHC8-400-MPD-{suffix}",
        manufacturer="De Havilland Canada",
        family="DHC-8",
        series="400",
        publication_code=f"PSM-1-84-7P-{suffix}",
        title="Q400 Maintenance Planning Document",
        publication_kind="MPD",
        status="ACTIVE",
        created_by_user_id=superuser.id,
    )
    db.add(publication)
    db.flush()
    revision = models.AircraftOemPublicationRevision(
        publication_id=publication.id,
        revision_code="52",
        status="CURRENT",
        effective_date=datetime.now(timezone.utc).date(),
        checksum_sha256=checksum,
        source_filename="q400-mpd.xlsx",
        storage_locator=f"controlled://oem/{suffix}/q400-mpd.xlsx",
        submitted_by_user_id=superuser.id,
        submitted_by_amo_id=amo.id,
        verified_by_user_id=superuser.id,
        verified_at=datetime.now(timezone.utc),
    )
    db.add(revision)
    db.commit()
    packs = services.bootstrap_source_intake_packs(db, user=superuser)
    pack = next(row for row in packs if row.code == "DHC8_400_MPD_SOURCE_INTAKE")
    return {
        "amo": amo,
        "superuser": superuser,
        "admin": admin,
        "publication": publication,
        "publication_revision": revision,
        "pack": pack,
        "content": content,
        "checksum": checksum,
    }


def test_oem_backend_database_contract_and_immutable_guards(engine):
    inspector = inspect(engine)
    tables = set(inspector.get_table_names())
    assert "aircraft_oem_source_intakes" in tables
    assert "aircraft_oem_source_intake_rows" in tables
    with engine.connect() as connection:
        tr_check_definition = connection.execute(
            text(
                """
                SELECT pg_get_constraintdef(oid)
                  FROM pg_constraint
                 WHERE conrelid = 'aircraft_oem_temporary_revisions'::regclass
                   AND contype = 'c'
                   AND pg_get_constraintdef(oid) LIKE '%CANDIDATE%'
                 LIMIT 1
                """
            )
        ).scalar_one()
        assert "CANDIDATE" in tr_check_definition
        assert "REJECTED" in tr_check_definition
        triggers = {
            row[0]
            for row in connection.execute(
                text(
                    """
                    SELECT tgname
                      FROM pg_trigger
                     WHERE NOT tgisinternal
                       AND tgrelid IN (
                         'aircraft_content_pack_revisions'::regclass,
                         'aircraft_content_pack_tasks'::regclass,
                         'aircraft_oem_publication_revisions'::regclass,
                         'aircraft_oem_temporary_revisions'::regclass,
                         'aircraft_oem_source_intakes'::regclass,
                         'aircraft_oem_source_intake_rows'::regclass
                       )
                    """
                )
            )
        }
    assert "trg_aircraft_content_pack_revision_oem_backend_controlled" in triggers
    assert "trg_aircraft_content_pack_tasks_controlled" in triggers
    assert "trg_aircraft_oem_publication_revision_controlled" in triggers
    assert "trg_aircraft_oem_temporary_revision_controlled" in triggers
    assert "trg_aircraft_oem_source_intake_controlled" in triggers
    assert "trg_aircraft_oem_source_intake_row_controlled" in triggers
    assert "trg_aircraft_oem_one_current_deferred" in triggers
    assert "trg_aircraft_content_one_published_deferred" in triggers


def test_full_q400_source_intake_materialization_and_publication_flow(sessions, engine):
    suffix = _suffix()
    with sessions() as db:
        fixture = _fixture(db, suffix)
        intake = backend_services.stage_intake(
            db,
            filename="q400-mpd.xlsx",
            content=fixture["content"],
            binding=backend_schemas.IntakeSourceBinding(
                publication_id=fixture["publication"].id,
                publication_revision_id=fixture["publication_revision"].id,
                pack_id=fixture["pack"].id,
                storage_locator=f"controlled://oem/{suffix}/q400-mpd.xlsx",
            ),
            user=fixture["admin"],
        )
        assert intake.status == "STAGED"
        assert intake.detected_profile == "DHC8_400_MPD_V1"
        assert len(intake.rows) == 9

        validation = backend_services.validate_intake(
            db,
            intake_id=intake.id,
            user=fixture["admin"],
        )
        assert validation.status == "VALIDATED"
        assert validation.task_rows == 4
        assert validation.resource_rows == 5
        assert validation.review_required_rows == 0
        assert validation.invalid_rows == 0

        approved = backend_services.approve_intake(
            db,
            intake_id=intake.id,
            payload=backend_schemas.OemSourceIntakeApproval(
                expected_normalization_hash=validation.normalization_hash,
                approval_note="Acceptance source reconciled against the controlled Q400 publication",
            ),
            user=fixture["superuser"],
        )
        assert approved.status == "APPROVED"

        content_revision = backend_services.materialize_intake(
            db,
            intake_id=intake.id,
            payload=backend_schemas.OemSourceIntakeMaterialize(
                revision_code=f"52-{suffix}",
                expected_normalization_hash=validation.normalization_hash,
                change_summary="Controlled Q400 acceptance baseline",
            ),
            user=fixture["superuser"],
        )
        assert content_revision.status == "DRAFT"
        assert len(content_revision.tasks) == 4
        assert len(content_revision.resources) == 5
        materialized_id = content_revision.id

        published = services.publish_revision(
            db,
            revision=content_revision,
            expected_content_hash=content_revision.content_hash,
            user=fixture["superuser"],
        )
        assert published.status == "PUBLISHED"
        assert published.pack.status == "ACTIVE"
        assert published.sources[0].publication_revision_id == fixture["publication_revision"].id
        task_id = published.tasks[0].id

    with pytest.raises(DBAPIError):
        with engine.begin() as connection:
            connection.execute(
                text(
                    "UPDATE aircraft_content_pack_tasks SET title = 'tampered' WHERE id = :id"
                ),
                {"id": task_id},
            )
    with pytest.raises(DBAPIError):
        with engine.begin() as connection:
            connection.execute(
                text("DELETE FROM aircraft_content_pack_revisions WHERE id = :id"),
                {"id": materialized_id},
            )


def test_temporary_revision_candidate_activation_and_base_revision_supersession(sessions):
    suffix = _suffix()
    with sessions() as db:
        fixture = _fixture(db, suffix)
        base = fixture["publication_revision"]
        candidate_tr = services.create_temporary_revision(
            db,
            publication_revision=base,
            payload=schemas.OemTemporaryRevisionCreate(
                temporary_revision_code=f"TR-{suffix}",
                checksum_sha256="b" * 64,
                source_filename=f"TR-{suffix}.pdf",
                storage_locator=f"controlled://oem/{suffix}/tr.pdf",
                change_summary="Acceptance Temporary Revision",
            ),
            user=fixture["admin"],
        )
        assert candidate_tr.status == "CANDIDATE"
        assert candidate_tr.verified_at is None

        active = governance.governed_temporary_revision_decision(
            db,
            temporary_revision=candidate_tr,
            payload=backend_schemas.OemTemporaryRevisionGovernanceDecision(
                action="ACTIVATE",
                decision_note="Verified against the controlled OEM source",
            ),
            user=fixture["superuser"],
        )
        assert active.status == "ACTIVE"
        assert active.verified_at is not None

        replacement = models.AircraftOemPublicationRevision(
            publication_id=fixture["publication"].id,
            revision_code="53",
            status="VERIFIED",
            effective_date=datetime.now(timezone.utc).date(),
            checksum_sha256="c" * 64,
            source_filename="q400-mpd-r53.xlsx",
            storage_locator=f"controlled://oem/{suffix}/q400-mpd-r53.xlsx",
            supersedes_revision_id=base.id,
            submitted_by_user_id=fixture["superuser"].id,
            submitted_by_amo_id=fixture["amo"].id,
            verified_by_user_id=fixture["superuser"].id,
            verified_at=datetime.now(timezone.utc),
        )
        db.add(replacement)
        db.commit()

        with pytest.raises(HTTPException, match="Active Temporary Revisions"):
            services.decide_oem_publication_revision(
                db,
                revision=replacement,
                payload=schemas.OemPublicationRevisionDecision(
                    action="MAKE_CURRENT",
                    decision_note="Attempt while active TR remains",
                ),
                user=fixture["superuser"],
            )
        db.rollback()

        governance.governed_temporary_revision_decision(
            db,
            temporary_revision=active,
            payload=backend_schemas.OemTemporaryRevisionGovernanceDecision(
                action="INCORPORATE",
                decision_note="TR incorporated into Revision 53",
            ),
            user=fixture["superuser"],
        )
        current = services.decide_oem_publication_revision(
            db,
            revision=replacement,
            payload=schemas.OemPublicationRevisionDecision(
                action="MAKE_CURRENT",
                decision_note="Revision 53 verified and incorporates active TR",
            ),
            user=fixture["superuser"],
        )
        assert current.status == "CURRENT"
        db.refresh(base)
        assert base.status == "SUPERSEDED"


def test_source_watch_change_detection_drives_currentness(sessions):
    suffix = _suffix()
    with sessions() as db:
        fixture = _fixture(db, suffix)
        watch = governance.create_governed_source_watch(
            db,
            publication=fixture["publication"],
            payload=backend_schemas.OemSourceWatchGovernanceCreate(
                channel_type="OEM_PORTAL",
                reference=f"controlled-oem-channel-{suffix}",
                check_interval_hours=24,
            ),
            user=fixture["superuser"],
        )
        checked = governance.record_governed_source_watch_check(
            db,
            watch=watch,
            payload=backend_schemas.OemSourceWatchGovernanceCheck(
                result_code="CHANGE_DETECTED",
                seen_marker="revision-53",
                detail="OEM portal exposes a newer source marker",
            ),
            user=fixture["superuser"],
        )
        governed = governance.governed_watch_read(checked)
        assert governed.last_result_code == "CHANGE_DETECTED"
        assert governed.consecutive_failures == 0
        currentness = governance.governed_publication_currentness(
            db,
            publication=fixture["publication"],
        )
        assert currentness.currentness_status == "SOURCE_CHANGE_DETECTED"
