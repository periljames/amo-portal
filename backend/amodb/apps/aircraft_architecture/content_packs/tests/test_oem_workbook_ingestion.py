from io import BytesIO

from openpyxl import Workbook

from amodb.apps.aircraft_architecture.content_packs import ingestion


def q400_sheet_names() -> list[str]:
    return [
        "Section 1",
        "Section 2",
        "Section 3",
        "Supplement 1",
        "Appendix A",
        "Appendix E",
        "Appendix F",
        "Appendix I",
        "Appendix J",
        "Appendix N",
        "Appendix P",
        "Appendix Q",
        "Appendix R",
        "Appendix T",
        "Appendix U",
    ]


def legacy_emp_sheet_names() -> list[str]:
    return ["Cross Reference", "L-Check", "Out of Phase", "Work Packages"]


def test_q400_profile_requires_core_sections_and_supporting_tables():
    profile, confidence, pack, warnings = ingestion.detect_profile(q400_sheet_names())
    assert profile == "DHC8_400_MPD_V1"
    assert confidence == "HIGH"
    assert pack == "DHC8_400_MPD_SOURCE_INTAKE"
    assert warnings == []


def test_series_100_200_300_legacy_mplm_filenames_map_to_separate_source_packs():
    expected = {
        "81MPLM.xls": ("DHC8_100_EMP_V1", "DHC8_100_MPD_SOURCE_INTAKE"),
        "82MPLM.xlsx": ("DHC8_200_EMP_V1", "DHC8_200_MPD_SOURCE_INTAKE"),
        "83MPLM.xls": ("DHC8_300_EMP_V1", "DHC8_300_MPD_SOURCE_INTAKE"),
    }
    for filename, (expected_profile, expected_pack) in expected.items():
        profile, confidence, pack, warnings = ingestion.detect_profile(
            legacy_emp_sheet_names(),
            filename=filename,
        )
        assert profile == expected_profile
        assert confidence == "MEDIUM"
        assert pack == expected_pack
        assert "must confirm series and revision" in warnings[0]


def test_legacy_emp_without_proven_series_is_preview_only():
    profile, confidence, pack, warnings = ingestion.detect_profile(
        legacy_emp_sheet_names(),
        filename="maintenance-data.xls",
    )
    assert profile == "DHC8_EMP_LEGACY_V1"
    assert confidence == "LOW"
    assert pack is None
    assert "preview-only" in warnings[0]


def test_unknown_workbook_is_preview_only():
    profile, confidence, pack, warnings = ingestion.detect_profile(["Status", "Hours"])
    assert profile == "UNMAPPED"
    assert confidence == "LOW"
    assert pack is None
    assert "must not materialize" in warnings[0]


def test_xlsx_preview_inventory_includes_hidden_sheets_without_executing_content():
    workbook = Workbook()
    first = workbook.active
    first.title = "Section 1"
    first.append(["MPD TASK", "INTERVAL"])
    first.append(["212500-201-A-00", "8000 FH"])
    for name in q400_sheet_names()[1:]:
        sheet = workbook.create_sheet(name)
        sheet.append([name, "controlled source"])
    workbook["Appendix U"].sheet_state = "hidden"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = ingestion.preview_oem_workbook(filename="MPD.xlsx", content=buffer.getvalue())

    assert result.detected_profile == "DHC8_400_MPD_V1"
    assert result.recommended_pack_code == "DHC8_400_MPD_SOURCE_INTAKE"
    assert len(result.checksum_sha256) == 64
    assert any(sheet.name == "Appendix U" and sheet.state == "HIDDEN" for sheet in result.sheets)
    assert any("Hidden workbook sheets" in warning for warning in result.warnings)
    assert result.source_manifest["materialization_allowed"] is True
    assert result.source_manifest["series_confirmation_required"] is False


def test_preview_renders_floating_cells_as_non_authoritative_strings():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Status"
    sheet.append(["Hours", 1250.25])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = ingestion.preview_oem_workbook(filename="source.xlsx", content=buffer.getvalue())

    assert result.sheets[0].sample_rows[0][1] == "1250.25"
    assert result.source_manifest["materialization_allowed"] is False
