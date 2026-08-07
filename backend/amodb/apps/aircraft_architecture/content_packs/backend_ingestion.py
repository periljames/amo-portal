from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from pydantic import ValidationError

from . import ingestion, normalization, schemas


MAX_CONTROLLED_ROWS = 250_000
MAX_CONTROLLED_COLUMNS = 80
CORE_TASK_SHEETS = {"section 1", "section 2", "section 3", "supplement 1"}
SUPPORTING_SHEETS = {
    "appendix a",
    "appendix e",
    "appendix f",
    "appendix i",
    "appendix j",
    "appendix n",
    "appendix p",
    "appendix q",
    "appendix r",
    "appendix t",
    "appendix u",
}

HEADER_ALIASES: dict[str, set[str]] = {
    "task_code": {
        "mpd task",
        "mpd task no",
        "mpd task number",
        "task",
        "task no",
        "task number",
        "task number configuration",
    },
    "description": {"description", "task description", "task description action", "action"},
    "ata": {"ata", "ata chapter", "ata no"},
    "task_type": {"task type", "type", "task code type"},
    "interval": {
        "interval",
        "maintenance interval",
        "task interval",
        "repeat interval",
        "interval requirement",
    },
    "threshold": {"threshold", "threshold interval", "t"},
    "initial": {"initial", "initial interval", "ii"},
    "repeat_cut_in": {"repeat cut in", "repeat cut-in", "rc"},
    "repeat": {"repeat", "repeat interval", "r"},
    "effectivity": {"effectivity", "applicability", "aircraft effectivity"},
    "source_requirement": {
        "source requirement",
        "source requirements",
        "mrb cmr",
        "mrb",
        "cmr",
        "mrm",
        "ali",
        "fec",
    },
    "task_card": {"task card", "task card no", "task card number"},
    "task_card_configuration": {"task card configuration", "configuration"},
    "amm": {"amm", "amm ref", "amm reference"},
    "zone": {"zone", "zones"},
    "panel": {"panel", "panels", "access panel", "access panels"},
    "general_reference": {"general reference", "general references", "reference", "references"},
    "skill": {"skill", "skill code"},
    "labour": {"labour hours", "labor hours", "man hours", "man-hours", "mh"},
    "persons": {"persons", "number of persons", "no of persons", "men"},
    "notes": {"notes", "note", "remarks", "programme notes", "program notes"},
}

UNIVERSAL_EFFECTIVITY = {
    "ALL",
    "ALL A/C",
    "ALL AC",
    "ALL AIRCRAFT",
    "ALL AIRPLANES",
    "ALL MSN",
}

MSN_RANGE_RE = re.compile(
    r"\bMSN\s*(?P<low>\d+)\s*(?:-|TO|THRU|THROUGH)\s*(?P<high>\d+)\b",
    re.IGNORECASE,
)
MSN_LIST_RE = re.compile(
    r"\bMSN\s*(?P<values>\d+(?:\s*[,/]\s*\d+)+)\b",
    re.IGNORECASE,
)
SB_RE = re.compile(
    r"\b(?P<state>PRE|POST)\s+SB\s*(?P<code>[A-Z0-9-]+)\b",
    re.IGNORECASE,
)
MODSUM_RE = re.compile(r"\bMODSUM\s*(?P<code>[A-Z0-9-]+)\b", re.IGNORECASE)
OPTION_RE = re.compile(r"\bOPTION\s*(?P<code>[A-Z0-9-]+)\b", re.IGNORECASE)


@dataclass(frozen=True)
class NormalizedCandidateRow:
    sheet_name: str
    row_number: int
    row_kind: str
    identity_key: str | None
    source_json: dict[str, Any]
    normalized_json: dict[str, Any]
    status: str
    issues: list[dict[str, Any]]

    @property
    def row_hash(self) -> str:
        payload = {
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "row_kind": self.row_kind,
            "source_json": self.source_json,
            "normalized_json": self.normalized_json,
            "status": self.status,
            "issues": self.issues,
        }
        encoded = json.dumps(
            payload,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
            default=str,
        ).encode("utf-8")
        return hashlib.sha256(encoded).hexdigest()


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _canonical_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return format(value, ".15g")
    if isinstance(value, Decimal):
        return format(value, "f")
    text = str(value).strip()
    return text if text else None


def _text(value: Any) -> str:
    canonical = _canonical_cell(value)
    return "" if canonical is None else str(canonical).strip()


def _split_controlled_list(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [
        part.strip()
        for part in re.split(r"[,;\n]+", text)
        if part.strip()
    ]


def _field_for_header(header: Any) -> str | None:
    normalized = _normalise_header(header)
    if not normalized:
        return None
    for field, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _header_mapping(row: Iterable[Any]) -> tuple[dict[int, str], int]:
    mapping: dict[int, str] = {}
    for index, value in enumerate(row):
        field = _field_for_header(value)
        if field and field not in mapping.values():
            mapping[index] = field
    score = len(set(mapping.values()))
    if "task_code" in mapping.values():
        score += 4
    if "description" in mapping.values():
        score += 3
    if any(field in mapping.values() for field in ("interval", "threshold", "repeat")):
        score += 2
    return mapping, score


def _find_task_header(rows: list[list[Any]]) -> tuple[int, dict[int, str]] | None:
    best: tuple[int, int, dict[int, str]] | None = None
    for index, row in enumerate(rows[:120], start=1):
        mapping, score = _header_mapping(row)
        fields = set(mapping.values())
        if "task_code" not in fields or "description" not in fields:
            continue
        if best is None or score > best[1]:
            best = (index, score, mapping)
    if best is None:
        return None
    return best[0], best[2]


def _row_dict(row: list[Any], mapping: dict[int, str]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for index, field in mapping.items():
        if index < len(row):
            value = _canonical_cell(row[index])
            if value not in (None, ""):
                result[field] = value
    return result


def _source_requirements(value: Any) -> list[dict[str, Any]]:
    text = _text(value)
    if not text:
        return []
    authorities = []
    for authority in ("MRB", "MRM", "CMR", "ALI"):
        if re.search(rf"\b{authority}\b", text, re.IGNORECASE):
            authorities.append({"authority": authority, "raw": text})
    return authorities or [{"authority": "OEM", "raw": text}]


def _sb_path(code: str) -> str:
    token = re.sub(r"[^a-z0-9]+", "_", code.casefold()).strip("_")
    return f"configuration.{token}"


def _effectivity_atom(text: str) -> tuple[dict[str, Any] | None, str]:
    working = " ".join(text.replace("\xa0", " ").split()).strip(" ,;")
    if not working or working.upper() in UNIVERSAL_EFFECTIVITY:
        return None, ""

    match = MSN_RANGE_RE.fullmatch(working)
    if match:
        return {
            "path": "aircraft.serial_number",
            "op": "between",
            "value": [int(match.group("low")), int(match.group("high"))],
            "label": "MSN",
        }, ""
    match = MSN_LIST_RE.fullmatch(working)
    if match:
        values = [int(value.strip()) for value in re.split(r"[,/]", match.group("values"))]
        return {
            "path": "aircraft.serial_number",
            "op": "in",
            "value": values,
            "label": "MSN",
        }, ""
    match = re.fullmatch(r"MSN\s*(\d+)", working, re.IGNORECASE)
    if match:
        return {
            "path": "aircraft.serial_number",
            "op": "eq",
            "value": int(match.group(1)),
            "label": "MSN",
        }, ""
    match = SB_RE.fullmatch(working)
    if match:
        code = match.group("code").upper()
        state = match.group("state").upper()
        return {
            "path": _sb_path(f"SB-{code}"),
            "op": "eq",
            "value": state,
            "label": f"SB {code}",
        }, ""
    match = MODSUM_RE.fullmatch(working)
    if match:
        code = match.group("code").upper()
        return {
            "path": "configuration.modsums",
            "op": "contains",
            "value": code,
            "label": f"Modsum {code}",
        }, ""
    match = OPTION_RE.fullmatch(working)
    if match:
        code = match.group("code").upper()
        return {
            "path": "configuration.options",
            "op": "contains",
            "value": code,
            "label": f"Option {code}",
        }, ""
    return None, working


def parse_effectivity_text(raw: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Parse only explicit, explainable Q400 effectivity forms.

    Unsupported wording is returned as a blocking review issue.  It is never
    silently converted to universal applicability.
    """

    text = " ".join((raw or "").replace("\xa0", " ").split()).strip()
    if not text or text.upper() in UNIVERSAL_EFFECTIVITY:
        return {}, []
    if "(" in text or ")" in text:
        return {}, [
            {
                "code": "EFFECTIVITY_REVIEW_REQUIRED",
                "message": "Parenthesized effectivity requires engineering mapping review",
                "raw": raw,
            }
        ]

    or_parts = [part.strip() for part in re.split(r"\s+OR\s+", text, flags=re.IGNORECASE)]
    any_groups: list[dict[str, Any]] = []
    unresolved: list[str] = []
    for or_part in or_parts:
        and_parts = [
            part.strip(" ,;")
            for part in re.split(r"\s+AND\s+|\s*;\s*", or_part, flags=re.IGNORECASE)
            if part.strip(" ,;")
        ]
        conditions: list[dict[str, Any]] = []
        for part in and_parts:
            atom, residue = _effectivity_atom(part)
            if atom:
                conditions.append(atom)
            elif residue:
                unresolved.append(residue)
        if conditions:
            any_groups.append(
                conditions[0]
                if len(conditions) == 1
                else {"operator": "ALL", "conditions": conditions}
            )

    if unresolved or not any_groups:
        return {}, [
            {
                "code": "EFFECTIVITY_REVIEW_REQUIRED",
                "message": "Effectivity wording could not be mapped without assumptions",
                "unresolved": unresolved or [text],
                "raw": raw,
            }
        ]
    expression = (
        any_groups[0]
        if len(any_groups) == 1
        else {"operator": "ANY", "conditions": any_groups}
    )
    return expression, []


def _interval_text(row: dict[str, Any]) -> str:
    direct = _text(row.get("interval"))
    if direct:
        return direct
    pieces: list[str] = []
    for field, prefix in (
        ("threshold", "T"),
        ("initial", "II"),
        ("repeat_cut_in", "RC"),
        ("repeat", "R"),
    ):
        value = _text(row.get(field))
        if value:
            pieces.append(f"{prefix} {value}")
    return "; ".join(pieces)


def _persons(value: Any) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        decimal = Decimal(text)
    except InvalidOperation:
        return None
    if decimal == decimal.to_integral_value() and decimal > 0:
        return int(decimal)
    return None


def _task_candidate(
    *,
    sheet_name: str,
    row_number: int,
    source: dict[str, Any],
    source_reference: str,
    source_revision: str,
    source_checksum_sha256: str,
) -> NormalizedCandidateRow:
    issues: list[dict[str, Any]] = []
    task_code = _text(source.get("task_code"))
    description = _text(source.get("description"))
    if not task_code or not description:
        return NormalizedCandidateRow(
            sheet_name=sheet_name,
            row_number=row_number,
            row_kind="UNMAPPED",
            identity_key=task_code or None,
            source_json=source,
            normalized_json={},
            status="REVIEW_REQUIRED",
            issues=[
                {
                    "code": "TASK_IDENTITY_REVIEW_REQUIRED",
                    "message": "Task code and description are required for controlled task materialization",
                }
            ],
        )

    raw_interval = _interval_text(source)
    intervals: dict[str, Any] = {}
    if raw_interval:
        try:
            intervals = normalization.parse_interval_text(raw_interval)
        except normalization.IntervalParseError as exc:
            issues.append(
                {
                    "code": "INTERVAL_REVIEW_REQUIRED",
                    "message": str(exc),
                    "raw": raw_interval,
                }
            )
    else:
        notes = _text(source.get("notes"))
        if "OPPORTUNITY" in notes.upper():
            intervals = {
                "schema": "MPD_INTERVAL_V1",
                "groups": [
                    {
                        "phase": "INTERVAL",
                        "mode": "OPPORTUNITY",
                        "reference": notes,
                    }
                ],
                "raw": notes,
            }
            raw_interval = notes
        else:
            issues.append(
                {
                    "code": "INTERVAL_REVIEW_REQUIRED",
                    "message": "No controlled interval was identified for the task row",
                }
            )

    raw_effectivity = _text(source.get("effectivity"))
    effectivity_expression, effectivity_issues = parse_effectivity_text(raw_effectivity)
    issues.extend(effectivity_issues)

    normalized: dict[str, Any] = {
        "task_code": task_code[:100],
        "title": description[:255],
        "description": description,
        "ata_chapter": _text(source.get("ata"))[:12] or None,
        "programme_section": sheet_name.upper().replace(" ", "_")[:40],
        "task_type": _text(source.get("task_type"))[:16] or None,
        "intervals_json": intervals,
        "raw_interval_text": raw_interval or None,
        "effectivity_expression_json": effectivity_expression,
        "raw_effectivity_text": raw_effectivity or None,
        "source_requirements_json": _source_requirements(source.get("source_requirement")),
        "task_card_number": _text(source.get("task_card"))[:120] or None,
        "task_card_configuration": _text(source.get("task_card_configuration"))[:120] or None,
        "amm_reference": _text(source.get("amm"))[:120] or None,
        "zones_json": _split_controlled_list(source.get("zone")),
        "panels_json": _split_controlled_list(source.get("panel")),
        "general_references_json": _split_controlled_list(source.get("general_reference")),
        "skill_code": _text(source.get("skill"))[:40] or None,
        "labour_hours": _text(source.get("labour"))[:24] or None,
        "number_of_persons": _persons(source.get("persons")),
        "program_notes_json": _split_controlled_list(source.get("notes")),
        "packaging_json": {},
        "source_page_ref": f"{sheet_name}!{row_number}"[:120],
        "source_reference": source_reference,
        "source_revision": source_revision,
        "source_checksum_sha256": source_checksum_sha256,
        "metadata_json": {
            "intake_sheet": sheet_name,
            "intake_row": row_number,
            "normalization_profile": "DHC8_400_MPD_V1",
        },
    }

    if not issues:
        try:
            validated = schemas.ContentTaskCreate.model_validate(normalized)
            normalized = validated.model_dump(mode="json")
        except ValidationError as exc:
            issues.append(
                {
                    "code": "TASK_SCHEMA_INVALID",
                    "message": "Normalized task does not satisfy the controlled task contract",
                    "detail": exc.errors(include_url=False),
                }
            )

    status = "VALID" if not issues else "REVIEW_REQUIRED"
    return NormalizedCandidateRow(
        sheet_name=sheet_name,
        row_number=row_number,
        row_kind="TASK",
        identity_key=task_code,
        source_json=source,
        normalized_json=normalized,
        status=status,
        issues=issues,
    )


def _resource_kind(sheet_name: str) -> str:
    return "Q400_" + re.sub(r"[^A-Z0-9]+", "_", sheet_name.upper()).strip("_")


def _resource_rows(
    *,
    sheet_name: str,
    rows: list[list[Any]],
    source_reference: str,
    source_revision: str,
    source_checksum_sha256: str,
) -> list[NormalizedCandidateRow]:
    if not rows:
        return []
    header_index = 0
    for index, row in enumerate(rows[:40]):
        labels = [_text(value) for value in row]
        nonempty = [label for label in labels if label]
        if len(nonempty) >= 2 and len(set(nonempty)) == len(nonempty):
            header_index = index
            break
    raw_headers = [_text(value) for value in rows[header_index]][:MAX_CONTROLLED_COLUMNS]
    headers: list[str] = []
    seen: set[str] = set()
    for index, header in enumerate(raw_headers, start=1):
        key = _normalise_header(header).replace(" ", "_") if header else f"col_{index}"
        if not key:
            key = f"col_{index}"
        base = key
        suffix = 2
        while key in seen:
            key = f"{base}_{suffix}"
            suffix += 1
        seen.add(key)
        headers.append(key)

    candidates: list[NormalizedCandidateRow] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = [_canonical_cell(value) for value in row[: len(headers)]]
        if not any(value not in (None, "") for value in values):
            continue
        payload = {
            headers[index]: value
            for index, value in enumerate(values)
            if value not in (None, "")
        }
        first_value = next((str(value) for value in values if value not in (None, "")), "")
        resource_code = f"{_resource_kind(sheet_name)}:{first_value or row_number}"[:140]
        normalized = schemas.ContentResourceCreate(
            resource_kind=_resource_kind(sheet_name)[:50],
            resource_code=resource_code,
            title=(first_value or f"{sheet_name} row {row_number}")[:255],
            payload_json=payload,
            source_reference=source_reference,
            source_revision=source_revision,
            source_checksum_sha256=source_checksum_sha256,
            source_page_ref=f"{sheet_name}!{row_number}"[:120],
            metadata_json={
                "intake_sheet": sheet_name,
                "intake_row": row_number,
                "normalization_profile": "DHC8_400_MPD_V1",
            },
        ).model_dump(mode="json")
        candidates.append(
            NormalizedCandidateRow(
                sheet_name=sheet_name,
                row_number=row_number,
                row_kind="RESOURCE",
                identity_key=resource_code,
                source_json=payload,
                normalized_json=normalized,
                status="VALID",
                issues=[],
            )
        )
    return candidates


def _xlsx_rows(content: bytes) -> dict[str, list[list[Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False, keep_links=False)
    result: dict[str, list[list[Any]]] = {}
    total = 0
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[Any]] = []
            max_col = min(max(int(worksheet.max_column or 1), 1), MAX_CONTROLLED_COLUMNS)
            for row in worksheet.iter_rows(values_only=True, max_col=max_col):
                rows.append(list(row))
                total += 1
                if total > MAX_CONTROLLED_ROWS:
                    raise ValueError("OEM workbook exceeds the controlled row limit")
            result[worksheet.title] = rows
    finally:
        workbook.close()
    return result


def _xls_rows(content: bytes) -> dict[str, list[list[Any]]]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    result: dict[str, list[list[Any]]] = {}
    total = 0
    try:
        for index, sheet_name in enumerate(workbook.sheet_names()):
            worksheet = workbook.sheet_by_index(index)
            rows: list[list[Any]] = []
            for row_index in range(worksheet.nrows):
                rows.append(
                    worksheet.row_values(
                        row_index,
                        start_colx=0,
                        end_colx=min(worksheet.ncols, MAX_CONTROLLED_COLUMNS),
                    )
                )
                total += 1
                if total > MAX_CONTROLLED_ROWS:
                    raise ValueError("OEM workbook exceeds the controlled row limit")
            result[sheet_name] = rows
            workbook.unload_sheet(index)
    finally:
        workbook.release_resources()
    return result


def workbook_rows(filename: str, content: bytes) -> dict[str, list[list[Any]]]:
    extension = Path(filename).suffix.lower()
    if extension == ".xls":
        return _xls_rows(content)
    if extension in {".xlsx", ".xlsm"}:
        return _xlsx_rows(content)
    raise ValueError("OEM workbook must be XLS, XLSX, or XLSM")


def normalize_oem_workbook(
    *,
    filename: str,
    content: bytes,
    source_reference: str,
    source_revision: str,
    source_checksum_sha256: str,
) -> tuple[ingestion.ingestion_schemas.OemWorkbookPreview, list[NormalizedCandidateRow]]:
    preview = ingestion.preview_oem_workbook(filename=filename, content=content)
    if preview.detected_profile != "DHC8_400_MPD_V1":
        raise ValueError(
            f"No governed materializer exists for profile {preview.detected_profile}; preview only"
        )

    sheets = workbook_rows(filename, content)
    candidates: list[NormalizedCandidateRow] = []
    for sheet_name, rows in sheets.items():
        normalized_name = " ".join(sheet_name.strip().casefold().split())
        if normalized_name in CORE_TASK_SHEETS:
            header = _find_task_header(rows)
            if not header:
                candidates.append(
                    NormalizedCandidateRow(
                        sheet_name=sheet_name,
                        row_number=1,
                        row_kind="UNMAPPED",
                        identity_key=None,
                        source_json={},
                        normalized_json={},
                        status="INVALID",
                        issues=[
                            {
                                "code": "CORE_SHEET_HEADER_UNRESOLVED",
                                "message": f"Controlled task header could not be identified in {sheet_name}",
                            }
                        ],
                    )
                )
                continue
            header_row, mapping = header
            for row_number, raw_row in enumerate(rows[header_row:], start=header_row + 1):
                source = _row_dict(raw_row, mapping)
                if not source:
                    continue
                if not _text(source.get("task_code")):
                    continue
                candidates.append(
                    _task_candidate(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        source=source,
                        source_reference=source_reference,
                        source_revision=source_revision,
                        source_checksum_sha256=source_checksum_sha256,
                    )
                )
        elif normalized_name in SUPPORTING_SHEETS:
            candidates.extend(
                _resource_rows(
                    sheet_name=sheet_name,
                    rows=rows,
                    source_reference=source_reference,
                    source_revision=source_revision,
                    source_checksum_sha256=source_checksum_sha256,
                )
            )

    if not any(row.row_kind == "TASK" for row in candidates):
        candidates.append(
            NormalizedCandidateRow(
                sheet_name="<workbook>",
                row_number=0,
                row_kind="UNMAPPED",
                identity_key=None,
                source_json={},
                normalized_json={},
                status="INVALID",
                issues=[
                    {
                        "code": "NO_CONTROLLED_TASKS_FOUND",
                        "message": "Q400 profile matched but no canonical task rows were materialized",
                    }
                ],
            )
        )
    return preview, candidates
