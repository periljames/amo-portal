from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook


SUPPORTED_DATASETS = {
    "AIRCRAFT_MASTER",
    "CONFIGURATION",
    "COMPONENTS",
    "LLP_STATUS",
    "UTILISATION",
    "AMP_STATUS",
    "AD_STATUS",
    "SB_STATUS",
    "MODIFICATIONS",
    "REPAIRS",
    "DEFERRALS",
    "MAINTENANCE_HISTORY",
    "DOCUMENT_INDEX",
}

DATASET_HINTS: dict[str, tuple[str, ...]] = {
    "AIRCRAFT_MASTER": ("aircraft", "master", "spec sheet", "info page"),
    "CONFIGURATION": ("configuration", "installed", "position", "assembly"),
    "COMPONENTS": ("component", "hard time", "safe life", "oc-cm", "llp"),
    "LLP_STATUS": ("life limit", "llp", "safe life", "hard time"),
    "UTILISATION": ("hours", "utilisation", "utilization", "flight log", "techlog"),
    "AMP_STATUS": ("amp", "smp", "inspection", "insp status", "maintenance status"),
    "AD_STATUS": ("ad status", "airworthiness directive", "ads"),
    "SB_STATUS": ("sb status", "service bulletin", "sbs"),
    "MODIFICATIONS": ("modification", "mods", "stc"),
    "REPAIRS": ("repair", "alteration", "damage"),
    "DEFERRALS": ("deferred", "deferral", "mel", "cddl"),
    "MAINTENANCE_HISTORY": ("logbook", "history", "accomplishment"),
    "DOCUMENT_INDEX": ("document", "index", "records"),
}

HEADER_HINTS: dict[str, tuple[str, ...]] = {
    "AIRCRAFT_MASTER": ("registration", "serial_number", "aircraft_model_code"),
    "CONFIGURATION": ("position", "part_number", "serial_number"),
    "COMPONENTS": ("part_number", "serial_number", "current_hours"),
    "LLP_STATUS": ("life_limit", "remaining", "part_number"),
    "UTILISATION": ("date", "techlog_no", "total_hours", "total_cycles"),
    "AMP_STATUS": ("task_code", "last_done", "next_due"),
    "AD_STATUS": ("ad_number", "compliance_status"),
    "SB_STATUS": ("sb_number", "compliance_status"),
    "MODIFICATIONS": ("modification", "stc", "embodied"),
    "REPAIRS": ("repair_reference", "description"),
    "DEFERRALS": ("deferral_reference", "due_date"),
    "MAINTENANCE_HISTORY": ("work_order", "accomplished_date"),
    "DOCUMENT_INDEX": ("document_type", "reference"),
}

MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_ROWS_PER_DATASET = 100_000


@dataclass(frozen=True)
class ParsedDataset:
    dataset: str
    source_name: str
    source_sheet: str | None
    headers: list[str]
    rows: list[dict[str, Any]]
    fingerprint: str


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    aliases = {
        "a_c_reg": "registration",
        "aircraft_registration": "registration",
        "registration_mark": "registration",
        "aircraft_serial": "serial_number",
        "manufacturer_serial_number": "serial_number",
        "msn": "serial_number",
        "tail": "registration",
        "tail_number": "registration",
        "model": "aircraft_model_code",
        "aircraft_model": "aircraft_model_code",
        "pn": "part_number",
        "p_n": "part_number",
        "sn": "serial_number",
        "s_n": "serial_number",
        "pos": "position",
        "airframe_hours": "total_hours",
        "ttaf": "total_hours",
        "airframe_cycles": "total_cycles",
        "tca": "total_cycles",
        "fh": "flight_hours",
        "fc": "flight_cycles",
        "tech_log": "techlog_no",
        "tech_log_no": "techlog_no",
        "task": "task_code",
        "task_number": "task_code",
        "ad": "ad_number",
        "sb": "sb_number",
    }
    return aliases.get(text, text)


def _clean_value(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def _row_from_values(headers: list[str], values: Iterable[Any]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for header, value in zip(headers, values):
        if not header:
            continue
        row[header] = _clean_value(value)
    return row


def fingerprint_dataset(source_system: str, dataset: str, headers: list[str], sheet_name: str | None = None) -> str:
    payload = {
        "source_system": (source_system or "GENERIC").strip().upper(),
        "dataset": dataset,
        "sheet": normalize_header(sheet_name or ""),
        "headers": sorted(set(headers)),
    }
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def classify_dataset(sheet_name: str | None, headers: list[str], requested: str | None = None) -> str:
    if requested:
        requested = requested.strip().upper()
        if requested not in SUPPORTED_DATASETS:
            raise HTTPException(status_code=400, detail=f"Unsupported dataset '{requested}'")
        return requested

    sheet = normalize_header(sheet_name or "").replace("_", " ")
    header_set = set(headers)
    scores: dict[str, int] = {code: 0 for code in SUPPORTED_DATASETS}
    for dataset, hints in DATASET_HINTS.items():
        scores[dataset] += sum(3 for hint in hints if hint in sheet)
    for dataset, hints in HEADER_HINTS.items():
        scores[dataset] += sum(2 for hint in hints if hint in header_set)

    winner = max(scores, key=scores.get)
    if scores[winner] <= 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={
                "code": "DATASET_CLASSIFICATION_REQUIRED",
                "message": "The source layout could not be classified. Select the dataset explicitly or save a mapping profile.",
                "sheet": sheet_name,
                "headers": headers,
            },
        )
    return winner


def _parse_csv(payload: bytes, filename: str, source_system: str, requested_dataset: str | None) -> list[ParsedDataset]:
    text = payload.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    raw_rows = list(reader)
    if not raw_rows:
        return []
    headers = [normalize_header(value) for value in raw_rows[0]]
    rows = [_row_from_values(headers, values) for values in raw_rows[1:] if any(value not in (None, "") for value in values)]
    dataset = classify_dataset(filename, headers, requested_dataset)
    return [
        ParsedDataset(
            dataset=dataset,
            source_name=filename,
            source_sheet=None,
            headers=headers,
            rows=rows[:MAX_ROWS_PER_DATASET],
            fingerprint=fingerprint_dataset(source_system, dataset, headers),
        )
    ]


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int:
    best_index = 0
    best_score = -1
    all_hints = {hint for hints in HEADER_HINTS.values() for hint in hints}
    for index, row in enumerate(rows[:25]):
        headers = [normalize_header(value) for value in row]
        score = sum(2 for value in headers if value in all_hints) + sum(1 for value in headers if value)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def _parse_excel(payload: bytes, filename: str, source_system: str, requested_dataset: str | None) -> list[ParsedDataset]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    datasets: list[ParsedDataset] = []
    for worksheet in workbook.worksheets:
        raw_rows = list(worksheet.iter_rows(values_only=True))
        if not raw_rows:
            continue
        header_index = _detect_header_row(raw_rows)
        headers = [normalize_header(value) for value in raw_rows[header_index]]
        if not any(headers):
            continue
        rows = [
            _row_from_values(headers, values)
            for values in raw_rows[header_index + 1 : header_index + 1 + MAX_ROWS_PER_DATASET]
            if any(value not in (None, "") for value in values)
        ]
        if not rows:
            continue
        try:
            dataset = classify_dataset(worksheet.title, headers, requested_dataset)
        except HTTPException:
            if requested_dataset:
                raise
            continue
        datasets.append(
            ParsedDataset(
                dataset=dataset,
                source_name=filename,
                source_sheet=worksheet.title,
                headers=headers,
                rows=rows,
                fingerprint=fingerprint_dataset(source_system, dataset, headers, worksheet.title),
            )
        )
    return datasets


async def parse_upload(file: UploadFile, source_system: str, requested_dataset: str | None = None) -> list[ParsedDataset]:
    filename = file.filename or "upload"
    payload = await file.read()
    if len(payload) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Upload exceeds 50 MB")
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        datasets = _parse_csv(payload, filename, source_system, requested_dataset)
    elif suffix in {".xlsx", ".xlsm"}:
        datasets = _parse_excel(payload, filename, source_system, requested_dataset)
    else:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Universal induction accepts CSV, XLSX, and XLSM source files. Binary XLSB files must be exported to XLSX or provided through a source-system adapter.",
        )
    if not datasets:
        raise HTTPException(status_code=422, detail="No classifiable induction datasets were found in the upload")
    return datasets
