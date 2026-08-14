from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


SEMANTIC_HEX = {
    "DUTY": "DCE6F1",
    "STANDBY": "FCE4D6",
    "TRAINING": "E4DFEC",
    "REST": "E2F0D9",
    "OFF": "E7E6E6",
    "LEAVE": "FFF2CC",
    "SICK": "F4CCCC",
    "OTHER": "D9EAD3",
}


def _parse(value: str) -> datetime:
    return datetime.fromisoformat(str(value).replace("Z", "+00:00"))


def _dates(snapshot: dict[str, Any]) -> list[date]:
    period = snapshot["period"]
    start = date.fromisoformat(period["starts_on"])
    end = date.fromisoformat(period["ends_on"])
    result: list[date] = []
    cursor = start
    while cursor <= end:
        result.append(cursor)
        cursor += timedelta(days=1)
    return result


def _assignment_dates(row: dict[str, Any]) -> list[date]:
    start = _parse(row["starts_at"])
    end = _parse(row["ends_at"])
    final = (end - timedelta(microseconds=1)).date() if end > start else start.date()
    cursor = start.date()
    values: list[date] = []
    while cursor <= final:
        values.append(cursor)
        cursor += timedelta(days=1)
    return values


def _cell_label(row: dict[str, Any]) -> str:
    shift = str(row.get("shift_code") or row.get("status") or "").strip()
    aircraft = str(row.get("aircraft_display_codes") or "").strip()
    return "\n".join(value for value in (shift, aircraft) if value)


def _semantic_by_code(snapshot: dict[str, Any]) -> dict[str, str]:
    return {
        str(item.get("code") or ""): str(item.get("duty_semantic") or "OTHER")
        for item in snapshot.get("legend", [])
    }


def _matrix(snapshot: dict[str, Any]) -> tuple[list[tuple[str, str, str, str]], dict[tuple[str, str, str, str], dict[date, list[dict[str, Any]]]]]:
    grouped: dict[tuple[str, str, str, str], dict[date, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for row in snapshot.get("rows", []):
        person = (
            str(row.get("staff_code") or ""),
            str(row.get("full_name") or ""),
            str(row.get("department_code") or ""),
            str(row.get("base_code") or ""),
        )
        for duty_date in _assignment_dates(row):
            grouped[person][duty_date].append(row)
    people = sorted(grouped, key=lambda item: (item[1].casefold(), item[0].casefold()))
    return people, grouped


def _document_title(snapshot: dict[str, Any]) -> str:
    period = snapshot["period"]
    return f"Duty Roster — {period['name']}"


def _revision_text(snapshot: dict[str, Any]) -> str:
    doc = snapshot["document"]
    parts = [f"Form {doc.get('form_number') or 'ROSTER'}"]
    if doc.get("revision_label"):
        parts.append(f"Rev {doc['revision_label']}")
    if doc.get("revision_date"):
        parts.append(f"Rev date {doc['revision_date']}")
    parts.append(f"Roster v{snapshot.get('version_no')}")
    parts.append(str(snapshot.get("status") or "DRAFT"))
    return " | ".join(parts)


def controlled_roster_xlsx(snapshot: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Controlled Roster"
    sheet.sheet_view.showGridLines = False
    days = _dates(snapshot)
    people, grouped = _matrix(snapshot)
    semantic_by_code = _semantic_by_code(snapshot)
    last_column = 4 + len(days)
    last_letter = get_column_letter(last_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(1, 1, _document_title(snapshot))
    sheet.cell(1, 1).font = Font(bold=True, size=16)
    sheet.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet.cell(2, 1, _revision_text(snapshot))
    sheet.cell(2, 1).font = Font(bold=True, size=9)
    sheet.cell(2, 1).alignment = Alignment(horizontal="center")

    doc = snapshot["document"]
    signoff = " | ".join(filter(None, [
        f"{doc.get('prepared_by_label') or 'Prepared by'}: {doc.get('prepared_by') or '—'}" + (f" ({doc.get('prepared_date')})" if doc.get('prepared_date') else ""),
        f"{doc.get('approved_by_label') or 'Approved by'}: {doc.get('approved_by') or '—'}" + (f" ({doc.get('approved_date')})" if doc.get('approved_date') else ""),
    ]))
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    sheet.cell(3, 1, signoff)
    sheet.cell(3, 1).alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.cell(3, 1).font = Font(size=9)

    if snapshot.get("status") != "PUBLISHED":
        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_column)
        sheet.cell(4, 1, "DRAFT — NOT A CONTROLLED PUBLISHED ROSTER")
        sheet.cell(4, 1).font = Font(bold=True, size=11)
        sheet.cell(4, 1).alignment = Alignment(horizontal="center")

    fixed = ["Staff Code", "Employee", "Department", "Base"]
    for column, label in enumerate(fixed, start=1):
        sheet.cell(5, column, label)
        sheet.merge_cells(start_row=5, start_column=column, end_row=6, end_column=column)
    for index, duty_date in enumerate(days, start=5):
        sheet.cell(5, index, duty_date.strftime("%a"))
        sheet.cell(6, index, duty_date.day)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    for row_number in (5, 6):
        for cell in sheet[row_number]:
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

    start_row = 7
    for offset, person in enumerate(people):
        row_number = start_row + offset
        for column, value in enumerate(person, start=1):
            sheet.cell(row_number, column, value)
            sheet.cell(row_number, column).alignment = Alignment(vertical="center", wrap_text=True)
        for day_index, duty_date in enumerate(days, start=5):
            assignments = grouped[person].get(duty_date, [])
            labels = list(dict.fromkeys(_cell_label(item) for item in assignments if _cell_label(item)))
            cell = sheet.cell(row_number, day_index, "\n".join(labels))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=8)
            if assignments:
                semantic = semantic_by_code.get(str(assignments[0].get("shift_code") or ""), "OTHER")
                cell.fill = PatternFill(fill_type="solid", fgColor=SEMANTIC_HEX.get(semantic, SEMANTIC_HEX["OTHER"]))
        sheet.row_dimensions[row_number].height = 30

    legend_start = start_row + len(people) + 2
    sheet.cell(legend_start, 1, "Roster code legend")
    sheet.cell(legend_start, 1).font = Font(bold=True, size=10)
    legend_headers = ["Code", "Meaning", "Default time", "Unpaid break", "Semantic", "Verification"]
    for column, label in enumerate(legend_headers, start=1):
        sheet.cell(legend_start + 1, column, label)
        sheet.cell(legend_start + 1, column).font = Font(bold=True, size=8)
        sheet.cell(legend_start + 1, column).fill = header_fill
    for offset, item in enumerate(snapshot.get("legend", []), start=2):
        time_label = ""
        if item.get("default_start_time") or item.get("default_end_time"):
            time_label = f"{item.get('default_start_time') or '—'}–{item.get('default_end_time') or '—'}"
        values = [
            item.get("code"),
            item.get("label"),
            time_label,
            f"{int(item.get('unpaid_break_minutes') or 0)} min",
            item.get("duty_semantic"),
            item.get("verification_status"),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(legend_start + offset, column, value)
            sheet.cell(legend_start + offset, column).alignment = Alignment(vertical="top", wrap_text=True)
        semantic = str(item.get("duty_semantic") or "OTHER")
        sheet.cell(legend_start + offset, 1).fill = PatternFill(fill_type="solid", fgColor=SEMANTIC_HEX.get(semantic, SEMANTIC_HEX["OTHER"]))

    footer_row = legend_start + len(snapshot.get("legend", [])) + 4
    if doc.get("footer_note"):
        sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=last_column)
        sheet.cell(footer_row, 1, doc["footer_note"])
        sheet.cell(footer_row, 1).alignment = Alignment(wrap_text=True)
        sheet.cell(footer_row, 1).font = Font(italic=True, size=8)
    if snapshot.get("legacy_reconstructed"):
        sheet.merge_cells(start_row=footer_row + 1, start_column=1, end_row=footer_row + 1, end_column=last_column)
        sheet.cell(footer_row + 1, 1, "Historical note: this snapshot was reconstructed from a roster published before controlled snapshot capture was enabled.")
        sheet.cell(footer_row + 1, 1).font = Font(italic=True, size=8)

    widths = {"A": 12, "B": 24, "C": 14, "D": 12}
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    for column in range(5, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 5.2

    sheet.freeze_panes = "E7"
    sheet.print_title_rows = "1:6"
    sheet.print_area = f"A1:{last_letter}{max(sheet.max_row, footer_row + 1)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3 if doc.get("page_size") == "A3" else sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = f"{doc.get('form_number') or 'ROSTER'} | Roster v{snapshot.get('version_no')} | Page &P of &N"
    if snapshot.get("status") != "PUBLISHED":
        sheet.oddHeader.center.text = "DRAFT — NOT CONTROLLED"

    detail = workbook.create_sheet("Assignment Detail")
    detail_headers = [
        "Staff Code", "Employee", "Department", "Base", "Shift", "Status",
        "Start", "End", "Planned Minutes", "Role", "Aircraft", "Aircraft Display", "Calendar Lineage",
    ]
    detail.append(detail_headers)
    for cell in detail[1]:
        cell.font = Font(bold=True)
    for row in snapshot.get("rows", []):
        detail.append([
            row.get("staff_code"), row.get("full_name"), row.get("department_code"), row.get("base_code"),
            row.get("shift_code"), row.get("status"), row.get("starts_at"), row.get("ends_at"),
            row.get("planned_minutes"), row.get("role_label"), row.get("aircraft_registrations"),
            row.get("aircraft_display_codes"), row.get("calendar_lineage"),
        ])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def controlled_roster_pdf(snapshot: dict[str, Any]) -> bytes:
    doc_meta = snapshot["document"]
    page = landscape(A3 if doc_meta.get("page_size") == "A3" else A4)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=page,
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=9 * mm,
        bottomMargin=10 * mm,
        title=_document_title(snapshot),
    )
    styles = getSampleStyleSheet()
    days = _dates(snapshot)
    people, grouped = _matrix(snapshot)
    semantic_by_code = _semantic_by_code(snapshot)

    story: list[Any] = [
        Paragraph(_document_title(snapshot), styles["Title"]),
        Paragraph(_revision_text(snapshot), styles["Normal"]),
        Spacer(1, 3 * mm),
    ]
    signoff_data = [[
        f"{doc_meta.get('prepared_by_label') or 'Prepared by'}: {doc_meta.get('prepared_by') or '—'}",
        f"Date: {doc_meta.get('prepared_date') or '—'}",
        f"{doc_meta.get('approved_by_label') or 'Approved by'}: {doc_meta.get('approved_by') or '—'}",
        f"Date: {doc_meta.get('approved_date') or '—'}",
    ]]
    signoff_table = Table(signoff_data, colWidths=[65 * mm, 32 * mm, 65 * mm, 32 * mm])
    signoff_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#94A3B8")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
    ]))
    story.extend([signoff_table, Spacer(1, 3 * mm)])

    table_data: list[list[Any]] = [["Staff", "Employee", "Dept", "Base", *[day.strftime("%d\n%a") for day in days]]]
    for person in people:
        values: list[Any] = [*person]
        for duty_date in days:
            labels = list(dict.fromkeys(_cell_label(item) for item in grouped[person].get(duty_date, []) if _cell_label(item)))
            values.append("\n".join(labels))
        table_data.append(values)

    page_width = page[0] - 14 * mm
    fixed_widths = [18 * mm, 45 * mm, 18 * mm, 16 * mm]
    remaining = max(page_width - sum(fixed_widths), 120 * mm)
    day_width = remaining / max(len(days), 1)
    table = Table(table_data, repeatRows=1, colWidths=[*fixed_widths, *([day_width] * len(days))])
    commands: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 6.5),
        ("ALIGN", (4, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (3, -1), 6.5),
        ("FONTSIZE", (4, 1), (-1, -1), 6),
        ("FONTNAME", (4, 1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94A3B8")),
        ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]
    for row_index, person in enumerate(people, start=1):
        for day_offset, duty_date in enumerate(days, start=4):
            assignments = grouped[person].get(duty_date, [])
            if not assignments:
                continue
            semantic = semantic_by_code.get(str(assignments[0].get("shift_code") or ""), "OTHER")
            commands.append(("BACKGROUND", (day_offset, row_index), (day_offset, row_index), colors.HexColor("#" + SEMANTIC_HEX.get(semantic, SEMANTIC_HEX["OTHER"]))))
    table.setStyle(TableStyle(commands))
    story.extend([table, Spacer(1, 4 * mm)])

    legend_rows = [["Code", "Meaning", "Time", "Break", "Semantic", "Status"]]
    for item in snapshot.get("legend", []):
        time_label = ""
        if item.get("default_start_time") or item.get("default_end_time"):
            time_label = f"{item.get('default_start_time') or '—'}–{item.get('default_end_time') or '—'}"
        legend_rows.append([
            item.get("code") or "",
            item.get("label") or "",
            time_label,
            f"{int(item.get('unpaid_break_minutes') or 0)} min",
            item.get("duty_semantic") or "",
            item.get("verification_status") or "",
        ])
    legend = Table(legend_rows, repeatRows=1, colWidths=[15 * mm, 55 * mm, 30 * mm, 22 * mm, 28 * mm, 32 * mm])
    legend.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#94A3B8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    legend_blocks: list[Any] = [Paragraph("Roster code legend", styles["Heading3"]), legend]
    if doc_meta.get("footer_note"):
        legend_blocks.extend([Spacer(1, 2 * mm), Paragraph(str(doc_meta["footer_note"]), styles["Normal"])])
    if snapshot.get("legacy_reconstructed"):
        legend_blocks.extend([Spacer(1, 2 * mm), Paragraph("Historical note: this snapshot was reconstructed from a roster published before controlled snapshot capture was enabled.", styles["Italic"])])
    story.append(KeepTogether(legend_blocks))

    def decorate(canvas: Canvas, report_doc) -> None:
        canvas.saveState()
        width, height = page
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(width / 2, 4 * mm, f"{doc_meta.get('form_number') or 'ROSTER'} | Roster v{snapshot.get('version_no')} | Page {canvas.getPageNumber()}")
        if snapshot.get("status") != "PUBLISHED":
            canvas.setFillColor(colors.Color(0.55, 0.55, 0.55, alpha=0.18))
            canvas.setFont("Helvetica-Bold", 54)
            canvas.translate(width / 2, height / 2)
            canvas.rotate(30)
            canvas.drawCentredString(0, 0, "DRAFT — NOT CONTROLLED")
        canvas.restoreState()

    document.build(story, onFirstPage=decorate, onLaterPages=decorate)
    return output.getvalue()
