from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from amodb.apps.rostering import exports


def _row(**overrides):
    row = {
        "assignment_id": "assignment-1",
        "version_id": "version-1",
        "user_id": "user-1",
        "staff_code": "ENG001",
        "full_name": "Amina Engineer",
        "department_code": "MAINT",
        "base_code": "NBO",
        "shift_code": "DAY",
        "status": "DUTY",
        "source": "MANUAL",
        "starts_at": "2026-07-23T08:00:00+00:00",
        "ends_at": "2026-07-23T17:00:00+00:00",
        "planned_minutes": 540,
        "role_label": "Certifying engineer",
        "team_code": "A",
        "location_label": "Hangar 1",
        "linked_task_count": 0,
        "linked_task_hours": 0,
        "change_reason": "Planner assignment",
    }
    row.update(overrides)
    return row


def test_xlsx_contains_week_grouped_person_by_day_matrix_and_detail_sheet():
    rows = [
        _row(),
        _row(
            assignment_id="assignment-2",
            shift_code="NIGHT",
            starts_at="2026-07-24T20:00:00+00:00",
            ends_at="2026-07-25T06:00:00+00:00",
            role_label="Night coverage",
        ),
    ]

    workbook = load_workbook(BytesIO(exports.assignment_xlsx(rows)))

    assert workbook.sheetnames == ["Duty Matrix", "Assignment Detail"]
    matrix = workbook["Duty Matrix"]
    assert matrix["A1"].value == "Staff Code"
    assert str(matrix["E1"].value).startswith("ISO Week")
    assert matrix["A3"].value == "ENG001"
    assert matrix["B3"].value == "Amina Engineer"
    assert matrix["E2"].value.date() == date(2026, 7, 23)
    assert "DAY" in matrix["E3"].value
    assert "NIGHT" in matrix["F3"].value
    assert "NIGHT" in matrix["G3"].value

    detail = workbook["Assignment Detail"]
    assert detail.max_row == 3
    assert detail["A2"].value == "assignment-1"
    assert detail["A3"].value == "assignment-2"


def test_xlsx_empty_export_remains_openable_and_explains_no_assignments():
    workbook = load_workbook(BytesIO(exports.assignment_xlsx([])))

    matrix = workbook["Duty Matrix"]
    assert matrix["A3"].value == "No published roster assignments in the selected range."
    assert workbook["Assignment Detail"].max_row == 1
