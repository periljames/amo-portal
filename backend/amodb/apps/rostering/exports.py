# backend/amodb/apps/rostering/exports.py
from __future__ import annotations

import csv
import io
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import schemas

UTC = timezone.utc


ASSIGNMENT_COLUMNS = [
    "assignment_id",
    "version_id",
    "staff_code",
    "full_name",
    "department_code",
    "base_code",
    "shift_code",
    "status",
    "source",
    "starts_at",
    "ends_at",
    "planned_minutes",
    "role_label",
    "team_code",
    "location_label",
    "linked_task_count",
    "linked_task_hours",
    "change_reason",
]


def assignment_csv(rows: Sequence[dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=ASSIGNMENT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def _parse_datetime(value: Any) -> datetime:
    parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=UTC)


def _assignment_dates(row: dict[str, Any]) -> list[date]:
    starts_at = _parse_datetime(row["starts_at"])
    ends_at = _parse_datetime(row["ends_at"])
    if ends_at <= starts_at:
        return [starts_at.date()]
    last_inclusive = ends_at - timedelta(microseconds=1)
    cursor = starts_at.date()
    final = last_inclusive.date()
    days: list[date] = []
    while cursor <= final:
        days.append(cursor)
        cursor += timedelta(days=1)
    return days


def _matrix_label(row: dict[str, Any]) -> str:
    shift = str(row.get("shift_code") or row.get("status") or "Duty")
    starts_at = _parse_datetime(row["starts_at"])
    ends_at = _parse_datetime(row["ends_at"])
    time_label = f"{starts_at:%H:%M}-{ends_at:%H:%M}"
    role = str(row.get("role_label") or "").strip()
    return " · ".join(value for value in (shift, time_label, role) if value)


def _build_matrix_sheet(workbook: Workbook, rows: Sequence[dict[str, Any]]) -> None:
    sheet = workbook.active
    sheet.title = "Duty Matrix"
    sheet.freeze_panes = "E3"
    sheet.sheet_view.showGridLines = False

    all_dates = sorted({day for row in rows for day in _assignment_dates(row)})
    fixed_headers = ["Staff Code", "Employee", "Department", "Base"]
    for column, heading in enumerate(fixed_headers, start=1):
        sheet.cell(row=1, column=column, value=heading)
        sheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)

    week_groups: list[tuple[int, int, str]] = []
    group_start = 5
    current_week: tuple[int, int] | None = None
    for offset, day in enumerate(all_dates, start=5):
        iso_year, iso_week, _ = day.isocalendar()
        week_key = (iso_year, iso_week)
        if current_week is None:
            current_week = week_key
            group_start = offset
        elif week_key != current_week:
            week_groups.append((group_start, offset - 1, f"ISO Week {current_week[1]} · {current_week[0]}"))
            current_week = week_key
            group_start = offset
        sheet.cell(row=2, column=offset, value=day)
        sheet.cell(row=2, column=offset).number_format = "ddd dd mmm"
    if current_week is not None:
        week_groups.append((group_start, 4 + len(all_dates), f"ISO Week {current_week[1]} · {current_week[0]}"))

    for start_column, end_column, title in week_groups:
        sheet.cell(row=1, column=start_column, value=title)
        if end_column > start_column:
            sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    week_fill = PatternFill(fill_type="solid", fgColor="B4C6E7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = week_fill if cell.column >= 5 else header_fill
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90 if cell.column >= 5 else 0)
        cell.fill = header_fill

    grouped: dict[tuple[str, str, str, str], dict[date, list[str]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        person_key = (
            str(row.get("staff_code") or ""),
            str(row.get("full_name") or row.get("user_id") or ""),
            str(row.get("department_code") or ""),
            str(row.get("base_code") or ""),
        )
        label = _matrix_label(row)
        for day in _assignment_dates(row):
            grouped[person_key][day].append(label)

    for person_key in sorted(grouped, key=lambda item: (item[1].casefold(), item[0].casefold())):
        day_values = grouped[person_key]
        output_row: list[Any] = [*person_key]
        for day in all_dates:
            output_row.append("\n".join(dict.fromkeys(day_values.get(day, []))))
        sheet.append(output_row)

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 14
    for column in range(5, 5 + len(all_dates)):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.auto_filter.ref = f"A2:{get_column_letter(max(sheet.max_column, 4))}{max(sheet.max_row, 2)}"
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 70

    if not rows:
        sheet.cell(row=3, column=1, value="No published roster assignments in the selected range.")
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)


def _build_detail_sheet(workbook: Workbook, rows: Sequence[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Assignment Detail")
    sheet.freeze_panes = "A2"
    sheet.append([column.replace("_", " ").title() for column in ASSIGNMENT_COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        sheet.append([row.get(column) for column in ASSIGNMENT_COLUMNS])
    for index, _column in enumerate(ASSIGNMENT_COLUMNS, start=1):
        values = [str(sheet.cell(row=row_number, column=index).value or "") for row_number in range(1, min(sheet.max_row, 250) + 1)]
        width = min(max(max((len(value) for value in values), default=10) + 2, 12), 34)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = sheet.dimensions


def assignment_xlsx(rows: Sequence[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    _build_matrix_sheet(workbook, rows)
    _build_detail_sheet(workbook, rows)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def assignment_pdf(rows: Sequence[dict[str, Any]], *, title: str, subtitle: str) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Paragraph(subtitle, styles["Normal"]), Spacer(1, 5 * mm)]
    columns = ["staff_code", "full_name", "department_code", "base_code", "shift_code", "status", "starts_at", "ends_at", "planned_minutes", "role_label"]
    data = [[column.replace("_", " ").title() for column in columns]]
    for row in rows:
        data.append([str(row.get(column) or "") for column in columns])
    table = Table(data, repeatRows=1, colWidths=[20 * mm, 34 * mm, 23 * mm, 18 * mm, 18 * mm, 18 * mm, 34 * mm, 34 * mm, 20 * mm, 30 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94A3B8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    document.build(story)
    return output.getvalue()


def report_csv(summary: schemas.RosterReportSummary) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    for field in (
        "from_date",
        "to_date",
        "planned_minutes",
        "attendance_minutes",
        "productive_minutes",
        "overtime_minutes",
        "assignment_count",
        "assigned_people",
        "leave_minutes",
        "training_minutes",
        "standby_minutes",
        "acknowledgement_rate",
        "blocker_count",
        "warning_count",
    ):
        writer.writerow([field, getattr(summary, field)])
    writer.writerow([])
    writer.writerow(["By base"])
    if summary.by_base:
        keys = list(summary.by_base[0].keys())
        writer.writerow(keys)
        writer.writerows([[row.get(key) for key in keys] for row in summary.by_base])
    writer.writerow([])
    writer.writerow(["By department"])
    if summary.by_department:
        keys = list(summary.by_department[0].keys())
        writer.writerow(keys)
        writer.writerows([[row.get(key) for key in keys] for row in summary.by_department])
    writer.writerow([])
    writer.writerow(["By person"])
    if summary.by_user:
        keys = list(summary.by_user[0].keys())
        writer.writerow(keys)
        writer.writerows([[row.get(key) for key in keys] for row in summary.by_user])
    return output.getvalue()


def _ics_escape(value: Any) -> str:
    text = str(value or "")
    return text.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def _ics_datetime(value: datetime) -> str:
    aware = value if value.tzinfo is not None else value.replace(tzinfo=UTC)
    return aware.astimezone(UTC).strftime("%Y%m%dT%H%M%SZ")


def assignment_ics(rows: Sequence[dict[str, Any]], *, calendar_name: str = "AMO Duty Roster") -> str:
    generated = datetime.now(UTC)
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//AMO Portal//Duty Rostering//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{_ics_escape(calendar_name)}",
    ]
    for row in rows:
        starts_at = datetime.fromisoformat(str(row["starts_at"]))
        ends_at = datetime.fromisoformat(str(row["ends_at"]))
        summary = f"{row.get('shift_code') or row.get('status') or 'Duty'} · {row.get('base_code') or 'Base unassigned'}"
        description = "\n".join(filter(None, [
            f"Status: {row.get('status')}",
            f"Role: {row.get('role_label')}" if row.get("role_label") else None,
            f"Team: {row.get('team_code')}" if row.get("team_code") else None,
            f"Location: {row.get('location_label')}" if row.get("location_label") else None,
        ]))
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{_ics_escape(row['assignment_id'])}@amo-portal",
            f"DTSTAMP:{_ics_datetime(generated)}",
            f"DTSTART:{_ics_datetime(starts_at)}",
            f"DTEND:{_ics_datetime(ends_at)}",
            f"SUMMARY:{_ics_escape(summary)}",
            f"DESCRIPTION:{_ics_escape(description)}",
            f"LOCATION:{_ics_escape(row.get('location_label') or row.get('base_code') or '')}",
            "STATUS:CONFIRMED",
            "END:VEVENT",
        ])
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"
