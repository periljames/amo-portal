"""Macro-disabled, hash-retained import of Reliability workbook source registers."""
from __future__ import annotations
import hashlib,json,re,zipfile
from collections import Counter
from datetime import date,datetime
from decimal import Decimal,InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from fastapi import APIRouter,Depends,File,Form,HTTPException,UploadFile
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from amodb.apps.accounts import models as accounts
from amodb.apps.fleet import models as fleet
from amodb.database import get_write_db
from amodb.security import get_current_active_user
from . import workbook_parity as wp
from . import workbook_parity_defaults as defaults
from . import workbook_parity_imports as legacy

NS="http://schemas.openxmlformats.org/spreadsheetml/2006/main"; REL="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
REQUIRED={
 "SAFARILINK-C208B-RP":{"ALRT CALC","DATA TYPE INDEX","AU","AI","FI","SB","SR","OS","RM","SM","PM"},
 "SAFARILINK-DHC8-RP":{"ALRT CALC HRS","ALRT CALC","UR","DATA TYPE INDEX","AU","AI","FI","OS","PM","SM","RM","SB","CS","AS","SR"}}
TECH={"AID-AIN":"aircraft_serial_number","AID-REG":"aircraft_registration","LBD-DOT":"event_date","AEV-IOD":"event_date","RLS-RED":"event_date","HCD-WOE":"event_date","HCD-WOD":"event_date","SBI-SBD":"accomplishment_date","HDR-RDT":"reporting_period_start","LBD-OEI":"reference_code","AEV-ERI":"reference_code","HCD-HRI":"reference_code","RCS-SFI":"reference_code","LBD-ATA":"ata_chapter","RLS-SER":"off_serial_number","RLS-MPN":"off_part_number","RLS-TTY":"removal_type","LBD-DCT":"defect_description","LBD-MNT":"action_taken","LBD-DOC":"report_type","EVC-DLY":"delay_indicator","EVC-CNX":"cancellation_indicator","EVC-SUB":"substitute_aircraft_indicator","AEV-ICD":"interruption_code","EVC-DTM":"delay_time_minutes","EVT-DCT":"defect_description","EVT-MNT":"event_corrective_action","HCD-WON":"workpack_reference","HCD-MRB":"programme_item_reference","SBI-MDX":"implementation_type","SBI-MTE":"document_type","SBI-MFR":"document_source","SBI-SBN":"service_bulletin_number","SBI-STN":"stc_mod_number","SBI-ARW":"airworthiness_directive_number","SBI-IDE":"issue_date","AID-AMC":"aircraft_model","AID-ASE":"aircraft_series","AID-CTH":"aircraft_total_hours","AID-CTY":"aircraft_total_cycles","OSI-CHD":"effective_change_date"}
ALIASES={
 "event_date":("EVENT DATE","OCCURRENCE DATE","REMOVAL DATE","CHECK COMPLETION DATE","OCCUR DATE"),"aircraft_serial_number":("MSN","AIRCRAFT SERIAL NUMBER","MANUFACTURER SERIAL NUMBER"),"aircraft_registration":("REGISTRATION","A/C REG","TAIL NUMBER"),"ata_chapter":("ATA","ATA CHAPTER","ATA CODE"),"reference_code":("REFERENCE","TECH LOG REF","WORK ORDER NUMBER"),"title":("TITLE","SUBJECT","COMPONENT REPLACED"),"description":("DESCRIPTION","REMARKS","NARRATIVE"),"quantity_per_aircraft":("QPA","QUANTITY PER AIRCRAFT"),"unit_hours":("UNIT HOURS","FLEET HOURS"),"unscheduled_removals":("UNSLD RMVLS","UNSCHEDULED REMOVALS"),"total_removals":("TOTAL RMVLS","TOTAL REMOVALS"),"reporting_period":("QUARTER","REPORTING PERIOD"),"fleet_variant":("FLEET VARIANT","AIRCRAFT SERIES"),"component_description":("COMPONENT","COMPONENT REPLACED","PART DESCRIPTION TEXT"),"part_number":("PART NO.","PART NO","PART NUMBER"),"manufacturer_serial_number":("AIRCRAFT MANUFACTURER SERIAL NUMBER",),"registration_number":("AIRCRAFT REGISTRATION NUMBER",),"effective_change_date":("EFFECTIVE CHANGE DATE",),"reporting_period_end":("REPORTING PERIOD END DATE",),"item_description":("ITEM DESCRIPTION","PART DESCRIPTION TEXT"),"currency":("CURRENCY",),"total_cost":("TOTAL COST",),"labour_cost":("LABOUR COST",),"material_cost":("MATERIAL COST",),"other_cost":("OTHER COST",)}

def norm(v:Any)->str:return re.sub(r"[^A-Z0-9]+"," ",str(v or "").upper()).strip()
def _profile_match(code:str,names:list[str])->dict[str,Any]:
    if code=="GENERIC-ANALYSIS-TEMPLATE": return {"matched":True,"missing_sheets":[],"unexpected_profile":False}
    missing=sorted(REQUIRED.get(code,set())-{norm(n) for n in names}); return {"matched":not missing,"missing_sheets":missing,"unexpected_profile":bool(missing)}

def audit(content:bytes)->dict[str,Any]:
    try:z=zipfile.ZipFile(BytesIO(content)); root=ET.fromstring(z.read("xl/workbook.xml"))
    except Exception as e: raise HTTPException(422,"Invalid Open XML workbook.") from e
    names=set(z.namelist()); rels={}
    if "xl/_rels/workbook.xml.rels" in names:
        rr=ET.fromstring(z.read("xl/_rels/workbook.xml.rels")); rels={r.attrib.get("Id"):r.attrib.get("Target","") for r in rr}
    sheets=[]; errors=Counter(); formulas=protected=0
    node=root.find(f"{{{NS}}}sheets")
    for s in node or []:
        target=rels.get(s.attrib.get(f"{{{REL}}}id"),"").replace("../",""); path=target if target.startswith("xl/") else f"xl/{target}"; path=path if path in names else f"xl/worksheets/{Path(target).name}"
        count=0; per=Counter(); guard=False; dim=None
        if path in names:
            x=ET.fromstring(z.read(path)); count=len(x.findall(f".//{{{NS}}}f")); guard=x.find(f"{{{NS}}}sheetProtection") is not None; d=x.find(f"{{{NS}}}dimension"); dim=d.attrib.get("ref") if d is not None else None
            for c in x.findall(f".//{{{NS}}}c"):
                if c.attrib.get("t")=="e":
                    v=c.find(f"{{{NS}}}v"); per[v.text if v is not None else "UNKNOWN"]+=1
        formulas+=count; errors.update(per); protected+=int(guard); sheets.append({"name":s.attrib.get("name"),"state":s.attrib.get("state","visible"),"dimension":dim,"formula_count":count,"cached_error_counts":dict(per),"protected":guard})
    broken=[]; dn=root.find(f"{{{NS}}}definedNames")
    for n in dn or []:
        if "#REF!" in (n.text or "") or "#NAME?" in (n.text or ""): broken.append(n.attrib.get("name",""))
    return {"has_vba":"xl/vbaProject.bin" in names,"has_vba_signature":any("vbaProjectSignature" in n for n in names),"external_link_count":sum(n.startswith("xl/externalLinks/") and n.endswith(".xml") and "/_rels/" not in n for n in names),"chart_count":sum(n.startswith("xl/charts/chart") and n.endswith(".xml") for n in names),"formula_count":formulas,"formula_error_counts":dict(errors),"broken_defined_names":broken,"hidden_sheet_count":sum(s["state"]!="visible" for s in sheets),"protected_sheet_count":protected,"sheets":sheets}

def aliases(db,amo,profile,code,sheet):
    out=legacy._candidate_aliases(db,amo,profile,code,sheet)
    for k,vs in ALIASES.items(): out.setdefault(k,{norm(k)}).update(norm(v) for v in vs)
    for k,v in TECH.items(): out.setdefault(v,{norm(v)}).add(norm(k))
    return out

def header(sheet,alias,requested):
    best=None; score=-10**9
    for r in ([requested] if requested>1 else range(1,min(sheet.max_row,12)+1)):
        rows=[]
        for n in (r,r+1): rows.append([c.value for c in next(sheet.iter_rows(min_row=n,max_row=n),())])
        mapping={}; used=set(); errs=[]
        for i in range(max(map(len,rows))):
            raw=[row[i] if i<len(row) else None for row in rows]; tech=next((TECH.get(str(v or "").strip().upper()) for v in raw if TECH.get(str(v or "").strip().upper())),None)
            candidates={k for k,vals in alias.items() if any(norm(v) in vals for v in raw if norm(v))}; key=tech or (next(iter(candidates)) if len(candidates)==1 else None)
            if len(candidates)>1 and not tech: errs.append(f"Column {i+1} is ambiguous: {', '.join(sorted(candidates))}.")
            if key and key not in used: mapping[str(i+1)]=key; used.add(key)
        s=len(mapping)*10-len(errs)*25
        if s>score: best=(r,rows[0],rows[1],mapping,errs); score=s
    if not best or not best[3]: raise HTTPException(422,"No controlled header row could be detected.")
    return best

def coerce(v,kind,label,epoch):
    if v in (None,""): return None
    if kind in {"date","datetime"}:
        if isinstance(v,(int,float,Decimal)): v=from_excel(v,epoch)
        if isinstance(v,datetime): return v.isoformat() if kind=="datetime" else v.date().isoformat()
        if isinstance(v,date): return v.isoformat()
        t=str(v).strip(); q=re.fullmatch(r"Q([1-4])\s+(\d{4})",t.upper())
        if q:return date(int(q[2]),(int(q[1])-1)*3+1,1).isoformat()
        return datetime.fromisoformat(t).isoformat() if kind=="datetime" else date.fromisoformat(t[:10]).isoformat()
    if kind=="integer":
        x=Decimal(str(v));
        if x!=x.to_integral_value(): raise ValueError(f"{label} must be a whole number.")
        return int(x)
    if kind=="decimal":
        try:return format(Decimal(str(v)),"f")
        except InvalidOperation as e: raise ValueError(f"{label} must be numeric.") from e
    if kind=="boolean": return str(v).strip().upper() in {"1","Y","YES","TRUE","X"}
    return str(v).strip()

def resolve_aircraft(db,amo,serial,registration):
    if not serial and not registration:return None,[]
    q=db.query(fleet.Aircraft).filter(fleet.Aircraft.amo_id==amo)
    rows=q.filter(or_(fleet.Aircraft.serial_number==serial,fleet.Aircraft.registration==registration)).all(); errs=[]
    if not rows: return None,["Aircraft identity is not registered in this tenant."]
    if len(rows)>1:return None,["Aircraft identity is ambiguous within this tenant."]
    row=rows[0]
    if serial and registration and (row.serial_number!=serial or row.registration!=registration): errs.append("Registration and manufacturer serial number do not identify the same tenant aircraft.")
    return row.serial_number,errs

def batch_dict(batch):
    d=legacy._batch_dict_original(batch) if hasattr(legacy,"_batch_dict_original") else legacy._batch_dict(batch)
    d["integrity"]=next((x.get("integrity") for x in batch.detected_sheets or [] if x.get("name")=="__WORKBOOK_INTEGRITY__"),None); return d


__all__=[name for name in globals() if not name.startswith("__")]
