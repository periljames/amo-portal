from __future__ import annotations

import csv
import hashlib
import io
import json
import os
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Literal, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator, model_validator
from sqlalchemy import (
    Boolean,
    CheckConstraint,
    Column,
    DateTime,
    ForeignKey,
    Index,
    Integer,
    JSON,
    Numeric,
    String,
    Text,
    UniqueConstraint,
    func,
)
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from sqlalchemy.types import TypeDecorator

from amodb.apps.accounts import models as account_models
from amodb.apps.fleet import models as fleet_models
from amodb.apps.technical_records import models as technical_models
from amodb.database import Base, get_write_db
from amodb.security import get_current_active_user
from amodb.utils.identifiers import generate_uuid7

from . import advanced_models as domain
from . import advanced_schemas as schemas
from . import advanced_services as services


UTC = timezone.utc
WORKBOOK_MAX_BYTES = int(os.getenv("RELIABILITY_WORKBOOK_MAX_BYTES", str(25 * 1024 * 1024)))
WORKBOOK_ROOT = Path(os.getenv("RELIABILITY_WORKBOOK_UPLOAD_DIR", "uploads/reliability/workbooks")).resolve()
WORKBOOK_EXTENSIONS = {".xlsx", ".csv"}
SEVERITIES = {"LOW", "MEDIUM", "HIGH", "CRITICAL"}
FLIGHT_TYPES = {
    "TECHNICAL_DELAY",
    "TECHNICAL_CANCELLATION",
    "RETURN_TO_GATE",
    "AIR_TURNBACK",
    "DIVERSION",
    "IN_FLIGHT_SHUTDOWN",
    "ABORTED_TAKEOFF",
}


def _utcnow() -> datetime:
    return datetime.now(UTC)


def _utc(value: datetime) -> datetime:
    return value.replace(tzinfo=UTC) if value.tzinfo is None else value.astimezone(UTC)


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return _utc(value).isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _safe_filename(value: str) -> str:
    name = Path(value or "workbook.xlsx").name
    return "".join(ch for ch in name if ch.isalnum() or ch in {"-", "_", "."}) or "workbook.xlsx"


class ExactAviationDecimal(TypeDecorator):
    """Exact database representation for regulated hours and life values."""

    impl = Numeric(20, 3)
    cache_ok = True

    def process_bind_param(self, value: Any, dialect):
        if value is None:
            return None
        try:
            parsed = Decimal(str(value))
        except Exception as exc:
            raise ValueError("Aviation time value must be numeric.") from exc
        if not parsed.is_finite():
            raise ValueError("Aviation time value must be finite.")
        return parsed.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

    def process_result_value(self, value: Any, dialect):
        return None if value is None else Decimal(str(value)).quantize(Decimal("0.001"))


class ExactAviationCount(TypeDecorator):
    """Exact integral database representation for cycles and landings."""

    impl = Numeric(20, 0)
    cache_ok = True

    def process_bind_param(self, value: Any, dialect):
        if value is None:
            return None
        try:
            parsed = Decimal(str(value))
        except Exception as exc:
            raise ValueError("Cycle value must be numeric.") from exc
        if not parsed.is_finite() or parsed != parsed.to_integral_value():
            raise ValueError("Cycle value must be a whole number.")
        return parsed

    def process_result_value(self, value: Any, dialect):
        return None if value is None else Decimal(str(value)).quantize(Decimal("1"))


def _patch_exact_column(model: Any, name: str, *, count: bool = False) -> None:
    column = model.__table__.c.get(name)
    if column is not None:
        column.type = ExactAviationCount() if count else ExactAviationDecimal()


def install_exact_aviation_types() -> None:
    hour_fields = {
        fleet_models.Aircraft: ["total_hours"],
        fleet_models.AircraftComponent: [
            "installed_hours", "current_hours", "tbo_hours", "hsi_hours", "last_overhaul_hours",
        ],
        fleet_models.AircraftUsage: [
            "block_hours", "ttaf_after", "ttesn_after", "ttsoh_after", "ttshsi_after",
            "pttsn_after", "pttso_after", "tscoa_after", "hours_to_mx",
        ],
        fleet_models.MaintenanceProgramItem: ["interval_hours"],
        fleet_models.MaintenanceStatus: ["last_done_hours", "next_due_hours", "remaining_hours"],
        technical_models.AircraftUtilisation: ["hours"],
        technical_models.AirworthinessItem: ["next_due_hours"],
        technical_models.AirworthinessComplianceEvent: ["next_due_hours"],
        technical_models.ComplianceAction: ["due_hours"],
    }
    count_fields = {
        fleet_models.Aircraft: ["total_cycles"],
        fleet_models.AircraftComponent: [
            "installed_cycles", "current_cycles", "tbo_cycles", "hsi_cycles", "last_overhaul_cycles",
        ],
        fleet_models.AircraftUsage: ["cycles", "tca_after", "tcesn_after", "tcsoh_after"],
        fleet_models.MaintenanceProgramItem: ["interval_cycles"],
        fleet_models.MaintenanceStatus: ["last_done_cycles", "next_due_cycles", "remaining_cycles"],
        technical_models.AircraftUtilisation: ["cycles"],
        technical_models.AirworthinessItem: ["next_due_cycles"],
        technical_models.AirworthinessComplianceEvent: ["next_due_cycles"],
        technical_models.ComplianceAction: ["due_cycles"],
    }
    for model, names in hour_fields.items():
        for name in names:
            _patch_exact_column(model, name)
    for model, names in count_fields.items():
        for name in names:
            _patch_exact_column(model, name, count=True)


class ReliabilityFlightOperation(Base):
    __tablename__ = "reliability_flight_operations"
    __table_args__ = (
        UniqueConstraint("amo_id", "record_number", name="uq_rel_flight_record"),
        Index("ix_rel_flight_amo_occurred", "amo_id", "occurred_at"),
        Index("ix_rel_flight_amo_status", "amo_id", "status"),
        CheckConstraint("delay_minutes IS NULL OR delay_minutes >= 0", name="ck_rel_flight_delay"),
    )

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    record_number = Column(String(80), nullable=False)
    revision = Column(Integer, nullable=False, default=1)
    event_type = Column(String(40), nullable=False, index=True)
    occurred_at = Column(DateTime(timezone=True), nullable=False, index=True)
    aircraft_serial_number = Column(String(50), ForeignKey("aircraft.serial_number", ondelete="RESTRICT"), nullable=False, index=True)
    flight_number = Column(String(24), nullable=False, index=True)
    origin_station = Column(String(8), nullable=True)
    destination_station = Column(String(8), nullable=True)
    scheduled_departure_at = Column(DateTime(timezone=True), nullable=True)
    actual_departure_at = Column(DateTime(timezone=True), nullable=True)
    delay_minutes = Column(Integer, nullable=True)
    dispatch_impact = Column(String(40), nullable=True)
    severity = Column(String(16), nullable=False, default="MEDIUM")
    ata_chapter = Column(String(20), nullable=True)
    description = Column(Text, nullable=False)
    status = Column(String(24), nullable=False, default="DRAFT", index=True)
    canonical_event_id = Column(Integer, ForeignKey("reliability_events.id", ondelete="SET NULL"), nullable=True, index=True)
    created_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_at = Column(DateTime(timezone=True), nullable=True)
    closed_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    closed_at = Column(DateTime(timezone=True), nullable=True)
    closure_note = Column(Text, nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)
    updated_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow, onupdate=_utcnow)


class ReliabilityMelCdlDeferral(Base):
    __tablename__ = "reliability_mel_cdl_deferrals"
    __table_args__ = (
        UniqueConstraint("amo_id", "deferral_number", name="uq_rel_deferral_number"),
        Index("ix_rel_deferral_amo_expiry", "amo_id", "expires_at"),
        Index("ix_rel_deferral_amo_status", "amo_id", "status"),
        CheckConstraint("expires_at >= applied_at", name="ck_rel_deferral_expiry"),
        CheckConstraint(
            "repetitive_inspection_minutes IS NULL OR repetitive_inspection_minutes > 0",
            name="ck_rel_deferral_repeat_interval",
        ),
    )

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    deferral_number = Column(String(80), nullable=False)
    revision = Column(Integer, nullable=False, default=1)
    deferral_type = Column(String(8), nullable=False, index=True)
    aircraft_serial_number = Column(String(50), ForeignKey("aircraft.serial_number", ondelete="RESTRICT"), nullable=False, index=True)
    defect_reference = Column(String(80), nullable=False)
    item_reference = Column(String(80), nullable=False)
    category = Column(String(16), nullable=True)
    applied_at = Column(DateTime(timezone=True), nullable=False)
    expires_at = Column(DateTime(timezone=True), nullable=False, index=True)
    control_basis = Column(Text, nullable=False)
    operational_procedure = Column(Text, nullable=True)
    maintenance_procedure = Column(Text, nullable=True)
    repetitive_inspection_minutes = Column(Integer, nullable=True)
    flight_number = Column(String(24), nullable=True)
    ata_chapter = Column(String(20), nullable=True)
    description = Column(Text, nullable=False)
    severity = Column(String(16), nullable=False, default="MEDIUM")
    status = Column(String(24), nullable=False, default="DRAFT", index=True)
    extension_history_json = Column(JSON, nullable=False, default=list)
    closure_evidence_json = Column(JSON, nullable=False, default=list)
    canonical_event_id = Column(Integer, ForeignKey("reliability_events.id", ondelete="SET NULL"), nullable=True, index=True)
    created_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_at = Column(DateTime(timezone=True), nullable=True)
    closed_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    closed_at = Column(DateTime(timezone=True), nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)
    updated_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow, onupdate=_utcnow)


class ReliabilityComponentShopFinding(Base):
    __tablename__ = "reliability_component_shop_findings"
    __table_args__ = (
        UniqueConstraint("amo_id", "shop_order_reference", name="uq_rel_shop_order"),
        Index("ix_rel_shop_amo_status", "amo_id", "status"),
        Index("ix_rel_shop_component", "amo_id", "part_number", "component_serial_number"),
        CheckConstraint(
            "event_type <> 'NO_FAULT_FOUND' OR confirmed_failure = false",
            name="ck_rel_shop_nff_false",
        ),
    )

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    shop_order_reference = Column(String(120), nullable=False)
    revision = Column(Integer, nullable=False, default=1)
    event_type = Column(String(32), nullable=False, index=True)
    component_id = Column(Integer, ForeignKey("aircraft_components.id", ondelete="SET NULL"), nullable=True, index=True)
    aircraft_serial_number = Column(String(50), ForeignKey("aircraft.serial_number", ondelete="SET NULL"), nullable=True, index=True)
    part_number = Column(String(80), nullable=False, index=True)
    component_serial_number = Column(String(80), nullable=False, index=True)
    received_at = Column(DateTime(timezone=True), nullable=False)
    inspected_at = Column(DateTime(timezone=True), nullable=False)
    ata_chapter = Column(String(20), nullable=True)
    confirmed_failure = Column(Boolean, nullable=True)
    test_result = Column(Text, nullable=False)
    disposition = Column(Text, nullable=False)
    release_reference = Column(String(120), nullable=True)
    description = Column(Text, nullable=False)
    severity = Column(String(16), nullable=False, default="MEDIUM")
    status = Column(String(24), nullable=False, default="DRAFT", index=True)
    canonical_event_id = Column(Integer, ForeignKey("reliability_events.id", ondelete="SET NULL"), nullable=True, index=True)
    created_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_at = Column(DateTime(timezone=True), nullable=True)
    released_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    released_at = Column(DateTime(timezone=True), nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)
    updated_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow, onupdate=_utcnow)


class ReliabilitySmsOccurrence(Base):
    __tablename__ = "reliability_sms_occurrences"
    __table_args__ = (
        UniqueConstraint("amo_id", "sms_reference", name="uq_rel_sms_reference"),
        Index("ix_rel_sms_amo_status", "amo_id", "status"),
        Index("ix_rel_sms_amo_occurred", "amo_id", "occurred_at"),
        CheckConstraint(
            "reliability_relevant = false OR reliability_link_reason IS NOT NULL",
            name="ck_rel_sms_link_reason",
        ),
    )

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    sms_reference = Column(String(120), nullable=False)
    revision = Column(Integer, nullable=False, default=1)
    occurred_at = Column(DateTime(timezone=True), nullable=False)
    aircraft_serial_number = Column(String(50), ForeignKey("aircraft.serial_number", ondelete="SET NULL"), nullable=True, index=True)
    hazard_reference = Column(String(120), nullable=True)
    risk_classification = Column(String(80), nullable=False)
    investigation_status = Column(String(40), nullable=False, default="OPEN")
    reliability_relevant = Column(Boolean, nullable=False, default=False)
    reliability_link_reason = Column(Text, nullable=True)
    ata_chapter = Column(String(20), nullable=True)
    description = Column(Text, nullable=False)
    severity = Column(String(16), nullable=False, default="MEDIUM")
    status = Column(String(24), nullable=False, default="DRAFT", index=True)
    canonical_event_id = Column(Integer, ForeignKey("reliability_events.id", ondelete="SET NULL"), nullable=True, index=True)
    created_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    assessed_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    assessed_at = Column(DateTime(timezone=True), nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)
    updated_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow, onupdate=_utcnow)


class ReliabilityWorkbookImport(Base):
    __tablename__ = "reliability_workbook_imports"
    __table_args__ = (
        UniqueConstraint("amo_id", "content_hash", name="uq_rel_workbook_hash"),
        Index("ix_rel_workbook_amo_status", "amo_id", "status"),
    )

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    original_filename = Column(String(255), nullable=False)
    content_hash = Column(String(64), nullable=False)
    storage_path = Column(Text, nullable=False)
    revision = Column(Integer, nullable=False, default=1)
    header_row = Column(Integer, nullable=False, default=1)
    mapping_json = Column(JSON, nullable=False, default=dict)
    defaults_json = Column(JSON, nullable=False, default=dict)
    status = Column(String(24), nullable=False, default="UPLOADED", index=True)
    total_rows = Column(Integer, nullable=False, default=0)
    valid_rows = Column(Integer, nullable=False, default=0)
    invalid_rows = Column(Integer, nullable=False, default=0)
    approved_rows = Column(Integer, nullable=False, default=0)
    rejected_rows = Column(Integer, nullable=False, default=0)
    ingested_rows = Column(Integer, nullable=False, default=0)
    created_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    approved_at = Column(DateTime(timezone=True), nullable=True)
    ingested_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    ingested_at = Column(DateTime(timezone=True), nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)
    updated_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow, onupdate=_utcnow)


class ReliabilityWorkbookRow(Base):
    __tablename__ = "reliability_workbook_rows"
    __table_args__ = (
        UniqueConstraint("import_id", "sheet_name", "source_row_number", name="uq_rel_workbook_row"),
        Index("ix_rel_workbook_row_status", "import_id", "status"),
        Index("ix_rel_workbook_row_hash", "amo_id", "row_hash"),
    )

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    import_id = Column(String(36), ForeignKey("reliability_workbook_imports.id", ondelete="CASCADE"), nullable=False, index=True)
    sheet_name = Column(String(255), nullable=False)
    source_row_number = Column(Integer, nullable=False)
    row_hash = Column(String(64), nullable=False)
    raw_json = Column(JSON, nullable=False)
    mapped_json = Column(JSON, nullable=False, default=dict)
    validation_errors_json = Column(JSON, nullable=False, default=list)
    status = Column(String(24), nullable=False, default="PENDING", index=True)
    decision_note = Column(Text, nullable=True)
    decided_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    decided_at = Column(DateTime(timezone=True), nullable=True)
    canonical_event_id = Column(Integer, ForeignKey("reliability_events.id", ondelete="SET NULL"), nullable=True, index=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)


class ReliabilitySourceRevisionEvent(Base):
    __tablename__ = "reliability_source_revision_events"
    __table_args__ = (Index("ix_rel_source_revision_chain", "amo_id", "source_type", "source_id", "created_at"),)

    id = Column(String(36), primary_key=True, default=generate_uuid7)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    source_type = Column(String(40), nullable=False)
    source_id = Column(String(36), nullable=False)
    revision = Column(Integer, nullable=False)
    action = Column(String(40), nullable=False)
    payload_json = Column(JSON, nullable=False, default=dict)
    actor_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=_utcnow)


class TenantModel(BaseModel):
    model_config = {"from_attributes": True}


class FlightOperationCreate(BaseModel):
    record_number: str = Field(min_length=1, max_length=80)
    event_type: str
    occurred_at: datetime
    aircraft_serial_number: str = Field(min_length=1, max_length=50)
    flight_number: str = Field(min_length=1, max_length=24)
    origin_station: Optional[str] = Field(default=None, max_length=8)
    destination_station: Optional[str] = Field(default=None, max_length=8)
    scheduled_departure_at: Optional[datetime] = None
    actual_departure_at: Optional[datetime] = None
    delay_minutes: Optional[int] = Field(default=None, ge=0)
    dispatch_impact: Optional[str] = Field(default=None, max_length=40)
    severity: str = "MEDIUM"
    ata_chapter: Optional[str] = Field(default=None, max_length=20)
    description: str = Field(min_length=3, max_length=12000)

    @field_validator("event_type")
    @classmethod
    def event_type_valid(cls, value: str) -> str:
        resolved = value.strip().upper()
        if resolved not in FLIGHT_TYPES:
            raise ValueError("Unsupported Flight Operations interruption type.")
        return resolved

    @field_validator("severity")
    @classmethod
    def severity_valid(cls, value: str) -> str:
        resolved = value.strip().upper()
        if resolved not in SEVERITIES:
            raise ValueError("Unsupported severity.")
        return resolved

    @model_validator(mode="after")
    def delay_required(self):
        if self.event_type == "TECHNICAL_DELAY" and self.delay_minutes is None:
            raise ValueError("Technical delay records require delay_minutes.")
        return self


class FlightOperationRead(TenantModel):
    id: str
    record_number: str
    revision: int
    event_type: str
    occurred_at: datetime
    aircraft_serial_number: str
    flight_number: str
    origin_station: Optional[str]
    destination_station: Optional[str]
    delay_minutes: Optional[int]
    dispatch_impact: Optional[str]
    severity: str
    ata_chapter: Optional[str]
    description: str
    status: str
    canonical_event_id: Optional[int]
    approved_at: Optional[datetime]
    closed_at: Optional[datetime]
    closure_note: Optional[str]


class DeferralCreate(BaseModel):
    deferral_number: str = Field(min_length=1, max_length=80)
    deferral_type: Literal["MEL", "CDL"]
    aircraft_serial_number: str = Field(min_length=1, max_length=50)
    defect_reference: str = Field(min_length=1, max_length=80)
    item_reference: str = Field(min_length=1, max_length=80)
    category: Optional[str] = Field(default=None, max_length=16)
    applied_at: datetime
    expires_at: datetime
    control_basis: str = Field(min_length=3, max_length=4000)
    operational_procedure: Optional[str] = Field(default=None, max_length=12000)
    maintenance_procedure: Optional[str] = Field(default=None, max_length=12000)
    repetitive_inspection_minutes: Optional[int] = Field(default=None, gt=0)
    flight_number: Optional[str] = Field(default=None, max_length=24)
    ata_chapter: Optional[str] = Field(default=None, max_length=20)
    description: str = Field(min_length=3, max_length=12000)
    severity: str = "MEDIUM"

    @model_validator(mode="after")
    def dates_valid(self):
        if _utc(self.expires_at) < _utc(self.applied_at):
            raise ValueError("Deferral expiry cannot precede application.")
        return self


class DeferralRead(TenantModel):
    id: str
    deferral_number: str
    revision: int
    deferral_type: str
    aircraft_serial_number: str
    defect_reference: str
    item_reference: str
    category: Optional[str]
    applied_at: datetime
    expires_at: datetime
    control_basis: str
    operational_procedure: Optional[str]
    maintenance_procedure: Optional[str]
    repetitive_inspection_minutes: Optional[int]
    flight_number: Optional[str]
    ata_chapter: Optional[str]
    description: str
    severity: str
    status: str
    extension_history_json: List[Dict[str, Any]]
    closure_evidence_json: List[Dict[str, Any]]
    canonical_event_id: Optional[int]


class DeferralExtension(BaseModel):
    new_expires_at: datetime
    reason: str = Field(min_length=5, max_length=4000)
    approval_reference: str = Field(min_length=1, max_length=160)


class CloseRequest(BaseModel):
    note: str = Field(min_length=5, max_length=4000)
    evidence: List[Dict[str, Any]] = Field(default_factory=list)


class ShopFindingCreate(BaseModel):
    shop_order_reference: str = Field(min_length=1, max_length=120)
    event_type: Literal["SHOP_FINDING", "NO_FAULT_FOUND"]
    component_id: Optional[int] = Field(default=None, ge=1)
    aircraft_serial_number: Optional[str] = Field(default=None, max_length=50)
    part_number: str = Field(min_length=1, max_length=80)
    component_serial_number: str = Field(min_length=1, max_length=80)
    received_at: datetime
    inspected_at: datetime
    ata_chapter: Optional[str] = Field(default=None, max_length=20)
    confirmed_failure: Optional[bool] = None
    test_result: str = Field(min_length=2, max_length=12000)
    disposition: str = Field(min_length=2, max_length=12000)
    release_reference: Optional[str] = Field(default=None, max_length=120)
    description: str = Field(min_length=3, max_length=12000)
    severity: str = "MEDIUM"

    @model_validator(mode="after")
    def disposition_valid(self):
        if _utc(self.inspected_at) < _utc(self.received_at):
            raise ValueError("Inspection cannot precede receipt.")
        if self.event_type == "NO_FAULT_FOUND" and self.confirmed_failure is not False:
            raise ValueError("NFF requires confirmed_failure=false.")
        return self


class ShopFindingRead(TenantModel):
    id: str
    shop_order_reference: str
    revision: int
    event_type: str
    component_id: Optional[int]
    aircraft_serial_number: Optional[str]
    part_number: str
    component_serial_number: str
    received_at: datetime
    inspected_at: datetime
    ata_chapter: Optional[str]
    confirmed_failure: Optional[bool]
    test_result: str
    disposition: str
    release_reference: Optional[str]
    description: str
    severity: str
    status: str
    canonical_event_id: Optional[int]


class ReleaseRequest(BaseModel):
    release_reference: str = Field(min_length=1, max_length=120)
    note: str = Field(min_length=3, max_length=4000)


class SmsOccurrenceCreate(BaseModel):
    sms_reference: str = Field(min_length=1, max_length=120)
    occurred_at: datetime
    aircraft_serial_number: Optional[str] = Field(default=None, max_length=50)
    hazard_reference: Optional[str] = Field(default=None, max_length=120)
    risk_classification: str = Field(min_length=1, max_length=80)
    investigation_status: str = Field(default="OPEN", min_length=1, max_length=40)
    ata_chapter: Optional[str] = Field(default=None, max_length=20)
    description: str = Field(min_length=3, max_length=12000)
    severity: str = "MEDIUM"


class SmsOccurrenceRead(TenantModel):
    id: str
    sms_reference: str
    revision: int
    occurred_at: datetime
    aircraft_serial_number: Optional[str]
    hazard_reference: Optional[str]
    risk_classification: str
    investigation_status: str
    reliability_relevant: bool
    reliability_link_reason: Optional[str]
    ata_chapter: Optional[str]
    description: str
    severity: str
    status: str
    canonical_event_id: Optional[int]


class SmsAssessment(BaseModel):
    reliability_relevant: bool
    reliability_link_reason: Optional[str] = Field(default=None, max_length=4000)
    investigation_status: str = Field(min_length=1, max_length=40)

    @model_validator(mode="after")
    def reason_required(self):
        if self.reliability_relevant and not (self.reliability_link_reason or "").strip():
            raise ValueError("A Reliability link reason is required for relevant SMS occurrences.")
        return self


class WorkbookMapRequest(BaseModel):
    mapping: Dict[str, str]
    defaults: Dict[str, Any] = Field(default_factory=dict)

    @model_validator(mode="after")
    def required_mapping(self):
        required = {"event_type", "occurred_at", "description"}
        available = set(self.mapping) | set(self.defaults)
        missing = sorted(required - available)
        if missing:
            raise ValueError(f"Missing canonical mapping fields: {', '.join(missing)}")
        return self


class WorkbookDecision(BaseModel):
    decision: Literal["APPROVE", "REJECT"]
    note: str = Field(min_length=3, max_length=4000)


class WorkbookImportRead(TenantModel):
    id: str
    original_filename: str
    content_hash: str
    revision: int
    header_row: int
    mapping_json: Dict[str, str]
    defaults_json: Dict[str, Any]
    status: str
    total_rows: int
    valid_rows: int
    invalid_rows: int
    approved_rows: int
    rejected_rows: int
    ingested_rows: int
    approved_at: Optional[datetime]
    ingested_at: Optional[datetime]
    created_at: datetime


class WorkbookRowRead(TenantModel):
    id: str
    import_id: str
    sheet_name: str
    source_row_number: int
    raw_json: Dict[str, Any]
    mapped_json: Dict[str, Any]
    validation_errors_json: List[Any]
    status: str
    decision_note: Optional[str]
    canonical_event_id: Optional[int]


class OperationalSummary(BaseModel):
    flight_operations: Dict[str, int]
    deferrals: Dict[str, int]
    component_shop: Dict[str, int]
    sms: Dict[str, int]
    workbooks: Dict[str, int]
    generated_at: datetime


def _context(
    current_user: account_models.User = Depends(get_current_active_user),
    db: Session = Depends(get_write_db),
) -> tuple[account_models.User, Session, str]:
    return current_user, db, services.tenant_id(current_user)


def _require(db: Session, user: account_models.User, capability: str) -> None:
    services.require_capability(db, user, capability)


def _aircraft(db: Session, amo_id: str, serial: Optional[str]) -> None:
    if not serial:
        return
    exists = db.query(fleet_models.Aircraft.serial_number).filter(
        fleet_models.Aircraft.amo_id == amo_id,
        fleet_models.Aircraft.serial_number == serial,
    ).first()
    if not exists:
        raise HTTPException(status_code=422, detail="Aircraft does not exist in this tenant.")


def _component(
    db: Session,
    amo_id: str,
    component_id: Optional[int],
    part_number: str,
    serial_number: str,
    aircraft_serial_number: Optional[str],
) -> None:
    if component_id is None:
        return
    item = db.query(fleet_models.AircraftComponent).filter(
        fleet_models.AircraftComponent.amo_id == amo_id,
        fleet_models.AircraftComponent.id == component_id,
    ).first()
    if not item:
        raise HTTPException(status_code=422, detail="Component does not exist in this tenant.")
    if item.part_number and item.part_number != part_number:
        raise HTTPException(status_code=422, detail="Part number conflicts with component master.")
    if item.serial_number and item.serial_number != serial_number:
        raise HTTPException(status_code=422, detail="Serial number conflicts with component master.")
    if aircraft_serial_number and item.aircraft_serial_number != aircraft_serial_number:
        raise HTTPException(status_code=422, detail="Aircraft conflicts with component master.")


def _revision_event(
    db: Session,
    *,
    amo_id: str,
    source_type: str,
    source_id: str,
    revision: int,
    action: str,
    payload: Dict[str, Any],
    actor_user_id: Optional[str],
) -> None:
    db.add(ReliabilitySourceRevisionEvent(
        amo_id=amo_id,
        source_type=source_type,
        source_id=source_id,
        revision=revision,
        action=action,
        payload_json=payload,
        actor_user_id=actor_user_id,
    ))
    services.append_audit(
        db,
        amo_id=amo_id,
        entity_type=source_type,
        entity_id=source_id,
        action=action,
        payload={"revision": revision, **payload},
        actor_user_id=actor_user_id,
    )


def _source(adapters: Any, db: Session, amo_id: str, code: str, actor_user_id: str):
    spec = next(item for item in adapters.ADAPTER_SPECS if item.code == code)
    return adapters._ensure_source(db, amo_id=amo_id, spec=spec, actor_user_id=actor_user_id)


def _external_id(code: str, record_id: str, revision: int) -> str:
    return f"{code}:{record_id}:{revision}"[:255]


def _find_event_id(db: Session, amo_id: str, source_id: str, external_id: str) -> Optional[int]:
    row = db.query(domain.ReliabilityIngestionRecord.normalized_event_id).filter(
        domain.ReliabilityIngestionRecord.amo_id == amo_id,
        domain.ReliabilityIngestionRecord.source_id == source_id,
        domain.ReliabilityIngestionRecord.external_id == external_id,
    ).first()
    return int(row[0]) if row and row[0] is not None else None


def _ingest(adapters: Any, db: Session, *, amo_id: str, code: str, record: Dict[str, Any], actor_user_id: str):
    source = _source(adapters, db, amo_id, code, actor_user_id)
    result = services.ingest_batch(
        db,
        amo_id=amo_id,
        source=source,
        payload=schemas.ReliabilityBatchIngest(
            records=[record],
            metadata_json={
                "adapter": "authoritative-operational-v1",
                "source_code": code,
                "source_record_id": record.get("authoritative_source_record_id"),
                "source_revision": record.get("authoritative_source_revision"),
            },
        ),
        actor_user_id=actor_user_id,
    )
    event_id = result.created_event_ids[0] if result.created_event_ids else _find_event_id(
        db, amo_id, source.id, str(record["external_id"])
    )
    return result, event_id


def _flight_record(row: ReliabilityFlightOperation) -> Dict[str, Any]:
    return {
        "external_id": _external_id("FLIGHT-OPERATIONS", row.id, row.revision),
        "event_type": row.event_type,
        "occurred_at": _utc(row.occurred_at).isoformat(),
        "aircraft_serial_number": row.aircraft_serial_number,
        "flight_number": row.flight_number,
        "origin_station": row.origin_station,
        "destination_station": row.destination_station,
        "scheduled_departure_at": _json_value(row.scheduled_departure_at),
        "actual_departure_at": _json_value(row.actual_departure_at),
        "delay_minutes": row.delay_minutes,
        "dispatch_impact": row.dispatch_impact,
        "severity": row.severity,
        "ata_chapter": row.ata_chapter,
        "reference_code": row.record_number,
        "description": row.description,
        "record_status": row.status,
        "authoritative_source_record_id": row.id,
        "authoritative_source_revision": str(row.revision),
    }


def _deferral_record(row: ReliabilityMelCdlDeferral) -> Dict[str, Any]:
    return {
        "external_id": _external_id("MEL-CDL", row.id, row.revision),
        "event_type": f"{row.deferral_type}_DEFERRAL",
        "occurred_at": _utc(row.applied_at).isoformat(),
        "aircraft_serial_number": row.aircraft_serial_number,
        "flight_number": row.flight_number,
        "mel_reference": row.item_reference if row.deferral_type == "MEL" else None,
        "cdl_reference": row.item_reference if row.deferral_type == "CDL" else None,
        "deferral_category": row.category,
        "deferred_until": _utc(row.expires_at).isoformat(),
        "control_basis": row.control_basis,
        "operational_procedure": row.operational_procedure,
        "maintenance_procedure": row.maintenance_procedure,
        "repetitive_inspection_minutes": row.repetitive_inspection_minutes,
        "severity": row.severity,
        "ata_chapter": row.ata_chapter,
        "reference_code": row.deferral_number,
        "description": row.description,
        "defect_reference": row.defect_reference,
        "record_status": row.status,
        "closure_evidence": row.closure_evidence_json,
        "authoritative_source_record_id": row.id,
        "authoritative_source_revision": str(row.revision),
    }


def _shop_record(row: ReliabilityComponentShopFinding) -> Dict[str, Any]:
    return {
        "external_id": _external_id("COMPONENT-SHOP-FINDINGS", row.id, row.revision),
        "event_type": row.event_type,
        "occurred_at": _utc(row.inspected_at).isoformat(),
        "aircraft_serial_number": row.aircraft_serial_number,
        "component_id": row.component_id,
        "part_number": row.part_number,
        "component_serial_number": row.component_serial_number,
        "shop_order_reference": row.shop_order_reference,
        "confirmed_failure": row.confirmed_failure,
        "test_result": row.test_result,
        "disposition": row.disposition,
        "release_reference": row.release_reference,
        "severity": row.severity,
        "ata_chapter": row.ata_chapter,
        "reference_code": row.shop_order_reference,
        "description": row.description,
        "record_status": row.status,
        "authoritative_source_record_id": row.id,
        "authoritative_source_revision": str(row.revision),
    }


def _sms_record(row: ReliabilitySmsOccurrence) -> Dict[str, Any]:
    return {
        "external_id": _external_id("SMS-EVENTS", row.id, row.revision),
        "event_type": "SAFETY_EVENT",
        "occurred_at": _utc(row.occurred_at).isoformat(),
        "aircraft_serial_number": row.aircraft_serial_number,
        "sms_reference": row.sms_reference,
        "hazard_reference": row.hazard_reference,
        "risk_classification": row.risk_classification,
        "investigation_status": row.investigation_status,
        "reliability_link_reason": row.reliability_link_reason,
        "severity": row.severity,
        "ata_chapter": row.ata_chapter,
        "reference_code": row.sms_reference,
        "description": row.description,
        "record_status": row.status,
        "authoritative_source_record_id": row.id,
        "authoritative_source_revision": str(row.revision),
    }


def _get_row(db: Session, model: Any, amo_id: str, row_id: str):
    row = db.query(model).filter(model.amo_id == amo_id, model.id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Operational source record not found.")
    return row


def _duplicate_error(exc: Exception, label: str) -> HTTPException:
    return HTTPException(status_code=409, detail=f"A {label} with that controlled reference already exists.")


def _status_counts(db: Session, model: Any, amo_id: str) -> Dict[str, int]:
    rows = db.query(model.status, func.count(model.id)).filter(model.amo_id == amo_id).group_by(model.status).all()
    return {str(key): int(value or 0) for key, value in rows}


def operational_summary(context=Depends(_context)) -> OperationalSummary:
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    return OperationalSummary(
        flight_operations=_status_counts(db, ReliabilityFlightOperation, amo_id),
        deferrals=_status_counts(db, ReliabilityMelCdlDeferral, amo_id),
        component_shop=_status_counts(db, ReliabilityComponentShopFinding, amo_id),
        sms=_status_counts(db, ReliabilitySmsOccurrence, amo_id),
        workbooks=_status_counts(db, ReliabilityWorkbookImport, amo_id),
        generated_at=_utcnow(),
    )


def list_flight_operations(
    record_status: Optional[str] = Query(default=None, alias="status"),
    limit: int = Query(default=200, ge=1, le=500),
    context=Depends(_context),
):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    query = db.query(ReliabilityFlightOperation).filter(ReliabilityFlightOperation.amo_id == amo_id)
    if record_status:
        query = query.filter(ReliabilityFlightOperation.status == record_status.upper())
    return query.order_by(ReliabilityFlightOperation.occurred_at.desc()).limit(limit).all()


def create_flight_operation(payload: FlightOperationCreate, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    _aircraft(db, amo_id, payload.aircraft_serial_number)
    row = ReliabilityFlightOperation(
        amo_id=amo_id,
        **payload.model_dump(),
        origin_station=payload.origin_station.upper() if payload.origin_station else None,
        destination_station=payload.destination_station.upper() if payload.destination_station else None,
        created_by_user_id=str(current_user.id),
    )
    db.add(row)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise _duplicate_error(exc, "Flight Operations record") from exc
    _revision_event(db, amo_id=amo_id, source_type="FLIGHT_OPERATION", source_id=row.id, revision=1,
                    action="CREATED", payload={"record_number": row.record_number}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def approve_flight_operation(row_id: str, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityFlightOperation, amo_id, row_id)
    if row.status != "DRAFT":
        raise HTTPException(status_code=409, detail="Only draft Flight Operations records can be approved.")
    row.status = "APPROVED"
    row.revision += 1
    row.approved_by_user_id = str(current_user.id)
    row.approved_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="FLIGHT_OPERATION", source_id=row.id, revision=row.revision,
                    action="APPROVED", payload={"record_number": row.record_number}, actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="FLIGHT-OPERATIONS",
                                        record=_flight_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def close_flight_operation(row_id: str, payload: CloseRequest, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityFlightOperation, amo_id, row_id)
    if row.status != "APPROVED":
        raise HTTPException(status_code=409, detail="Only approved Flight Operations records can be closed.")
    row.status = "CLOSED"
    row.revision += 1
    row.closed_by_user_id = str(current_user.id)
    row.closed_at = _utcnow()
    row.closure_note = payload.note
    _revision_event(db, amo_id=amo_id, source_type="FLIGHT_OPERATION", source_id=row.id, revision=row.revision,
                    action="CLOSED", payload={"closure_note": payload.note, "evidence": payload.evidence}, actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="FLIGHT-OPERATIONS",
                                        record=_flight_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def list_deferrals(
    record_status: Optional[str] = Query(default=None, alias="status"),
    limit: int = Query(default=200, ge=1, le=500),
    context=Depends(_context),
):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    query = db.query(ReliabilityMelCdlDeferral).filter(ReliabilityMelCdlDeferral.amo_id == amo_id)
    if record_status:
        query = query.filter(ReliabilityMelCdlDeferral.status == record_status.upper())
    return query.order_by(ReliabilityMelCdlDeferral.expires_at.asc()).limit(limit).all()


def create_deferral(payload: DeferralCreate, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    _aircraft(db, amo_id, payload.aircraft_serial_number)
    row = ReliabilityMelCdlDeferral(amo_id=amo_id, **payload.model_dump(), created_by_user_id=str(current_user.id))
    db.add(row)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise _duplicate_error(exc, "deferral") from exc
    _revision_event(db, amo_id=amo_id, source_type="MEL_CDL_DEFERRAL", source_id=row.id, revision=1,
                    action="CREATED", payload={"deferral_number": row.deferral_number}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def approve_deferral(row_id: str, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityMelCdlDeferral, amo_id, row_id)
    if row.status != "DRAFT":
        raise HTTPException(status_code=409, detail="Only draft deferrals can be approved.")
    row.status = "OPEN"
    row.revision += 1
    row.approved_by_user_id = str(current_user.id)
    row.approved_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="MEL_CDL_DEFERRAL", source_id=row.id, revision=row.revision,
                    action="APPROVED", payload={"expires_at": _utc(row.expires_at).isoformat()}, actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="MEL-CDL",
                                        record=_deferral_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def extend_deferral(row_id: str, payload: DeferralExtension, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityMelCdlDeferral, amo_id, row_id)
    if row.status not in {"OPEN", "EXTENDED"}:
        raise HTTPException(status_code=409, detail="Only open deferrals can be extended.")
    if _utc(payload.new_expires_at) <= _utc(row.expires_at):
        raise HTTPException(status_code=422, detail="Extended expiry must be later than the current expiry.")
    previous = _utc(row.expires_at).isoformat()
    history = list(row.extension_history_json or [])
    history.append({
        "from": previous,
        "to": _utc(payload.new_expires_at).isoformat(),
        "reason": payload.reason,
        "approval_reference": payload.approval_reference,
        "approved_by_user_id": str(current_user.id),
        "approved_at": _utcnow().isoformat(),
    })
    row.extension_history_json = history
    row.expires_at = payload.new_expires_at
    row.status = "EXTENDED"
    row.revision += 1
    _revision_event(db, amo_id=amo_id, source_type="MEL_CDL_DEFERRAL", source_id=row.id, revision=row.revision,
                    action="EXTENDED", payload=history[-1], actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="MEL-CDL",
                                        record=_deferral_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def close_deferral(row_id: str, payload: CloseRequest, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityMelCdlDeferral, amo_id, row_id)
    if row.status not in {"OPEN", "EXTENDED", "EXPIRED"}:
        raise HTTPException(status_code=409, detail="This deferral cannot be closed from its current state.")
    row.status = "CLOSED"
    row.revision += 1
    row.closed_by_user_id = str(current_user.id)
    row.closed_at = _utcnow()
    row.closure_evidence_json = [{"note": payload.note}, *payload.evidence]
    _revision_event(db, amo_id=amo_id, source_type="MEL_CDL_DEFERRAL", source_id=row.id, revision=row.revision,
                    action="CLOSED", payload={"note": payload.note, "evidence": payload.evidence}, actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="MEL-CDL",
                                        record=_deferral_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def list_shop_findings(limit: int = Query(default=200, ge=1, le=500), context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    return db.query(ReliabilityComponentShopFinding).filter(
        ReliabilityComponentShopFinding.amo_id == amo_id
    ).order_by(ReliabilityComponentShopFinding.inspected_at.desc()).limit(limit).all()


def create_shop_finding(payload: ShopFindingCreate, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    _aircraft(db, amo_id, payload.aircraft_serial_number)
    _component(db, amo_id, payload.component_id, payload.part_number, payload.component_serial_number, payload.aircraft_serial_number)
    row = ReliabilityComponentShopFinding(amo_id=amo_id, **payload.model_dump(), created_by_user_id=str(current_user.id))
    db.add(row)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise _duplicate_error(exc, "shop finding") from exc
    _revision_event(db, amo_id=amo_id, source_type="COMPONENT_SHOP_FINDING", source_id=row.id, revision=1,
                    action="CREATED", payload={"shop_order_reference": row.shop_order_reference}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def approve_shop_finding(row_id: str, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityComponentShopFinding, amo_id, row_id)
    if row.status != "DRAFT":
        raise HTTPException(status_code=409, detail="Only draft shop findings can be approved.")
    row.status = "APPROVED"
    row.revision += 1
    row.approved_by_user_id = str(current_user.id)
    row.approved_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="COMPONENT_SHOP_FINDING", source_id=row.id, revision=row.revision,
                    action="APPROVED", payload={"disposition": row.disposition}, actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="COMPONENT-SHOP-FINDINGS",
                                        record=_shop_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def release_shop_finding(row_id: str, payload: ReleaseRequest, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilityComponentShopFinding, amo_id, row_id)
    if row.status != "APPROVED":
        raise HTTPException(status_code=409, detail="Only approved shop findings can be released.")
    row.status = "RELEASED"
    row.revision += 1
    row.release_reference = payload.release_reference
    row.released_by_user_id = str(current_user.id)
    row.released_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="COMPONENT_SHOP_FINDING", source_id=row.id, revision=row.revision,
                    action="RELEASED", payload=payload.model_dump(), actor_user_id=str(current_user.id))
    _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="COMPONENT-SHOP-FINDINGS",
                                        record=_shop_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def list_sms(limit: int = Query(default=200, ge=1, le=500), context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    return db.query(ReliabilitySmsOccurrence).filter(
        ReliabilitySmsOccurrence.amo_id == amo_id
    ).order_by(ReliabilitySmsOccurrence.occurred_at.desc()).limit(limit).all()


def create_sms(payload: SmsOccurrenceCreate, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    _aircraft(db, amo_id, payload.aircraft_serial_number)
    row = ReliabilitySmsOccurrence(amo_id=amo_id, **payload.model_dump(), created_by_user_id=str(current_user.id))
    db.add(row)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise _duplicate_error(exc, "SMS occurrence") from exc
    _revision_event(db, amo_id=amo_id, source_type="SMS_OCCURRENCE", source_id=row.id, revision=1,
                    action="CREATED", payload={"sms_reference": row.sms_reference}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def assess_sms(row_id: str, payload: SmsAssessment, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    row = _get_row(db, ReliabilitySmsOccurrence, amo_id, row_id)
    if row.status not in {"DRAFT", "ASSESSED"}:
        raise HTTPException(status_code=409, detail="SMS occurrence cannot be reassessed from its current state.")
    row.reliability_relevant = payload.reliability_relevant
    row.reliability_link_reason = payload.reliability_link_reason.strip() if payload.reliability_link_reason else None
    row.investigation_status = payload.investigation_status
    row.status = "ASSESSED"
    row.revision += 1
    row.assessed_by_user_id = str(current_user.id)
    row.assessed_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="SMS_OCCURRENCE", source_id=row.id, revision=row.revision,
                    action="ASSESSED", payload=payload.model_dump(), actor_user_id=str(current_user.id))
    if row.reliability_relevant:
        _, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="SMS-EVENTS",
                                            record=_sms_record(row), actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def _read_csv(content: bytes, header_row: int) -> Iterable[tuple[str, int, Dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    if len(rows) < header_row:
        return []
    headers = [str(value).strip() or f"COLUMN_{index + 1}" for index, value in enumerate(rows[header_row - 1])]
    output = []
    for source_row, values in enumerate(rows[header_row:], start=header_row + 1):
        if not any(str(value).strip() for value in values):
            continue
        output.append(("CSV", source_row, {headers[index]: _json_value(value) for index, value in enumerate(values) if index < len(headers)}))
    return output


def _read_xlsx(content: bytes, header_row: int) -> Iterable[tuple[str, int, Dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    output: List[tuple[str, int, Dict[str, Any]]] = []
    try:
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue
            header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
            if not header_values:
                continue
            headers = [str(value).strip() if value not in (None, "") else f"COLUMN_{index + 1}" for index, value in enumerate(header_values)]
            for source_row, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not any(value not in (None, "") for value in values):
                    continue
                raw = {headers[index]: _json_value(value) for index, value in enumerate(values) if index < len(headers)}
                output.append((sheet.title, source_row, raw))
    finally:
        workbook.close()
    return output


async def upload_workbook(
    file: UploadFile = File(...),
    header_row: int = Form(default=1, ge=1, le=100),
    context=Depends(_context),
):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    filename = _safe_filename(file.filename or "workbook.xlsx")
    extension = Path(filename).suffix.lower()
    if extension not in WORKBOOK_EXTENSIONS:
        raise HTTPException(status_code=415, detail="Only .xlsx and .csv Reliability workbooks are accepted.")
    content = await file.read(WORKBOOK_MAX_BYTES + 1)
    if not content:
        raise HTTPException(status_code=422, detail="Workbook is empty.")
    if len(content) > WORKBOOK_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Workbook exceeds the configured upload limit.")
    content_hash = hashlib.sha256(content).hexdigest()
    existing = db.query(ReliabilityWorkbookImport.id).filter(
        ReliabilityWorkbookImport.amo_id == amo_id,
        ReliabilityWorkbookImport.content_hash == content_hash,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="This workbook has already been registered for the tenant.")
    try:
        parsed_rows = list(_read_csv(content, header_row) if extension == ".csv" else _read_xlsx(content, header_row))
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Workbook could not be parsed: {exc}") from exc
    if not parsed_rows:
        raise HTTPException(status_code=422, detail="Workbook contains no data rows after the selected header row.")
    import_id = generate_uuid7()
    directory = (WORKBOOK_ROOT / amo_id / import_id).resolve()
    if not str(directory).startswith(str(WORKBOOK_ROOT)):
        raise HTTPException(status_code=400, detail="Invalid workbook storage path.")
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / filename
    path.write_bytes(content)
    item = ReliabilityWorkbookImport(
        id=import_id,
        amo_id=amo_id,
        original_filename=filename,
        content_hash=content_hash,
        storage_path=str(path),
        header_row=header_row,
        total_rows=len(parsed_rows),
        created_by_user_id=str(current_user.id),
    )
    db.add(item)
    db.flush()
    for sheet_name, source_row_number, raw in parsed_rows:
        row_hash = hashlib.sha256(json.dumps(raw, sort_keys=True, default=str).encode("utf-8")).hexdigest()
        db.add(ReliabilityWorkbookRow(
            amo_id=amo_id,
            import_id=item.id,
            sheet_name=sheet_name,
            source_row_number=source_row_number,
            row_hash=row_hash,
            raw_json=raw,
        ))
    _revision_event(db, amo_id=amo_id, source_type="WORKBOOK_IMPORT", source_id=item.id, revision=1,
                    action="UPLOADED", payload={"filename": filename, "rows": len(parsed_rows), "sha256": content_hash}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(item)
    return item


def list_workbooks(limit: int = Query(default=100, ge=1, le=500), context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    return db.query(ReliabilityWorkbookImport).filter(
        ReliabilityWorkbookImport.amo_id == amo_id
    ).order_by(ReliabilityWorkbookImport.created_at.desc()).limit(limit).all()


def list_workbook_rows(
    import_id: str,
    row_status: Optional[str] = Query(default=None, alias="status"),
    limit: int = Query(default=500, ge=1, le=2000),
    context=Depends(_context),
):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.read")
    _get_row(db, ReliabilityWorkbookImport, amo_id, import_id)
    query = db.query(ReliabilityWorkbookRow).filter(
        ReliabilityWorkbookRow.amo_id == amo_id,
        ReliabilityWorkbookRow.import_id == import_id,
    )
    if row_status:
        query = query.filter(ReliabilityWorkbookRow.status == row_status.upper())
    return query.order_by(ReliabilityWorkbookRow.sheet_name, ReliabilityWorkbookRow.source_row_number).limit(limit).all()


def _mapped_value(raw: Dict[str, Any], mapping: Dict[str, str], defaults: Dict[str, Any], field: str) -> Any:
    column = mapping.get(field)
    return raw.get(column) if column else defaults.get(field)


def map_workbook(import_id: str, payload: WorkbookMapRequest, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    item = _get_row(db, ReliabilityWorkbookImport, amo_id, import_id)
    if item.status in {"APPROVED", "INGESTED"}:
        raise HTTPException(status_code=409, detail="Approved or ingested workbooks cannot be remapped.")
    rows = db.query(ReliabilityWorkbookRow).filter(
        ReliabilityWorkbookRow.amo_id == amo_id,
        ReliabilityWorkbookRow.import_id == import_id,
    ).all()
    valid = invalid = 0
    for row in rows:
        raw = dict(row.raw_json or {})
        mapped = {
            key: _mapped_value(raw, payload.mapping, payload.defaults, key)
            for key in set(payload.mapping) | set(payload.defaults)
        }
        mapped.update({
            "source_record_id": row.id,
            "source_revision": str(item.revision + 1),
            "source_workbook": item.original_filename,
            "source_sheet": row.sheet_name,
            "source_row_number": row.source_row_number,
            "mapping_profile": f"workbook:{item.id}:r{item.revision + 1}",
            "reconciliation_status": "MAPPED",
            "reconciliation_note": "Mapped through controlled Reliability workbook reconciliation.",
        })
        errors: List[str] = []
        try:
            historical = _ADAPTERS.HistoricalOccurrence.model_validate(mapped)
            mapped = historical.model_dump(mode="json")
            _aircraft(db, amo_id, historical.aircraft_serial_number)
        except Exception as exc:
            errors.append(str(exc))
        duplicate = db.query(ReliabilityWorkbookRow.id).filter(
            ReliabilityWorkbookRow.amo_id == amo_id,
            ReliabilityWorkbookRow.row_hash == row.row_hash,
            ReliabilityWorkbookRow.import_id != import_id,
            ReliabilityWorkbookRow.status.in_(["APPROVED", "INGESTED"]),
        ).first()
        if duplicate:
            errors.append("Duplicate historical row already approved or ingested in another workbook.")
        row.mapped_json = mapped
        row.validation_errors_json = errors
        row.status = "INVALID" if errors else "VALID"
        if errors:
            invalid += 1
        else:
            valid += 1
    item.mapping_json = payload.mapping
    item.defaults_json = payload.defaults
    item.valid_rows = valid
    item.invalid_rows = invalid
    item.approved_rows = 0
    item.rejected_rows = 0
    item.status = "MAPPED"
    item.revision += 1
    _revision_event(db, amo_id=amo_id, source_type="WORKBOOK_IMPORT", source_id=item.id, revision=item.revision,
                    action="MAPPED", payload={"mapping": payload.mapping, "valid": valid, "invalid": invalid}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(item)
    return item


def decide_workbook_row(import_id: str, row_id: str, payload: WorkbookDecision, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    item = _get_row(db, ReliabilityWorkbookImport, amo_id, import_id)
    if item.status not in {"MAPPED", "IN_REVIEW"}:
        raise HTTPException(status_code=409, detail="Workbook is not open for row reconciliation.")
    row = db.query(ReliabilityWorkbookRow).filter(
        ReliabilityWorkbookRow.amo_id == amo_id,
        ReliabilityWorkbookRow.import_id == import_id,
        ReliabilityWorkbookRow.id == row_id,
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Workbook row not found.")
    if payload.decision == "APPROVE" and row.status not in {"VALID", "APPROVED"}:
        raise HTTPException(status_code=422, detail="Only validated rows can be approved.")
    row.status = "APPROVED" if payload.decision == "APPROVE" else "REJECTED"
    row.decision_note = payload.note
    row.decided_by_user_id = str(current_user.id)
    row.decided_at = _utcnow()
    item.status = "IN_REVIEW"
    item.revision += 1
    item.approved_rows = db.query(func.count(ReliabilityWorkbookRow.id)).filter(
        ReliabilityWorkbookRow.import_id == import_id,
        ReliabilityWorkbookRow.status == "APPROVED",
    ).scalar() or 0
    item.rejected_rows = db.query(func.count(ReliabilityWorkbookRow.id)).filter(
        ReliabilityWorkbookRow.import_id == import_id,
        ReliabilityWorkbookRow.status == "REJECTED",
    ).scalar() or 0
    _revision_event(db, amo_id=amo_id, source_type="WORKBOOK_IMPORT", source_id=item.id, revision=item.revision,
                    action="ROW_DECISION", payload={"row_id": row.id, **payload.model_dump()}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(row)
    return row


def approve_workbook(import_id: str, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    item = _get_row(db, ReliabilityWorkbookImport, amo_id, import_id)
    if item.status not in {"MAPPED", "IN_REVIEW"}:
        raise HTTPException(status_code=409, detail="Workbook is not open for approval.")
    counts = {str(key): int(value or 0) for key, value in db.query(
        ReliabilityWorkbookRow.status, func.count(ReliabilityWorkbookRow.id)
    ).filter(ReliabilityWorkbookRow.import_id == import_id).group_by(ReliabilityWorkbookRow.status).all()}
    unresolved = sum(counts.get(key, 0) for key in ("PENDING", "VALID", "INVALID"))
    if unresolved:
        raise HTTPException(status_code=422, detail="Every workbook row must be approved or rejected before import approval.")
    if counts.get("APPROVED", 0) == 0:
        raise HTTPException(status_code=422, detail="At least one validated row must be approved.")
    item.status = "APPROVED"
    item.revision += 1
    item.approved_rows = counts.get("APPROVED", 0)
    item.rejected_rows = counts.get("REJECTED", 0)
    item.approved_by_user_id = str(current_user.id)
    item.approved_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="WORKBOOK_IMPORT", source_id=item.id, revision=item.revision,
                    action="APPROVED", payload={"approved_rows": item.approved_rows, "rejected_rows": item.rejected_rows}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(item)
    return item


def ingest_workbook(import_id: str, context=Depends(_context)):
    current_user, db, amo_id = context
    _require(db, current_user, "reliability.ingest")
    item = _get_row(db, ReliabilityWorkbookImport, amo_id, import_id)
    if item.status != "APPROVED":
        raise HTTPException(status_code=409, detail="Workbook must be approved before ingestion.")
    rows = db.query(ReliabilityWorkbookRow).filter(
        ReliabilityWorkbookRow.amo_id == amo_id,
        ReliabilityWorkbookRow.import_id == import_id,
        ReliabilityWorkbookRow.status == "APPROVED",
    ).order_by(ReliabilityWorkbookRow.sheet_name, ReliabilityWorkbookRow.source_row_number).all()
    source = _source(_ADAPTERS, db, amo_id, "WORKBOOK-HISTORY", str(current_user.id))
    records = []
    external_to_row: Dict[str, ReliabilityWorkbookRow] = {}
    for row in rows:
        values = dict(row.mapped_json or {})
        values.update({
            "source_revision": str(item.revision),
            "reconciliation_status": "APPROVED",
            "reconciliation_note": row.decision_note or "Approved through row reconciliation.",
        })
        historical = _ADAPTERS.HistoricalOccurrence.model_validate(values)
        record = _ADAPTERS._record_payload("WORKBOOK-HISTORY", historical)
        records.append(record)
        external_to_row[str(record["external_id"])] = row
    result = services.ingest_batch(
        db,
        amo_id=amo_id,
        source=source,
        payload=schemas.ReliabilityBatchIngest(
            records=records,
            metadata_json={
                "adapter": "historical-workbook-v1",
                "workbook_import_id": item.id,
                "workbook_sha256": item.content_hash,
                "mapping": item.mapping_json,
            },
        ),
        actor_user_id=str(current_user.id),
    )
    ingestion_rows = db.query(domain.ReliabilityIngestionRecord).filter(
        domain.ReliabilityIngestionRecord.amo_id == amo_id,
        domain.ReliabilityIngestionRecord.source_id == source.id,
        domain.ReliabilityIngestionRecord.external_id.in_(list(external_to_row)),
    ).all()
    for ingestion in ingestion_rows:
        workbook_row = external_to_row.get(ingestion.external_id)
        if workbook_row:
            workbook_row.status = "INGESTED"
            workbook_row.canonical_event_id = ingestion.normalized_event_id
    item.status = "INGESTED"
    item.revision += 1
    item.ingested_rows = len(ingestion_rows)
    item.ingested_by_user_id = str(current_user.id)
    item.ingested_at = _utcnow()
    _revision_event(db, amo_id=amo_id, source_type="WORKBOOK_IMPORT", source_id=item.id, revision=item.revision,
                    action="INGESTED", payload={"rows": item.ingested_rows, "batch_id": result.batch.id}, actor_user_id=str(current_user.id))
    db.commit()
    db.refresh(item)
    return item


def _source_stats(db: Session, amo_id: str, model: Any, *, relevant_filter: Any = None):
    query = db.query(func.count(model.id), func.max(model.updated_at if hasattr(model, "updated_at") else model.created_at)).filter(model.amo_id == amo_id)
    if relevant_filter is not None:
        query = query.filter(relevant_filter)
    count, latest = query.one()
    return int(count or 0), latest


def _readiness(db: Session, *, amo_id: str):
    response = _ORIGINAL_READINESS(db, amo_id=amo_id)
    stats = {
        "FLIGHT-OPERATIONS": _source_stats(db, amo_id, ReliabilityFlightOperation, relevant_filter=ReliabilityFlightOperation.status.in_(["APPROVED", "CLOSED"])),
        "MEL-CDL": _source_stats(db, amo_id, ReliabilityMelCdlDeferral, relevant_filter=ReliabilityMelCdlDeferral.status.in_(["OPEN", "EXTENDED", "EXPIRED", "CLOSED"])),
        "COMPONENT-SHOP-FINDINGS": _source_stats(db, amo_id, ReliabilityComponentShopFinding, relevant_filter=ReliabilityComponentShopFinding.status.in_(["APPROVED", "RELEASED"])),
        "SMS-EVENTS": _source_stats(db, amo_id, ReliabilitySmsOccurrence, relevant_filter=ReliabilitySmsOccurrence.reliability_relevant.is_(True)),
        "WORKBOOK-HISTORY": _source_stats(db, amo_id, ReliabilityWorkbookRow, relevant_filter=ReliabilityWorkbookRow.status.in_(["APPROVED", "INGESTED"])),
    }
    for item in response.items:
        if item.code not in stats:
            continue
        available, latest = stats[item.code]
        item.available_record_count = available
        item.latest_available_at = latest
        if not item.source_id:
            item.connection_state = "CONFIGURATION_REQUIRED"
        elif available == 0:
            item.connection_state = "NO_DATA"
        elif item.ingested_record_count >= available and item.last_success_at:
            item.connection_state = "WIRED"
        else:
            item.connection_state = "SYNC_REQUIRED"
        item.detail = "Authoritative source register, revision history, approval controls and canonical ingestion are active."
    return response


def _harvest_operational(db: Session, *, amo_id: str, actor_user_id: str):
    results = []
    now = _utcnow()
    expiring = db.query(ReliabilityMelCdlDeferral).filter(
        ReliabilityMelCdlDeferral.amo_id == amo_id,
        ReliabilityMelCdlDeferral.status.in_(["OPEN", "EXTENDED"]),
        ReliabilityMelCdlDeferral.expires_at < now,
    ).all()
    for row in expiring:
        row.status = "EXPIRED"
        row.revision += 1
        _revision_event(db, amo_id=amo_id, source_type="MEL_CDL_DEFERRAL", source_id=row.id, revision=row.revision,
                        action="EXPIRED", payload={"expires_at": _utc(row.expires_at).isoformat()}, actor_user_id=actor_user_id)
        result, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="MEL-CDL",
                                                 record=_deferral_record(row), actor_user_id=actor_user_id)
        results.append(result)
    pending_sets = [
        (ReliabilityFlightOperation, "FLIGHT-OPERATIONS", _flight_record, ["APPROVED", "CLOSED"]),
        (ReliabilityMelCdlDeferral, "MEL-CDL", _deferral_record, ["OPEN", "EXTENDED", "EXPIRED", "CLOSED"]),
        (ReliabilityComponentShopFinding, "COMPONENT-SHOP-FINDINGS", _shop_record, ["APPROVED", "RELEASED"]),
    ]
    for model, code, builder, statuses in pending_sets:
        rows = db.query(model).filter(model.amo_id == amo_id, model.status.in_(statuses), model.canonical_event_id.is_(None)).all()
        for row in rows:
            result, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code=code,
                                                     record=builder(row), actor_user_id=actor_user_id)
            results.append(result)
    sms_rows = db.query(ReliabilitySmsOccurrence).filter(
        ReliabilitySmsOccurrence.amo_id == amo_id,
        ReliabilitySmsOccurrence.status == "ASSESSED",
        ReliabilitySmsOccurrence.reliability_relevant.is_(True),
        ReliabilitySmsOccurrence.canonical_event_id.is_(None),
    ).all()
    for row in sms_rows:
        result, row.canonical_event_id = _ingest(_ADAPTERS, db, amo_id=amo_id, code="SMS-EVENTS",
                                                 record=_sms_record(row), actor_user_id=actor_user_id)
        results.append(result)
    db.commit()
    return results


def activate_authoritative_adapters(adapters: Any) -> None:
    global _ADAPTERS, _ORIGINAL_READINESS
    _ADAPTERS = adapters
    for spec in adapters.ADAPTER_SPECS:
        if spec.code == "FLIGHT-OPERATIONS":
            spec.module = "Reliability / Flight Operations"
            spec.authoritative_tables = ["reliability_flight_operations", "reliability_source_revision_events"]
        elif spec.code == "MEL-CDL":
            spec.module = "Reliability / Defect Control"
            spec.authoritative_tables = ["reliability_mel_cdl_deferrals", "reliability_source_revision_events"]
        elif spec.code == "COMPONENT-SHOP-FINDINGS":
            spec.module = "Reliability / Component Shop"
            spec.authoritative_tables = ["reliability_component_shop_findings", "reliability_source_revision_events"]
        elif spec.code == "SMS-EVENTS":
            spec.module = "Reliability / Safety Occurrences"
            spec.authoritative_tables = ["reliability_sms_occurrences", "reliability_source_revision_events"]
        elif spec.code == "WORKBOOK-HISTORY":
            spec.module = "Reliability / Data Migration"
            spec.authoritative_tables = ["reliability_workbook_imports", "reliability_workbook_rows"]
        else:
            continue
        spec.connection_mode = "INTERNAL_LINK"
        spec.implementation_state = "READY"
        spec.detail = "Authoritative register and controlled canonical ingestion are available."
    _ORIGINAL_READINESS = adapters.adapter_readiness
    adapters.adapter_readiness = _readiness
    original_harvest = services.harvest_internal_sources
    if not getattr(original_harvest, "_operational_sources_wrapper", False):
        def harvest_with_operational(db: Session, *, amo_id: str, actor_user_id: Optional[str]):
            results = list(original_harvest(db, amo_id=amo_id, actor_user_id=actor_user_id))
            if actor_user_id:
                results.extend(_harvest_operational(db, amo_id=amo_id, actor_user_id=str(actor_user_id)))
            return results
        setattr(harvest_with_operational, "_operational_sources_wrapper", True)
        services.harvest_internal_sources = harvest_with_operational


_ADAPTERS: Any = None
_ORIGINAL_READINESS: Any = None
_REGISTERED = False


def register(router: APIRouter) -> None:
    global _REGISTERED
    if _REGISTERED:
        return
    install_exact_aviation_types()
    _ADAPTERS._install_workpack_decorator()
    router.add_api_route(
        "/authoritative-sources/configure",
        _ADAPTERS.configure_route,
        methods=["POST"],
        response_model=_ADAPTERS.AdapterReadinessResponse,
        summary="Configure authoritative Reliability operational sources",
    )
    router.add_api_route(
        "/authoritative-sources/readiness",
        _ADAPTERS.readiness_route,
        methods=["GET"],
        response_model=_ADAPTERS.AdapterReadinessResponse,
        summary="Show authoritative Reliability operational source readiness",
    )
    router.add_api_route(
        "/authoritative-sources/qms/findings/{finding_id}/link",
        _ADAPTERS.qms_link_route,
        methods=["POST"],
        response_model=schemas.ReliabilityIngestionResult,
        status_code=status.HTTP_201_CREATED,
        summary="Link a selected QMS finding to Reliability",
    )
    prefix = "/operational-sources"
    router.add_api_route(f"{prefix}/summary", operational_summary, methods=["GET"], response_model=OperationalSummary)
    router.add_api_route(f"{prefix}/flight-operations", list_flight_operations, methods=["GET"], response_model=List[FlightOperationRead])
    router.add_api_route(f"{prefix}/flight-operations", create_flight_operation, methods=["POST"], response_model=FlightOperationRead, status_code=status.HTTP_201_CREATED)
    router.add_api_route(f"{prefix}/flight-operations/{{row_id}}/approve", approve_flight_operation, methods=["POST"], response_model=FlightOperationRead)
    router.add_api_route(f"{prefix}/flight-operations/{{row_id}}/close", close_flight_operation, methods=["POST"], response_model=FlightOperationRead)
    router.add_api_route(f"{prefix}/deferrals", list_deferrals, methods=["GET"], response_model=List[DeferralRead])
    router.add_api_route(f"{prefix}/deferrals", create_deferral, methods=["POST"], response_model=DeferralRead, status_code=status.HTTP_201_CREATED)
    router.add_api_route(f"{prefix}/deferrals/{{row_id}}/approve", approve_deferral, methods=["POST"], response_model=DeferralRead)
    router.add_api_route(f"{prefix}/deferrals/{{row_id}}/extend", extend_deferral, methods=["POST"], response_model=DeferralRead)
    router.add_api_route(f"{prefix}/deferrals/{{row_id}}/close", close_deferral, methods=["POST"], response_model=DeferralRead)
    router.add_api_route(f"{prefix}/component-shop", list_shop_findings, methods=["GET"], response_model=List[ShopFindingRead])
    router.add_api_route(f"{prefix}/component-shop", create_shop_finding, methods=["POST"], response_model=ShopFindingRead, status_code=status.HTTP_201_CREATED)
    router.add_api_route(f"{prefix}/component-shop/{{row_id}}/approve", approve_shop_finding, methods=["POST"], response_model=ShopFindingRead)
    router.add_api_route(f"{prefix}/component-shop/{{row_id}}/release", release_shop_finding, methods=["POST"], response_model=ShopFindingRead)
    router.add_api_route(f"{prefix}/sms", list_sms, methods=["GET"], response_model=List[SmsOccurrenceRead])
    router.add_api_route(f"{prefix}/sms", create_sms, methods=["POST"], response_model=SmsOccurrenceRead, status_code=status.HTTP_201_CREATED)
    router.add_api_route(f"{prefix}/sms/{{row_id}}/assess", assess_sms, methods=["POST"], response_model=SmsOccurrenceRead)
    router.add_api_route(f"{prefix}/workbooks", list_workbooks, methods=["GET"], response_model=List[WorkbookImportRead])
    router.add_api_route(f"{prefix}/workbooks/upload", upload_workbook, methods=["POST"], response_model=WorkbookImportRead, status_code=status.HTTP_201_CREATED)
    router.add_api_route(f"{prefix}/workbooks/{{import_id}}/rows", list_workbook_rows, methods=["GET"], response_model=List[WorkbookRowRead])
    router.add_api_route(f"{prefix}/workbooks/{{import_id}}/map", map_workbook, methods=["POST"], response_model=WorkbookImportRead)
    router.add_api_route(f"{prefix}/workbooks/{{import_id}}/rows/{{row_id}}/decision", decide_workbook_row, methods=["POST"], response_model=WorkbookRowRead)
    router.add_api_route(f"{prefix}/workbooks/{{import_id}}/approve", approve_workbook, methods=["POST"], response_model=WorkbookImportRead)
    router.add_api_route(f"{prefix}/workbooks/{{import_id}}/ingest", ingest_workbook, methods=["POST"], response_model=WorkbookImportRead)
    _REGISTERED = True
