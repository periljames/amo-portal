from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import BigInteger, CheckConstraint, Column, DateTime, ForeignKey, Index, Integer, String, Text, UniqueConstraint, func
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from amodb.apps.accounts import models as account_models
from amodb.database import Base, get_write_db
from amodb.security import get_current_active_user

from . import workbook_parity as wp
from .workbook_parity_defaults import WORKBOOK_PROFILES

UTC = timezone.utc
MAX_UPLOAD_BYTES = 25 * 1024 * 1024
MAX_PREVIEW_ROWS = 10_000
MAX_COMMIT_CHUNK = 250
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}
ALLOWED_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/octet-stream",
}
IMPORT_ROLES = {"SUPERUSER", "AMO_ADMIN", "QUALITY_MANAGER", "SAFETY_MANAGER", "PLANNING_ENGINEER"}


class ReliabilityWorkbookImportBatch(Base):
    __tablename__ = "reliability_workbook_import_batches"
    __table_args__ = (
        UniqueConstraint("amo_id", "profile_code", "dataset_code", "selected_sheet", "source_hash", name="uq_rel_workbook_import_source"),
        CheckConstraint("file_size_bytes > 0", name="ck_rel_workbook_import_size_positive"),
        CheckConstraint("header_row > 0", name="ck_rel_workbook_import_header_positive"),
        CheckConstraint("status IN ('PREVIEW_READY','PROCESSING','COMPLETED','PARTIAL_FAILED','FAILED')", name="ck_rel_workbook_import_status"),
        Index("ix_rel_workbook_import_scope", "amo_id", "status", "created_at"),
        Index("ix_rel_workbook_import_hash", "amo_id", "source_hash"),
    )

    id = Column(Integer, primary_key=True)
    amo_id = Column(String(36), ForeignKey("amos.id", ondelete="CASCADE"), nullable=False, index=True)
    profile_code = Column(String(64), nullable=False, index=True)
    dataset_code = Column(String(24), nullable=False, index=True)
    original_filename = Column(String(255), nullable=False)
    sanitized_filename = Column(String(255), nullable=False)
    file_extension = Column(String(8), nullable=False)
    file_size_bytes = Column(BigInteger, nullable=False)
    source_hash = Column(String(64), nullable=False, index=True)
    status = Column(String(32), nullable=False, default="PREVIEW_READY", index=True)
    detected_sheets = Column(JSONB, nullable=False, default=list)
    selected_sheet = Column(String(128), nullable=False)
    header_row = Column(Integer, nullable=False, default=1)
    header_map = Column(JSONB, nullable=False, default=dict)
    total_rows = Column(Integer, nullable=False, default=0)
    valid_rows = Column(Integer, nullable=False, default=0)
    invalid_rows = Column(Integer, nullable=False, default=0)
    committed_rows = Column(Integer, nullable=False, default=0)
    failed_rows = Column(Integer, nullable=False, default=0)
    created_by_user_id = Column(String(36), ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=lambda: datetime.now(UTC))
    updated_at = Column(DateTime(timezone=True), nullable=False, default=lambda: datetime.now(UTC), onupdate=lambda: datetime.now(UTC))
    completed_at = Column(DateTime(timezone=True), nullable=True)


class ReliabilityWorkbookImportRowResult(Base):
    __tablename__ = "reliability_workbook_import_row_results"
    __table_args__ = (
        UniqueConstraint("batch_id", "row_number", name="uq_rel_workbook_import_row"),
        CheckConstraint("row_number > 0", name="ck_rel_workbook_import_row_positive"),
        CheckConstraint("attempt_count >= 0", name="ck_rel_workbook_import_attempt_nonnegative"),
        CheckConstraint("status IN ('VALID','INVALID','COMMITTED','FAILED')", name="ck_rel_workbook_import_row_status"),
        Index("ix_rel_workbook_import_row_queue", "batch_id", "status", "row_number"),
        Index("ix_rel_workbook_import_row_hash", "row_source_hash"),
    )

    id = Column(Integer, primary_key=True)
    batch_id = Column(Integer, ForeignKey("reliability_workbook_import_batches.id", ondelete="CASCADE"), nullable=False, index=True)
    row_number = Column(Integer, nullable=False)
    row_source_hash = Column(String(64), nullable=False, index=True)
    raw_values = Column(JSONB, nullable=False, default=dict)
    mapped_values = Column(JSONB, nullable=False, default=dict)
    errors = Column(JSONB, nullable=False, default=list)
    status = Column(String(24), nullable=False, default="VALID", index=True)
    attempt_count = Column(Integer, nullable=False, default=0)
    last_error = Column(Text, nullable=True)
    workbook_record_id = Column(Integer, ForeignKey("reliability_workbook_records.id", ondelete="SET NULL"), nullable=True, index=True)
    created_at = Column(DateTime(timezone=True), nullable=False, default=lambda: datetime.now(UTC))
    updated_at = Column(DateTime(timezone=True), nullable=False, default=lambda: datetime.now(UTC), onupdate=lambda: datetime.now(UTC))


class ImportCommitRequest(BaseModel):
    chunk_size: int = Field(default=100, ge=1, le=MAX_COMMIT_CHUNK)


class ImportRetryRequest(BaseModel):
    failed_only: bool = True


def _amo_id(user: account_models.User) -> str:
    amo_id = user.effective_amo_id
    if not amo_id:
        raise HTTPException(status_code=403, detail="A tenant context is required.")
    return str(amo_id)


def _require_import_permission(user: account_models.User) -> None:
    role = str(getattr(getattr(user, "role", None), "value", getattr(user, "role", ""))).upper()
    if bool(getattr(user, "is_superuser", False)) or bool(getattr(user, "is_amo_admin", False)) or role in IMPORT_ROLES:
        return
    raise HTTPException(status_code=403, detail="Workbook import requires Reliability data-governance permission.")


def _sanitize_filename(filename: str) -> str:
    base = Path(filename or "workbook.xlsx").name
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._")
    return (cleaned or "workbook.xlsx")[:255]


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").strip().upper()).strip()


COMMON_ALIASES: dict[str, tuple[str, ...]] = {
    "event_date": ("EVENT DATE", "DATE", "REPORT DATE", "OCCURRENCE DATE", "PERIOD DATE"),
    "event_end_date": ("END DATE", "EVENT END DATE", "TO DATE"),
    "aircraft_serial_number": ("AIRCRAFT", "AIRCRAFT REGISTRATION", "REGISTRATION", "TAIL", "TAIL NUMBER", "A C REG"),
    "ata_chapter": ("ATA", "ATA CHAPTER", "ATA CODE"),
    "reference_code": ("REFERENCE", "REFERENCE CODE", "REF", "REPORT NUMBER", "TECH LOG REF"),
    "title": ("TITLE", "SUBJECT", "EVENT TITLE"),
    "description": ("DESCRIPTION", "REMARKS", "NARRATIVE"),
}


def _candidate_aliases(
    db: Session,
    amo_id: str,
    profile_code: str,
    dataset_code: wp.WorkbookDatasetCode,
    selected_sheet: str,
) -> dict[str, set[str]]:
    definition = wp.DATASET_CATALOG[dataset_code]
    aliases: dict[str, set[str]] = {
        canonical: {_normalise_header(canonical), *(_normalise_header(value) for value in values)}
        for canonical, values in COMMON_ALIASES.items()
    }
    for field in definition.fields:
        aliases[field.key] = {_normalise_header(field.key), _normalise_header(field.label)}
    rows = db.query(wp.ReliabilityWorkbookFieldMapping).filter(
        wp.ReliabilityWorkbookFieldMapping.amo_id == amo_id,
        wp.ReliabilityWorkbookFieldMapping.profile_code == profile_code,
        wp.ReliabilityWorkbookFieldMapping.dataset_code == dataset_code.value,
        wp.ReliabilityWorkbookFieldMapping.active.is_(True),
    ).all()
    for row in rows:
        if _normalise_header(row.source_sheet) not in {_normalise_header(selected_sheet), *(_normalise_header(name) for name in definition.workbook_sheet_names)}:
            continue
        bucket = aliases.setdefault(row.canonical_field, set())
        bucket.add(_normalise_header(row.source_column))
        bucket.update(_normalise_header(value) for value in (row.aliases or []))
    return aliases


def _match_headers(headers: list[Any], aliases: dict[str, set[str]]) -> tuple[dict[str, str], list[str]]:
    header_map: dict[str, str] = {}
    errors: list[str] = []
    used_fields: dict[str, int] = {}
    for index, raw in enumerate(headers, start=1):
        header = _normalise_header(raw)
        if not header:
            continue
        candidates = sorted(canonical for canonical, values in aliases.items() if header in values)
        if len(candidates) > 1:
            errors.append(f"Column {index} '{raw}' is ambiguous: {', '.join(candidates)}.")
            continue
        if not candidates:
            continue
        canonical = candidates[0]
        if canonical in used_fields:
            errors.append(f"Columns {used_fields[canonical]} and {index} both map to '{canonical}'.")
            continue
        used_fields[canonical] = index
        header_map[str(index)] = canonical
    return header_map, errors


def _date_value(value: Any, label: str, *, allow_datetime: bool = False) -> str:
    if isinstance(value, datetime):
        return value.isoformat() if allow_datetime else value.date().isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).replace(tzinfo=UTC).isoformat() if allow_datetime else value.isoformat()
    text = str(value or "").strip()
    try:
        if allow_datetime:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).isoformat()
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise ValueError(f"{label} must be an ISO date{'-time' if allow_datetime else ''}.") from exc


def _coerce_value(value: Any, data_type: str, label: str, options: list[str] | None = None) -> Any:
    if value in (None, ""):
        return None
    if data_type == "date":
        return _date_value(value, label)
    if data_type == "datetime":
        return _date_value(value, label, allow_datetime=True)
    if data_type == "decimal":
        try:
            number = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ValueError(f"{label} must be a valid number.") from exc
        if not number.is_finite():
            raise ValueError(f"{label} must be finite.")
        return format(number, "f")
    if data_type == "integer":
        try:
            number = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ValueError(f"{label} must be a whole number.") from exc
        if number != number.to_integral_value():
            raise ValueError(f"{label} must be a whole number.")
        return int(number)
    if data_type == "boolean":
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in {"true", "yes", "y", "1"}:
            return True
        if text in {"false", "no", "n", "0"}:
            return False
        raise ValueError(f"{label} must be true or false.")
    text = str(value).strip()
    if data_type == "select":
        text = text.upper()
        if options and text not in options:
            raise ValueError(f"{label} must be one of: {', '.join(options)}.")
    return text


def _build_preview_row(
    dataset_code: wp.WorkbookDatasetCode,
    definition: wp.DatasetDefinition,
    row_number: int,
    cells: tuple[Any, ...],
    header_map: dict[str, str],
    batch_hash: str,
    sheet_name: str,
) -> tuple[dict[str, Any], dict[str, Any], list[str], str]:
    raw_values: dict[str, Any] = {}
    mapped_common: dict[str, Any] = {}
    mapped_payload: dict[str, Any] = {}
    errors: list[str] = []
    field_lookup = {field.key: field for field in definition.fields}

    for index_text, canonical in header_map.items():
        index = int(index_text)
        if index > len(cells):
            continue
        cell = cells[index - 1]
        value = cell.value
        raw_values[str(index)] = value.isoformat() if isinstance(value, (date, datetime)) else value
        if getattr(cell, "data_type", None) == "f" or (isinstance(value, str) and value.startswith("=")):
            errors.append(f"Column {index} contains a formula. Formula results are not imported as controlled values.")
            continue
        if getattr(cell, "data_type", None) == "e":
            errors.append(f"Column {index} contains a workbook formula error.")
            continue
        try:
            if canonical in field_lookup:
                field = field_lookup[canonical]
                mapped_payload[canonical] = _coerce_value(value, field.data_type, field.label, field.options)
            elif canonical in {"event_date", "event_end_date"}:
                mapped_common[canonical] = _coerce_value(value, "date", canonical.replace("_", " ").title())
            else:
                mapped_common[canonical] = _coerce_value(value, "text", canonical.replace("_", " ").title())
        except ValueError as exc:
            errors.append(str(exc))

    required_common = {"event_date"}
    if dataset_code != wp.WorkbookDatasetCode.AI:
        required_common.add("aircraft_serial_number")
    for canonical in sorted(required_common):
        if mapped_common.get(canonical) in (None, ""):
            errors.append(f"{canonical.replace('_', ' ').title()} is required.")
    for field in definition.fields:
        if field.required and mapped_payload.get(field.key) in (None, ""):
            errors.append(f"{field.label} is required.")

    title = str(mapped_common.get("title") or mapped_common.get("reference_code") or f"{dataset_code.value} imported row {row_number}")
    mapped = {
        "dataset_code": dataset_code.value,
        "event_date": mapped_common.get("event_date"),
        "event_end_date": mapped_common.get("event_end_date"),
        "aircraft_serial_number": mapped_common.get("aircraft_serial_number"),
        "ata_chapter": mapped_common.get("ata_chapter"),
        "reference_code": mapped_common.get("reference_code"),
        "title": title,
        "description": mapped_common.get("description"),
        "payload": mapped_payload,
    }
    row_hash = hashlib.sha256(json.dumps({"batch": batch_hash, "sheet": sheet_name, "row": row_number, "mapped": mapped}, sort_keys=True, default=str).encode()).hexdigest()
    return raw_values, mapped, list(dict.fromkeys(errors)), row_hash


def _batch_dict(batch: ReliabilityWorkbookImportBatch) -> dict[str, Any]:
    return {
        "id": batch.id,
        "profile_code": batch.profile_code,
        "dataset_code": batch.dataset_code,
        "original_filename": batch.original_filename,
        "sanitized_filename": batch.sanitized_filename,
        "file_extension": batch.file_extension,
        "file_size_bytes": batch.file_size_bytes,
        "source_hash": batch.source_hash,
        "status": batch.status,
        "detected_sheets": batch.detected_sheets,
        "selected_sheet": batch.selected_sheet,
        "header_row": batch.header_row,
        "header_map": batch.header_map,
        "total_rows": batch.total_rows,
        "valid_rows": batch.valid_rows,
        "invalid_rows": batch.invalid_rows,
        "committed_rows": batch.committed_rows,
        "failed_rows": batch.failed_rows,
        "created_at": batch.created_at,
        "updated_at": batch.updated_at,
        "completed_at": batch.completed_at,
    }


def _refresh_counts(db: Session, batch: ReliabilityWorkbookImportBatch) -> None:
    counts = dict(db.query(ReliabilityWorkbookImportRowResult.status, func.count(ReliabilityWorkbookImportRowResult.id)).filter(ReliabilityWorkbookImportRowResult.batch_id == batch.id).group_by(ReliabilityWorkbookImportRowResult.status).all())
    batch.total_rows = sum(int(value) for value in counts.values())
    batch.valid_rows = int(counts.get("VALID", 0))
    batch.invalid_rows = int(counts.get("INVALID", 0))
    batch.committed_rows = int(counts.get("COMMITTED", 0))
    batch.failed_rows = int(counts.get("FAILED", 0))
    if batch.valid_rows > 0:
        batch.status = "PROCESSING" if batch.committed_rows or batch.failed_rows else "PREVIEW_READY"
        batch.completed_at = None
    elif batch.failed_rows > 0:
        batch.status = "PARTIAL_FAILED" if batch.committed_rows else "FAILED"
        batch.completed_at = datetime.now(UTC)
    elif batch.invalid_rows and not batch.committed_rows:
        batch.status = "FAILED"
        batch.completed_at = datetime.now(UTC)
    else:
        batch.status = "COMPLETED"
        batch.completed_at = datetime.now(UTC)


def register(router: APIRouter) -> None:
    @router.post("/workbook-parity/imports/preview", status_code=201)
    async def preview_import(
        profile_code: str = Form(...),
        dataset_code: wp.WorkbookDatasetCode = Form(...),
        source_sheet: str | None = Form(default=None),
        header_row: int = Form(default=1, ge=1, le=100),
        workbook: UploadFile = File(...),
        current_user: account_models.User = Depends(get_current_active_user),
        db: Session = Depends(get_write_db),
    ):
        amo_id = _amo_id(current_user)
        _require_import_permission(current_user)
        profile_codes = {profile["code"] for profile in WORKBOOK_PROFILES}
        if profile_code not in profile_codes:
            raise HTTPException(status_code=422, detail="Unknown workbook profile. Select a controlled profile before preview.")
        sanitized = _sanitize_filename(workbook.filename or "workbook.xlsx")
        extension = Path(sanitized).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=415, detail="Only .xlsx and .xlsm workbooks are accepted.")
        content_type = (workbook.content_type or "").lower()
        if content_type and content_type not in ALLOWED_MIME_TYPES:
            raise HTTPException(status_code=415, detail="The upload MIME type is not an accepted Excel workbook type.")
        content = await workbook.read(MAX_UPLOAD_BYTES + 1)
        if not content:
            raise HTTPException(status_code=422, detail="The uploaded workbook is empty.")
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail="Workbook uploads are limited to 25 MiB.")
        source_hash = hashlib.sha256(content).hexdigest()
        try:
            book = load_workbook(filename=BytesIO(content), read_only=True, data_only=False, keep_vba=False, keep_links=False)
        except Exception as exc:
            raise HTTPException(status_code=422, detail="The workbook could not be parsed safely as an Open XML workbook.") from exc
        detected_sheets = [{"name": sheet.title, "state": sheet.sheet_state, "max_row": sheet.max_row, "max_column": sheet.max_column} for sheet in book.worksheets]
        definition = wp.DATASET_CATALOG[dataset_code]
        if source_sheet:
            if source_sheet not in book.sheetnames:
                raise HTTPException(status_code=422, detail="The selected source sheet is not present in the workbook.")
            selected_sheet = source_sheet
        else:
            expected = {_normalise_header(name) for name in definition.workbook_sheet_names}
            matches = [name for name in book.sheetnames if _normalise_header(name) in expected]
            if len(matches) != 1:
                raise HTTPException(status_code=422, detail={"message": "Workbook profile detection requires review.", "candidate_sheets": matches, "available_sheets": book.sheetnames})
            selected_sheet = matches[0]
        existing = db.query(ReliabilityWorkbookImportBatch).filter(
            ReliabilityWorkbookImportBatch.amo_id == amo_id,
            ReliabilityWorkbookImportBatch.profile_code == profile_code,
            ReliabilityWorkbookImportBatch.dataset_code == dataset_code.value,
            ReliabilityWorkbookImportBatch.selected_sheet == selected_sheet,
            ReliabilityWorkbookImportBatch.source_hash == source_hash,
        ).one_or_none()
        if existing:
            raise HTTPException(status_code=409, detail={"message": "This workbook source has already been previewed for the selected profile, dataset and sheet.", "batch_id": existing.id, "status": existing.status})
        sheet = book[selected_sheet]
        header_cells = next(sheet.iter_rows(min_row=header_row, max_row=header_row), ())
        headers = [cell.value for cell in header_cells]
        aliases = _candidate_aliases(db, amo_id, profile_code, dataset_code, selected_sheet)
        header_map, header_errors = _match_headers(headers, aliases)
        required = {field.key for field in definition.fields if field.required} | {"event_date"}
        if dataset_code != wp.WorkbookDatasetCode.AI:
            required.add("aircraft_serial_number")
        mapped_fields = set(header_map.values())
        missing = sorted(required - mapped_fields)
        if header_errors or missing:
            raise HTTPException(status_code=422, detail={"message": "Workbook headers do not satisfy the controlled mapping.", "header_errors": header_errors, "missing_required_fields": missing, "header_map": header_map, "headers": headers})

        batch = ReliabilityWorkbookImportBatch(
            amo_id=amo_id,
            profile_code=profile_code,
            dataset_code=dataset_code.value,
            original_filename=(workbook.filename or sanitized)[:255],
            sanitized_filename=sanitized,
            file_extension=extension,
            file_size_bytes=len(content),
            source_hash=source_hash,
            status="PREVIEW_READY",
            detected_sheets=detected_sheets,
            selected_sheet=selected_sheet,
            header_row=header_row,
            header_map=header_map,
            created_by_user_id=current_user.id,
        )
        db.add(batch)
        try:
            db.flush()
        except IntegrityError as exc:
            db.rollback()
            raise HTTPException(status_code=409, detail="This workbook source has already been previewed.") from exc

        total = valid = invalid = 0
        for row_number, cells in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            if all(cell.value in (None, "") for cell in cells):
                continue
            total += 1
            if total > MAX_PREVIEW_ROWS:
                db.rollback()
                raise HTTPException(status_code=422, detail="Workbook preview is limited to 10,000 non-empty rows. Split the source into controlled batches.")
            raw, mapped, errors, row_hash = _build_preview_row(dataset_code, definition, row_number, cells, header_map, source_hash, selected_sheet)
            status = "INVALID" if errors else "VALID"
            valid += int(status == "VALID")
            invalid += int(status == "INVALID")
            db.add(ReliabilityWorkbookImportRowResult(batch_id=batch.id, row_number=row_number, row_source_hash=row_hash, raw_values=raw, mapped_values=mapped, errors=errors, status=status))
        batch.total_rows = total
        batch.valid_rows = valid
        batch.invalid_rows = invalid
        if total == 0:
            db.rollback()
            raise HTTPException(status_code=422, detail="The selected sheet contains no non-empty data rows after the header.")
        db.commit()
        db.refresh(batch)
        preview_rows = db.query(ReliabilityWorkbookImportRowResult).filter(ReliabilityWorkbookImportRowResult.batch_id == batch.id).order_by(ReliabilityWorkbookImportRowResult.row_number).limit(200).all()
        return {**_batch_dict(batch), "preview_rows": [{"id": row.id, "row_number": row.row_number, "status": row.status, "raw_values": row.raw_values, "mapped_values": row.mapped_values, "errors": row.errors, "row_source_hash": row.row_source_hash} for row in preview_rows], "preview_truncated": total > 200}

    @router.get("/workbook-parity/imports")
    def list_imports(
        status: str | None = None,
        profile_code: str | None = None,
        dataset_code: wp.WorkbookDatasetCode | None = None,
        limit: int = Query(default=50, ge=1, le=250),
        offset: int = Query(default=0, ge=0),
        current_user: account_models.User = Depends(get_current_active_user),
        db: Session = Depends(get_write_db),
    ):
        query = db.query(ReliabilityWorkbookImportBatch).filter(ReliabilityWorkbookImportBatch.amo_id == _amo_id(current_user))
        if status:
            query = query.filter(ReliabilityWorkbookImportBatch.status == status.upper())
        if profile_code:
            query = query.filter(ReliabilityWorkbookImportBatch.profile_code == profile_code)
        if dataset_code:
            query = query.filter(ReliabilityWorkbookImportBatch.dataset_code == dataset_code.value)
        total = query.count()
        rows = query.order_by(ReliabilityWorkbookImportBatch.created_at.desc(), ReliabilityWorkbookImportBatch.id.desc()).offset(offset).limit(limit).all()
        return {"total": total, "offset": offset, "limit": limit, "items": [_batch_dict(row) for row in rows]}

    @router.get("/workbook-parity/imports/{batch_id}")
    def read_import(
        batch_id: int,
        row_status: str | None = None,
        limit: int = Query(default=200, ge=1, le=500),
        offset: int = Query(default=0, ge=0),
        current_user: account_models.User = Depends(get_current_active_user),
        db: Session = Depends(get_write_db),
    ):
        batch = db.query(ReliabilityWorkbookImportBatch).filter(ReliabilityWorkbookImportBatch.id == batch_id, ReliabilityWorkbookImportBatch.amo_id == _amo_id(current_user)).one_or_none()
        if not batch:
            raise HTTPException(status_code=404, detail="Workbook import batch not found.")
        query = db.query(ReliabilityWorkbookImportRowResult).filter(ReliabilityWorkbookImportRowResult.batch_id == batch.id)
        if row_status:
            query = query.filter(ReliabilityWorkbookImportRowResult.status == row_status.upper())
        total = query.count()
        rows = query.order_by(ReliabilityWorkbookImportRowResult.row_number).offset(offset).limit(limit).all()
        return {**_batch_dict(batch), "row_total": total, "row_offset": offset, "row_limit": limit, "rows": [{"id": row.id, "row_number": row.row_number, "status": row.status, "raw_values": row.raw_values, "mapped_values": row.mapped_values, "errors": row.errors, "attempt_count": row.attempt_count, "last_error": row.last_error, "workbook_record_id": row.workbook_record_id, "row_source_hash": row.row_source_hash} for row in rows]}

    @router.post("/workbook-parity/imports/{batch_id}/commit")
    def commit_import(
        batch_id: int,
        request: ImportCommitRequest,
        current_user: account_models.User = Depends(get_current_active_user),
        db: Session = Depends(get_write_db),
    ):
        amo_id = _amo_id(current_user)
        _require_import_permission(current_user)
        batch = db.query(ReliabilityWorkbookImportBatch).filter(ReliabilityWorkbookImportBatch.id == batch_id, ReliabilityWorkbookImportBatch.amo_id == amo_id).with_for_update().one_or_none()
        if not batch:
            raise HTTPException(status_code=404, detail="Workbook import batch not found.")
        if batch.status == "COMPLETED":
            return {**_batch_dict(batch), "processed": 0, "message": "Import batch is already complete."}
        rows = db.query(ReliabilityWorkbookImportRowResult).filter(ReliabilityWorkbookImportRowResult.batch_id == batch.id, ReliabilityWorkbookImportRowResult.status == "VALID").order_by(ReliabilityWorkbookImportRowResult.row_number).limit(request.chunk_size).with_for_update(skip_locked=True).all()
        processed = 0
        for row in rows:
            row.attempt_count += 1
            try:
                with db.begin_nested():
                    existing_record = db.query(wp.ReliabilityWorkbookRecord).filter(wp.ReliabilityWorkbookRecord.amo_id == amo_id, wp.ReliabilityWorkbookRecord.source_hash == row.row_source_hash).one_or_none()
                    if existing_record:
                        row.status = "COMMITTED"
                        row.workbook_record_id = existing_record.id
                        row.last_error = None
                        processed += 1
                        continue
                    create = wp.WorkbookRecordCreate.model_validate(row.mapped_values)
                    wp._validate_aircraft(db, amo_id, create.aircraft_serial_number)
                    definition = wp.DATASET_CATALOG[create.dataset_code]
                    normalised, derived = wp._normalise_payload(definition, create.payload)
                    record = wp.ReliabilityWorkbookRecord(
                        amo_id=amo_id,
                        dataset_code=create.dataset_code.value,
                        record_number=wp._record_number(db, amo_id, create.dataset_code.value, create.event_date),
                        revision=1,
                        status=wp.WorkbookRecordStatus.DRAFT.value,
                        event_date=create.event_date,
                        event_end_date=create.event_end_date,
                        aircraft_serial_number=create.aircraft_serial_number,
                        ata_chapter=create.ata_chapter,
                        reference_code=create.reference_code,
                        title=create.title,
                        description=create.description,
                        payload=normalised,
                        derived_values=derived,
                        source_workbook=batch.sanitized_filename,
                        source_sheet=batch.selected_sheet,
                        source_row_number=row.row_number,
                        source_hash=row.row_source_hash,
                        created_by_user_id=current_user.id,
                    )
                    db.add(record)
                    db.flush()
                    row.status = "COMMITTED"
                    row.workbook_record_id = record.id
                    row.last_error = None
                    processed += 1
            except Exception as exc:
                row.status = "FAILED"
                row.last_error = str(getattr(exc, "detail", exc))[:2000]
        db.flush()
        _refresh_counts(db, batch)
        db.commit()
        db.refresh(batch)
        return {**_batch_dict(batch), "processed": processed, "remaining_valid_rows": batch.valid_rows}

    @router.post("/workbook-parity/imports/{batch_id}/retry")
    def retry_import(
        batch_id: int,
        request: ImportRetryRequest,
        current_user: account_models.User = Depends(get_current_active_user),
        db: Session = Depends(get_write_db),
    ):
        amo_id = _amo_id(current_user)
        _require_import_permission(current_user)
        batch = db.query(ReliabilityWorkbookImportBatch).filter(ReliabilityWorkbookImportBatch.id == batch_id, ReliabilityWorkbookImportBatch.amo_id == amo_id).with_for_update().one_or_none()
        if not batch:
            raise HTTPException(status_code=404, detail="Workbook import batch not found.")
        query = db.query(ReliabilityWorkbookImportRowResult).filter(ReliabilityWorkbookImportRowResult.batch_id == batch.id)
        query = query.filter(ReliabilityWorkbookImportRowResult.status == "FAILED") if request.failed_only else query.filter(ReliabilityWorkbookImportRowResult.status.in_(["FAILED", "VALID"]))
        reset = query.update({"status": "VALID", "last_error": None}, synchronize_session=False)
        db.flush()
        _refresh_counts(db, batch)
        db.commit()
        db.refresh(batch)
        return {**_batch_dict(batch), "reset_rows": reset}
