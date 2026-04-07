from __future__ import annotations

import io
import os
from dataclasses import dataclass
from datetime import datetime, date, timezone
from typing import Any, Optional

from sqlalchemy import func
from sqlalchemy.orm import Session

from . import models, schemas
from .services import get_password_hash

STATUS_ACTIVE = "Active"
STATUS_DORMANT = "Dormant"


@dataclass
class ParsedRow:
    row_number: int
    person_id: str
    first_name: str
    last_name: str
    full_name: Optional[str]
    national_id: Optional[str]
    amel_no: Optional[str]
    internal_certification_stamp_no: Optional[str]
    initial_authorization_date: Optional[date]
    department: Optional[str]
    position_title: Optional[str]
    phone_number: Optional[str]
    secondary_phone: Optional[str]
    email: Optional[str]
    hire_date: Optional[date]
    employment_status: Optional[str]
    status: str
    date_of_birth: Optional[date]
    birth_place: Optional[str]


def _to_clean_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date value '{text}'")


def _normalize_status(value: Optional[str]) -> str:
    raw = (value or "").strip()
    if not raw:
        raise ValueError("Status is required")
    lowered = raw.lower()
    if lowered == "active":
        return STATUS_ACTIVE
    if lowered == "dormant":
        return STATUS_DORMANT
    return raw


def parse_people_sheet(file_bytes: bytes, *, filename: str, sheet_name: str = "People") -> list[dict[str, Any]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in {"xlsx", "xlsm", "xltx", "xltm"}:
        raise ValueError("Only Excel .xlsx/.xlsm files are supported for personnel import.")

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for personnel import. Install openpyxl.") from exc

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    payload_rows: list[dict[str, Any]] = []
    for idx, values in enumerate(rows[1:], start=2):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        person_id = _to_clean_str(row.get("PersonID"))
        if person_id and person_id.lower() == "total":
            continue
        payload_rows.append({"row_number": idx, **row})

    return payload_rows


def _build_parsed_row(raw: dict[str, Any]) -> ParsedRow:
    person_id = _to_clean_str(raw.get("PersonID"))
    first_name = _to_clean_str(raw.get("FIRSTNAME"))
    last_name = _to_clean_str(raw.get("LASTNAME"))
    status = _normalize_status(_to_clean_str(raw.get("Status")))

    if not person_id:
        raise ValueError("PersonID is required")
    if not first_name:
        raise ValueError("FIRSTNAME is required")
    if not last_name:
        raise ValueError("LASTNAME is required")

    email = _to_clean_str(raw.get("Email"))
    if email:
        email = email.lower()

    return ParsedRow(
        row_number=int(raw["row_number"]),
        person_id=person_id,
        first_name=first_name,
        last_name=last_name,
        full_name=_to_clean_str(raw.get("PersonName")),
        national_id=_to_clean_str(raw.get("nid")),
        amel_no=_to_clean_str(raw.get("AMEL NO:")),
        internal_certification_stamp_no=_to_clean_str(raw.get("Internal Certification Stamp No:")),
        initial_authorization_date=_parse_date(raw.get("initial_auth")),
        department=_to_clean_str(raw.get("Department")),
        position_title=_to_clean_str(raw.get("Position")),
        phone_number=_to_clean_str(raw.get("PhoneNumber")),
        secondary_phone=_to_clean_str(raw.get("secondary_phone")),
        email=email,
        hire_date=_parse_date(raw.get("HireDate")),
        employment_status=_to_clean_str(raw.get("Employment_Status")),
        status=status,
        date_of_birth=_parse_date(raw.get("DOB")),
        birth_place=_to_clean_str(raw.get("birthplace")),
    )


def import_personnel_rows(
    db: Session,
    *,
    amo_id: str,
    rows: list[dict[str, Any]],
    dry_run: bool,
) -> schemas.PersonnelImportSummary:
    issues: list[schemas.PersonnelImportRowIssue] = []
    created_personnel = 0
    updated_personnel = 0
    created_accounts = 0
    updated_accounts = 0
    skipped_accounts = 0
    skipped_rows = 0

    default_password = (os.getenv("PERSONNEL_IMPORT_DEFAULT_PASSWORD") or "").strip()
    if not default_password and not dry_run:
        raise ValueError("PERSONNEL_IMPORT_DEFAULT_PASSWORD is required for live import.")

    seen_person_ids: set[str] = set()
    seen_emails: set[str] = set()

    for raw in rows:
        try:
            parsed = _build_parsed_row(raw)
        except ValueError as exc:
            issues.append(schemas.PersonnelImportRowIssue(row_number=int(raw.get("row_number") or 0), reason=str(exc)))
            continue

        if parsed.person_id.lower() in seen_person_ids:
            issues.append(schemas.PersonnelImportRowIssue(row_number=parsed.row_number, person_id=parsed.person_id, reason="Duplicate PersonID inside import file."))
            continue
        seen_person_ids.add(parsed.person_id.lower())

        if parsed.email:
            email_key = parsed.email.lower()
            if email_key in seen_emails:
                issues.append(schemas.PersonnelImportRowIssue(row_number=parsed.row_number, person_id=parsed.person_id, reason="Duplicate Email inside import file."))
                continue
            seen_emails.add(email_key)

        profile = (
            db.query(models.PersonnelProfile)
            .filter(models.PersonnelProfile.amo_id == amo_id, models.PersonnelProfile.person_id == parsed.person_id)
            .first()
        )

        now = datetime.now(timezone.utc)
        is_new_profile = profile is None
        if is_new_profile:
            profile = models.PersonnelProfile(amo_id=amo_id, person_id=parsed.person_id, created_at=now, updated_at=now)

        profile.first_name = parsed.first_name
        profile.last_name = parsed.last_name
        profile.full_name = parsed.full_name or f"{parsed.first_name} {parsed.last_name}".strip()
        profile.national_id = parsed.national_id
        profile.amel_no = parsed.amel_no
        profile.internal_certification_stamp_no = parsed.internal_certification_stamp_no
        profile.initial_authorization_date = parsed.initial_authorization_date
        profile.department = parsed.department
        profile.position_title = parsed.position_title
        profile.phone_number = parsed.phone_number
        profile.secondary_phone = parsed.secondary_phone
        profile.email = parsed.email
        profile.hire_date = parsed.hire_date
        profile.employment_status = parsed.employment_status
        profile.status = parsed.status
        profile.date_of_birth = parsed.date_of_birth
        profile.birth_place = parsed.birth_place
        profile.updated_at = now

        if is_new_profile:
            db.add(profile)
            created_personnel += 1
        else:
            updated_personnel += 1

        if not parsed.email:
            skipped_accounts += 1
            issues.append(schemas.PersonnelImportRowIssue(row_number=parsed.row_number, person_id=parsed.person_id, reason="Account creation skipped: missing email."))
            continue

        existing_user = None
        if profile.user_id:
            existing_user = (
                db.query(models.User)
                .filter(models.User.id == profile.user_id)
                .first()
            )

        if not existing_user:
            existing_user = (
                db.query(models.User)
                .filter(models.User.amo_id == amo_id, func.lower(models.User.email) == parsed.email.lower())
                .first()
            )

        if not existing_user:
            if dry_run:
                created_accounts += 1
            else:
                existing_user = models.User(
                    amo_id=amo_id,
                    department_id=None,
                    staff_code=parsed.person_id,
                    email=parsed.email,
                    first_name=parsed.first_name,
                    last_name=parsed.last_name,
                    full_name=parsed.full_name or f"{parsed.first_name} {parsed.last_name}".strip(),
                    role=models.AccountRole.TECHNICIAN,
                    position_title=parsed.position_title,
                    phone=parsed.phone_number,
                    secondary_phone=parsed.secondary_phone,
                    hashed_password=get_password_hash(default_password),
                    is_active=parsed.status == STATUS_ACTIVE,
                    is_amo_admin=False,
                    is_auditor=False,
                    must_change_password=True,
                    password_changed_at=None,
                )
                db.add(existing_user)
                db.flush()
                created_accounts += 1
        else:
            updated_accounts += 1
            if not dry_run:
                existing_user.first_name = parsed.first_name
                existing_user.last_name = parsed.last_name
                existing_user.full_name = parsed.full_name or f"{parsed.first_name} {parsed.last_name}".strip()
                existing_user.position_title = parsed.position_title
                existing_user.phone = parsed.phone_number
                existing_user.secondary_phone = parsed.secondary_phone
                existing_user.is_active = parsed.status == STATUS_ACTIVE
                if existing_user.staff_code != parsed.person_id:
                    issues.append(schemas.PersonnelImportRowIssue(row_number=parsed.row_number, person_id=parsed.person_id, reason=f"User exists with email but different staff_code ({existing_user.staff_code}); staff_code not changed."))

        if not dry_run and existing_user:
            profile.user_id = existing_user.id

    if dry_run:
        db.rollback()
    else:
        db.commit()

    rejected_rows = len([i for i in issues if "skipped" not in i.reason.lower() and "different staff_code" not in i.reason.lower()])
    skipped_rows = len([i for i in issues if "skipped" in i.reason.lower()]) + skipped_rows

    return schemas.PersonnelImportSummary(
        dry_run=dry_run,
        rows_processed=len(rows),
        created_personnel=created_personnel,
        updated_personnel=updated_personnel,
        created_accounts=created_accounts,
        updated_accounts=updated_accounts,
        skipped_accounts=skipped_accounts,
        rejected_rows=rejected_rows,
        skipped_rows=skipped_rows,
        issues=issues,
    )
